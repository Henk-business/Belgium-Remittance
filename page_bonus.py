"""
Bonus & Payout Tools
  Tab 1 — Customer matching: compare SAP export vs bonus file, highlight differences,
           add missing SAP accounts into the bonus file.
  Tab 2 — Payout & block checker: scan export for X payouts (clean, no B/U blocks),
           flag open invoices on or before a chosen date.
"""
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime


# ── COLOURS ───────────────────────────────────────────────────────────────────
DK_BLUE  = "FF1F3864"
MD_BLUE  = "FF2E75B6"
WHITE    = "FFFFFFFF"
GREY     = "FFF2F2F2"
BLACK    = "FF000000"
GREEN_HL = "FFE2EFDA"   # light green  – in both files
ORANGE_HL= "FFFCE4D6"   # light orange – in bonus file but not SAP
YELLOW_HL= "FFFFF2CC"   # yellow       – added from SAP (missing in bonus)
RED_HL   = "FFFFC7CE"   # light red    – flagged / problem


def _fill(rgb): return PatternFill("solid", fgColor=rgb)
def _font(bold=False, color=BLACK, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _align(ha="left"):
    return Alignment(horizontal=ha, vertical="center", wrap_text=False)
def _thin():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)
def _mw(ws, row, c1, c2, val, bold=False, bg=WHITE, fg=BLACK, sz=10, ha="left"):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    cell = ws.cell(row=row, column=c1, value=val)
    cell.font = _font(bold=bold, color=fg, size=sz)
    cell.fill = _fill(bg); cell.alignment = _align(ha); cell.border = _thin()
    for c in range(c1+1, c2+1):
        ws.cell(row, c).fill = _fill(bg); ws.cell(row, c).border = _thin()


def _read_account_col(df) -> str | None:
    """Find the account/customer number column."""
    for col in df.columns:
        low = col.lower()
        if any(kw in low for kw in ("account","konto","debitor","customer","klant","client","nr","number")):
            # Confirm it looks like numbers
            sample = df[col].dropna().astype(str).str.strip()
            if sample.str.match(r"^\d{5,12}$").mean() > 0.3:
                return col
    # Fallback: first column with numeric-looking values
    for col in df.columns:
        sample = df[col].dropna().astype(str).str.strip()
        if sample.str.match(r"^\d{5,12}$").mean() > 0.5:
            return col
    return None


def _clean_acc(val) -> str:
    """Normalise an account number to plain digits."""
    s = str(val).strip().split(".")[0]
    return s.lstrip("0").zfill(8) if s.isdigit() else s


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — CUSTOMER MATCHING
# ══════════════════════════════════════════════════════════════════════════════

def _build_match_report(sap_df, bonus_df, sap_acc_col, bonus_acc_col,
                         today_str) -> bytes:
    """
    Produce an Excel with:
    Sheet 1 — Annotated bonus file (all original columns kept)
              + Status column: Match / Not in SAP / Added from SAP
    Sheet 2 — Summary
    """
    # Normalise account sets
    sap_accs = {_clean_acc(v) for v in sap_df[sap_acc_col].dropna()}

    bonus_df = bonus_df.copy()
    bonus_df["__acc_norm__"] = bonus_df[bonus_acc_col].apply(_clean_acc)

    # Find SAP accounts missing from bonus file
    bonus_accs = set(bonus_df["__acc_norm__"])
    missing_in_bonus = sap_accs - bonus_accs

    # Build rows to add from SAP
    sap_df = sap_df.copy()
    sap_df["__acc_norm__"] = sap_df[sap_acc_col].apply(_clean_acc)
    sap_extra = sap_df[sap_df["__acc_norm__"].isin(missing_in_bonus)].copy()

    # Status for bonus rows
    bonus_df["Status"] = bonus_df["__acc_norm__"].apply(
        lambda a: "✓ Match" if a in sap_accs else "⚠ Not in SAP export"
    )

    # Build combined df: original bonus rows + added SAP rows
    # For SAP rows, map sap_acc_col → bonus_acc_col
    add_rows = []
    for _, row in sap_extra.iterrows():
        new_row = {c: "" for c in bonus_df.columns}
        new_row[bonus_acc_col] = row[sap_acc_col]
        new_row["__acc_norm__"] = row["__acc_norm__"]
        new_row["Status"] = "➕ Added from SAP"
        add_rows.append(new_row)

    if add_rows:
        add_df = pd.DataFrame(add_rows, columns=bonus_df.columns)
        combined = pd.concat([bonus_df, add_df], ignore_index=True)
    else:
        combined = bonus_df.copy()

    # Drop internal col
    combined = combined.drop(columns=["__acc_norm__"], errors="ignore")

    # ── Build Excel ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bonus Matching"

    display_cols = list(combined.columns)
    ncols = len(display_cols)
    status_ci = display_cols.index("Status") + 1 if "Status" in display_cols else None

    # Title
    _mw(ws, 1, 1, ncols,
        f"Bonus Customer Match  —  {today_str}  ·  "
        f"{len(bonus_df)} bonus accounts  ·  {len(sap_accs)} SAP accounts",
        bold=True, bg=DK_BLUE, fg=WHITE, sz=13)
    ws.row_dimensions[1].height = 32

    # Legend row
    _mw(ws, 2, 1, ncols,
        "🟢 Match   🟠 Not in SAP   🟡 Added from SAP",
        bold=False, bg=MD_BLUE, fg=WHITE, sz=9)
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # Headers
    for ci, h in enumerate(display_cols, 1):
        cell = ws.cell(4, ci, value=h)
        cell.font = _font(bold=True, color=WHITE, size=9)
        cell.fill = _fill(MD_BLUE)
        cell.alignment = _align("center")
        cell.border = _thin()
        ws.column_dimensions[get_column_letter(ci)].width = max(len(h)+2, 14)
    ws.row_dimensions[4].height = 16

    # Data rows
    for ri, (_, row) in enumerate(combined.iterrows()):
        r = 5 + ri
        status = str(row.get("Status", ""))
        if "Added" in status:
            row_fill = YELLOW_HL
        elif "Not in SAP" in status:
            row_fill = ORANGE_HL
        else:
            row_fill = GREEN_HL if ri % 2 == 0 else WHITE

        for ci, col in enumerate(display_cols, 1):
            val = row[col]
            if pd.isna(val): val = ""
            elif isinstance(val, float) and val == int(val): val = int(val)
            cell = ws.cell(r, ci, value=val)
            cell.font = _font(size=9)
            cell.fill = _fill(row_fill)
            cell.alignment = _align("left")
            cell.border = _thin()
        ws.row_dimensions[r].height = 13

    ws.freeze_panes = "A5"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    summary_rows = [
        ("Total SAP accounts",          len(sap_accs)),
        ("Total bonus file accounts",   len(bonus_accs)),
        ("Matched (in both)",           len(sap_accs & bonus_accs)),
        ("In bonus file, NOT in SAP",   len(bonus_accs - sap_accs)),
        ("In SAP, NOT in bonus file",   len(missing_in_bonus)),
        ("Rows added to bonus file",    len(add_rows)),
    ]
    _mw(ws2, 1, 1, 2, f"Bonus Match Summary  —  {today_str}",
        bold=True, bg=DK_BLUE, fg=WHITE, sz=13)
    ws2.row_dimensions[1].height = 30
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 14
    for ri, (label, val) in enumerate(summary_rows, 3):
        ws2.cell(ri, 1, value=label).font = _font(size=10)
        ws2.cell(ri, 2, value=val).font   = _font(bold=True, size=10)
        ws2.cell(ri, 1).fill = _fill(GREY if ri % 2 == 0 else WHITE)
        ws2.cell(ri, 2).fill = _fill(GREY if ri % 2 == 0 else WHITE)
        ws2.row_dimensions[ri].height = 16

    # Missing from bonus — list them
    if missing_in_bonus:
        ws2.cell(len(summary_rows)+4, 1,
                 value="SAP accounts missing from bonus file:").font = _font(bold=True, size=10)
        for i, acc in enumerate(sorted(missing_in_bonus), len(summary_rows)+5):
            ws2.cell(i, 1, value=acc).font = _font(size=9)
            ws2.cell(i, 1).fill = _fill(YELLOW_HL)

    out = BytesIO()
    wb.save(out); out.seek(0)
    return out.read()


def _show_matching():
    st.markdown("### 📋 Customer matching")
    st.caption(
        "Compare your SAP customer list with a bonus file. "
        "Highlights which accounts match, which are missing, and adds SAP-only accounts to the output."
    )

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Your SAP export** (FBL5N or customer list)")
        sap_file = st.file_uploader("SAP", type=["xlsx","xls","xlsm"],
                                    label_visibility="collapsed", key="bon_sap")
    with c2:
        st.markdown("**Bonus / partner file**")
        bonus_file = st.file_uploader("Bonus", type=["xlsx","xls","xlsm"],
                                      label_visibility="collapsed", key="bon_bonus")

    if not sap_file or not bonus_file:
        return

    try:
        sap_df   = pd.read_excel(sap_file,   dtype=str, engine="openpyxl")
        bonus_df = pd.read_excel(bonus_file, dtype=str, engine="openpyxl")
    except Exception as e:
        st.error(f"Could not read files: {e}")
        return

    sap_acc_col   = _read_account_col(sap_df)
    bonus_acc_col = _read_account_col(bonus_df)

    sc1, sc2 = st.columns(2)
    with sc1:
        sap_acc_col = st.selectbox(
            "Account column in SAP file",
            sap_df.columns.tolist(),
            index=sap_df.columns.tolist().index(sap_acc_col) if sap_acc_col else 0,
            key="bon_sap_col"
        )
    with sc2:
        bonus_acc_col = st.selectbox(
            "Account column in bonus file",
            bonus_df.columns.tolist(),
            index=bonus_df.columns.tolist().index(bonus_acc_col) if bonus_acc_col else 0,
            key="bon_bonus_col"
        )

    if st.button("🔍  Run matching", type="primary", key="bon_run",
                 use_container_width=True):
        with st.spinner("Matching accounts…"):
            try:
                today_str = datetime.date.today().strftime("%d/%m/%Y")
                result = _build_match_report(
                    sap_df, bonus_df, sap_acc_col, bonus_acc_col, today_str
                )
                st.session_state["bon_result"] = result
            except Exception as e:
                st.error(f"Error: {e}")
                import traceback
                with st.expander("Detail"): st.code(traceback.format_exc())

    if "bon_result" not in st.session_state:
        return

    result = st.session_state["bon_result"]

    # Quick stats
    wb   = openpyxl.load_workbook(BytesIO(result))
    ws2  = wb["Summary"]
    stats = {}
    for r in range(3, 9):
        label = ws2.cell(r,1).value
        val   = ws2.cell(r,2).value
        if label: stats[label] = val

    st.markdown("---")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("SAP accounts",    stats.get("Total SAP accounts","—"))
    m2.metric("Bonus accounts",  stats.get("Total bonus file accounts","—"))
    m3.metric("✓ Matched",       stats.get("Matched (in both)","—"))
    m4.metric("⚠ Differences",
              (stats.get("In bonus file, NOT in SAP",0) or 0) +
              (stats.get("In SAP, NOT in bonus file",0) or 0))

    today = datetime.date.today().strftime("%Y%m%d")
    st.download_button(
        "⬇  Download match report",
        data=result,
        file_name=f"BonusMatch_{today}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, key="bon_dl",
    )

    added = stats.get("Rows added to bonus file", 0) or 0
    missing_sap = stats.get("In bonus file, NOT in SAP", 0) or 0
    if added:
        st.info(f"➕ {added} SAP account(s) were missing from the bonus file and have been added (highlighted yellow).")
    if missing_sap:
        st.warning(f"⚠ {missing_sap} account(s) in the bonus file were not found in your SAP export (highlighted orange). Check if these are valid customers.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — PAYOUT & BLOCK CHECKER
# ══════════════════════════════════════════════════════════════════════════════

def _is_bonus_file(df) -> bool:
    """Detect if this is the dedicated bonus/payout file (has Bonus, Payout Y/N cols)."""
    cols_lower = [c.lower() for c in df.columns]
    return any("bonus" in c for c in cols_lower) and any("payout" in c for c in cols_lower)


def _build_payout_report(df, cutoff_date, today_str) -> tuple:
    """
    Analyse a SAP FBL5N export:
    - X payouts: flag B or U payment blocks
    - B-blocked items: any row with Payment Block B
    - Open invoices on or before cutoff date
    """
    is_bonus = False  # Tab 2 always treats file as SAP export
    col_map   = {c.lower().strip(): c for c in df.columns}

    # Column detection
    pay_meth_col  = next((col_map[k] for k in col_map if "payment method" in k), None)
    pay_block_col = next((col_map[k] for k in col_map if "payment block" in k), None)
    acc_col       = next((col_map[k] for k in col_map
                          if k in ("account","customer","konto","debitor")), None)
    name_col      = next((col_map[k] for k in col_map if k == "name"), None)
    amt_col       = next((col_map[k] for k in col_map if "amount" in k), None)
    if not amt_col:
        amt_col = next((col_map[k] for k in col_map if "amount" in k or "balance" in k), None)
    due_col       = next((col_map[k] for k in col_map
                          if "next payout" in k or "net due" in k or "due date" in k), None)
    status_col    = next((col_map[k] for k in col_map if k == "status"), None)
    payout_col    = next((col_map[k] for k in col_map if "payout y" in k), None)
    open_col      = next((col_map[k] for k in col_map
                          if "other open" in k or "open item" in k), None)
    doc_type_col  = next((col_map[k] for k in col_map
                          if "document type" in k or "belegtyp" in k), None)

    df = df.copy()
    if due_col:
        df[due_col] = pd.to_datetime(df[due_col], errors="coerce")
    if amt_col:
        df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce")
    if open_col:
        df[open_col] = pd.to_numeric(df[open_col], errors="coerce")

    cutoff_ts = pd.Timestamp(cutoff_date)

    def _get(row, col):
        if not col or col not in row.index: return ""
        v = row[col]
        return str(v).strip().upper() if pd.notna(v) else ""

    x_clean    = []
    x_blocked  = []
    b_blocked  = []
    open_items = []

    for _, row in df.iterrows():
        pm   = _get(row, pay_meth_col)
        pb   = _get(row, pay_block_col)
        acc  = _get(row, acc_col)
        name = str(row[name_col]).strip() if name_col and pd.notna(row.get(name_col)) else ""
        amt  = row[amt_col] if amt_col and pd.notna(row.get(amt_col)) else None
        due  = row[due_col] if due_col and pd.notna(row.get(due_col)) else None
        stat = _get(row, status_col)
        pout = _get(row, payout_col)
        oi   = row[open_col] if open_col and pd.notna(row.get(open_col)) else 0

        base = {"Account": acc, "Name": name, "Amount": amt, "Due": due,
                "Status": stat, "Payout Y/N": pout}

        # X payout rows
        if pm == "X":
            if pb in ("B", "U"):
                x_blocked.append({**base,
                    "Payment Block": pb,
                    "Issue": f"X payout blocked — {pb} block must be removed before payout"})
            else:
                x_clean.append({**base,
                    "Payment Block": pb or "—",
                    "Issue": "OK — no block"})

        # B-blocked items
        if pb == "B":
            b_blocked.append({**base,
                "Payment Method": pm,
                "Issue": "Payment Block B"})

        # Open items by cutoff (non-X rows)
        if (due is not None and due <= cutoff_ts and
                amt is not None and float(amt) > 0 and pm != "X"):
            open_items.append({**base,
                "Issue": f"Open invoice due on or before {cutoff_date.strftime('%d/%m/%Y')}"})

    # ── X-account blockers ────────────────────────────────────────────────────
    # For every account that has an X row, check if there are open RV invoices,
    # positive RS (re-invoices/debit corrections) or RU rows on the same account
    # — these would likely prevent a clean payout
    x_accounts = {str(r.get("Account","")) for r in x_clean + x_blocked}
    x_account_blockers = []
    if acc_col and doc_type_col:
        for _, row in df.iterrows():
            acc  = str(row[acc_col]).strip().split(".")[0].lstrip("0").zfill(8) if acc_col else ""
            if acc not in x_accounts:
                continue
            dt   = str(row.get(doc_type_col,"") or "").strip().upper() if doc_type_col else ""
            pm2  = str(row.get(pay_meth_col,"") or "").strip().upper() if pay_meth_col else ""
            amt2 = float(row[amt_col]) if amt_col and pd.notna(row.get(amt_col)) else 0
            due2 = row[due_col] if due_col and pd.notna(row.get(due_col)) else None
            name = str(row[name_col]).strip() if name_col and pd.notna(row.get(name_col)) else ""

            if pm2 == "X":
                continue  # skip the X row itself

            blocker = None
            if dt == "RV" and amt2 > 0:
                blocker = f"Open invoice (RV) €{amt2:,.2f}"
            elif dt.startswith("RS") and amt2 > 0:
                blocker = f"Positive RS (re-invoice/debit) €{amt2:,.2f}"
            elif dt.startswith("RU") and amt2 > 0:
                blocker = f"Positive RU €{amt2:,.2f}"

            if blocker:
                due_str = pd.Timestamp(due2).strftime("%d/%m/%Y") if due2 is not None else "—"
                x_account_blockers.append({
                    "Account": acc, "Name": name,
                    "Document Type": dt, "Amount": amt2, "Due": due_str,
                    "Issue": f"X account has blocker: {blocker}"
                })

    summary = {
        "X payouts — clean (no block)":       len(x_clean),
        "X payouts — BLOCKED (B or U)":       len(x_blocked),
        "B-blocked items":                     len(b_blocked),
        "Open items on/before cutoff":         len(open_items),
        "X accounts with open blockers":       len({r["Account"] for r in x_account_blockers}),
    }

    # ── Build Excel ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _write_sheet(title, rows, cols, row_fill_fn):
        ws = wb.create_sheet(title=title[:31])
        ncols = len(cols)
        _mw(ws, 1, 1, ncols,
            f"{title}  —  {today_str}  ·  {len(rows)} item(s)",
            bold=True, bg=DK_BLUE, fg=WHITE, sz=12)
        ws.row_dimensions[1].height = 28
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(2, ci, value=h)
            cell.font = _font(bold=True, color=WHITE, size=9)
            cell.fill = _fill(MD_BLUE); cell.alignment = _align("center"); cell.border = _thin()
            ws.column_dimensions[get_column_letter(ci)].width = max(len(h)+2, 14)
        ws.row_dimensions[2].height = 15
        if not rows:
            _mw(ws, 3, 1, ncols, "No items found — all clear ✓", bg=GREEN_HL, sz=10)
            ws.row_dimensions[3].height = 20
            return
        for ri, row_data in enumerate(rows):
            r = 3 + ri
            rf = row_fill_fn(ri)
            for ci, col in enumerate(cols, 1):
                val = row_data.get(col, "")
                if isinstance(val, pd.Timestamp): val = val.strftime("%d/%m/%Y")
                elif not isinstance(val, str) and pd.isna(val): val = ""
                elif isinstance(val, float) and val == int(val): val = int(val)
                is_amt = col in ("Amount","Open Amount","Bonus")
                cell = ws.cell(r, ci, value=val)
                cell.font = _font(size=9,
                    color="FFC00000" if is_amt and isinstance(val,(int,float)) and val > 0
                    else "FF375623" if is_amt and isinstance(val,(int,float)) and val < 0
                    else BLACK)
                cell.fill = _fill(rf)
                cell.alignment = _align("right" if is_amt else "left")
                cell.border = _thin()
                if is_amt and isinstance(val,(int,float)):
                    cell.number_format = "#,##0.00"
            ws.row_dimensions[r].height = 13
        ws.freeze_panes = "A3"

    _write_sheet("X Payouts — OK", x_clean,
        ["Account","Name","Amount","Due","Payment Block","Status","Payout Y/N","Issue"],
        lambda i: GREEN_HL)
    _write_sheet("X Payouts — BLOCKED", x_blocked,
        ["Account","Name","Amount","Due","Payment Block","Status","Payout Y/N","Issue"],
        lambda i: RED_HL)
    _write_sheet("B-Blocked Items", b_blocked,
        ["Account","Name","Amount","Due","Payment Method","Status","Issue"],
        lambda i: ORANGE_HL)
    _write_sheet("Open Items by Cutoff", open_items,
        ["Account","Name","Amount","Due","Status","Payout Y/N","Issue"],
        lambda i: YELLOW_HL if i % 2 == 0 else WHITE)
    _write_sheet("X Account Blockers", x_account_blockers,
        ["Account","Name","Document Type","Amount","Due","Issue"],
        lambda i: RED_HL)

    # Summary sheet
    ws_s = wb.create_sheet("Summary", 0)
    _mw(ws_s, 1, 1, 2, f"Payout & Block Check  —  {today_str}  ·  Cutoff: {cutoff_date.strftime('%d/%m/%Y')}",
        bold=True, bg=DK_BLUE, fg=WHITE, sz=13)
    ws_s.row_dimensions[1].height = 30
    ws_s.column_dimensions["A"].width = 35
    ws_s.column_dimensions["B"].width = 14
    for ri, (label, val) in enumerate(summary.items(), 3):
        has_issue = val > 0 and ("BLOCKED" in label or "B-blocked" in label or "Open" in label)
        rf = RED_HL if has_issue else (GREEN_HL if val == 0 else GREY)
        ws_s.cell(ri, 1, value=label).font = _font(size=10)
        ws_s.cell(ri, 2, value=val).font   = _font(bold=True, size=10)
        ws_s.cell(ri, 1).fill = _fill(rf)
        ws_s.cell(ri, 2).fill = _fill(rf)
        ws_s.row_dimensions[ri].height = 18

    out = BytesIO(); wb.save(out); out.seek(0)
    return out.read(), summary


def _show_payout_checker():
    st.markdown("### 💸 Payout & block checker")
    st.caption(
        "Upload your SAP export or bonus/payout file. "
        "Checks X payouts for blocks, flags B-blocked items, and shows open items by cutoff date."
    )

    sap_file = st.file_uploader(
        "SAP export or bonus/payout file", type=["xlsx","xls","xlsm"],
        label_visibility="collapsed", key="pbc_sap"
    )

    p1, p2, _ = st.columns([1, 1, 2])
    with p1:
        cutoff = st.date_input(
            "Cutoff date for open items",
            value=datetime.date.today().replace(day=21),
            key="pbc_cutoff",
            help="Flag open items with due date on or before this date"
        )
    with p2:
        st.markdown("")
        st.markdown("")
        run = st.button("▶  Run check", type="primary",
                        key="pbc_run", use_container_width=True)

    if run:
        if not sap_file:
            st.error("Please upload a file.")
        else:
            with st.spinner("Analysing…"):
                try:
                    df = pd.read_excel(sap_file, dtype=str, engine="openpyxl")
                    today_str = datetime.date.today().strftime("%d/%m/%Y")
                    result, summary = _build_payout_report(df, cutoff, today_str)
                    st.session_state["pbc_result"]  = result
                    st.session_state["pbc_summary"] = summary
                    st.session_state["pbc_cutoff"]  = cutoff
                except Exception as e:
                    st.error(f"Error: {e}")
                    import traceback
                    with st.expander("Detail"): st.code(traceback.format_exc())

    if "pbc_result" not in st.session_state:
        return

    result  = st.session_state["pbc_result"]
    summary = st.session_state["pbc_summary"]
    cutoff  = st.session_state.get("pbc_cutoff", cutoff)
    st.markdown("---")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("✅ X payouts clean",   summary.get("X payouts — clean (no block)", 0))
    m2.metric("🚫 X payouts blocked", summary.get("X payouts — BLOCKED (B or U)", 0),
              delta="needs action" if summary.get("X payouts — BLOCKED (B or U)",0) > 0 else None,
              delta_color="inverse")
    m3.metric("🔴 B-blocked items",   summary.get("B-blocked items", 0),
              delta="needs action" if summary.get("B-blocked items",0) > 0 else None,
              delta_color="inverse")
    m4.metric(f"📅 Open by {cutoff.strftime('%d/%m')}",
              summary.get("Open items on/before cutoff", 0),
              delta="needs review" if summary.get("Open items on/before cutoff",0) > 0 else None,
              delta_color="inverse")

    _, m5 = st.columns([3,1])
    m5.metric("⛔ X acct blockers",
              summary.get("X accounts with open blockers", 0),
              delta="check before payout" if summary.get("X accounts with open blockers",0) > 0 else None,
              delta_color="inverse")

    if summary.get("X payouts — BLOCKED (B or U)", 0) > 0:
        st.error(f"🚫 {summary['X payouts — BLOCKED (B or U)']} X payout(s) have a B or U block — remove the block before the payout run.")
    if summary.get("B-blocked items", 0) > 0:
        st.warning(f"⚠ {summary['B-blocked items']} item(s) have a Payment Block B.")
    if summary.get("Open items on/before cutoff", 0) > 0:
        st.warning(f"📅 {summary['Open items on/before cutoff']} open item(s) due on or before {cutoff.strftime('%d/%m/%Y')}.")
    if summary.get("X accounts with open blockers", 0) > 0:
        n_acc = summary['X accounts with open blockers']
        st.error(f"⛔ {n_acc} account(s) with an X payout have open invoices, positive RS or RU rows — these may need to be resolved before payout. See 'X Account Blockers' sheet.")
    if all(v == 0 for k,v in summary.items() if k != "X payouts — clean (no block)"):
        st.success("✅ All clear — no blocked payouts, no B-blocks, no overdue open items.")

    today_fn = datetime.date.today().strftime("%Y%m%d")
    st.download_button(
        "⬇  Download full report",
        data=result,
        file_name=f"PayoutCheck_{today_fn}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, key="pbc_dl",
    )


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def show():
    st.markdown("## 🎁 Bonus & Payout Tools")
    st.caption("Customer matching against bonus files, and payout / payment block checker.")

    tab1, tab2 = st.tabs([
        "📋 Customer matching",
        "💸 Payout & block checker",
    ])

    with tab1:
        _show_matching()

    with tab2:
        _show_payout_checker()
