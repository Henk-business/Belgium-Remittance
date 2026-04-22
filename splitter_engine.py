"""Account splitter engine."""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import warnings
warnings.filterwarnings("ignore")

from common import BG, FG, c, mr, col_w, hdr_row, fd, auto_widths, clean_id
from template_manager import apply_template

# ── DOCUMENT TYPE TRANSLATION ─────────────────────────────────────────────────
_DOC_LABELS = {
    "en": {"RV+":"Invoice","RV-":"Credit note","ZP":"Payment","DZ":"Payment",
           "RS+":"Re-invoice (bonus correction)","RS-":"Bonus",
           "AB":"Clearing","X_PAY":"Payout to customer"},
    "nl": {"RV+":"Factuur","RV-":"Creditnota","ZP":"Betaling","DZ":"Betaling",
           "RS+":"Refactuur (bonuscorrectie)","RS-":"Bonus",
           "AB":"Verrekening","X_PAY":"Uitbetaling aan klant"},
    "fr": {"RV+":"Facture","RV-":"Note de crédit","ZP":"Paiement","DZ":"Paiement",
           "RS+":"Re-facturation (correction bonus)","RS-":"Bonus",
           "AB":"Ajustement comptable","X_PAY":"Virement au client"},
}

def translate_doc_types(df: "pd.DataFrame", lang: str = "en") -> "pd.DataFrame":
    """
    Replace Document Type column values with human-readable descriptions.
    Uses the same logic as the Customer Overview.
    Returns a copy of df with the column values replaced.
    """
    import pandas as _pd
    doc_type_col = next((c for c in df.columns if "document type" in c.lower()
                         or "belegtyp" in c.lower()), None)
    amt_col      = next((c for c in df.columns if "amount" in c.lower()
                         or "bedrag" in c.lower()), None)
    pay_col      = next((c for c in df.columns if "payment method" in c.lower()), None)
    if not doc_type_col:
        return df
    df = df.copy()
    labels = _DOC_LABELS.get(lang, _DOC_LABELS["en"])
    def _lbl(row):
        dt  = str(row.get(doc_type_col,"") or "").strip().upper()
        amt = float(row.get(amt_col, 0) or 0) if amt_col else 0
        pm  = str(row.get(pay_col,"") or "").strip().upper() if pay_col else ""
        if pm == "X":        return labels["X_PAY"]
        if dt == "RV":       return labels["RV+"] if amt >= 0 else labels["RV-"]
        if dt in ("ZP","DZ"):return labels["ZP"]
        if dt == "RS":       return labels["RS+"] if amt >= 0 else labels["RS-"]
        if dt == "AB":       return labels["AB"]
        return labels.get(dt, dt)
    df[doc_type_col] = df.apply(_lbl, axis=1)
    return df


CUSTOMER_CONFIG_KEY = "customer_configs"

# ── CHUNKED ACCOUNTS ──────────────────────────────────────────────────────────
# Accounts that need their output split into ~chunk_size batches.
# Add more accounts here as needed: "account_number": chunk_size
# Columns shown for chunked accounts (Account column removed).
CHUNKED_ACCOUNTS = {
    "30111788": {
        "chunk_size": 40000,
        "columns": [
            "Assignment", "Document Number", "Reference Key 3",
            "Document Date", "Net due date", "Document Type",
            "Amount in local currency",
        ],
    },
}

# Columns that are always stripped from the output — internal SAP fields
# that are not relevant to send to customers.
STRIP_COLS = {
    "Reason code",
    "Clerk Abbreviation",
    "Cleared/open items symbol",
    "Case ID",
    "Status",
    "Dunning Block",
    "Disputed item",
    "Payment Block",
    "Payment Method",
    "Net due date symbol",
    "G/L Account",
    "Text",
    "Clearing date",
    "Clearing Document",
    "Dunning Level",
    "Last Dunned",
    "Reversed with",
    "Document Header Text",
    "User Name",
    "Special G/L ind.",
    "Billing Document",
    "Reference Key 1",
    # normalised versions (after SAP column map)
    "text",
    "clearing_date",
    "clearing_doc",
    "header_text",
    "sap_class",
    "is_open",
    "ref",
    "doc_number_str",
}


def get_configs(state):
    return state.get(CUSTOMER_CONFIG_KEY, {})


def save_config(state, account_id, config):
    if CUSTOMER_CONFIG_KEY not in state:
        state[CUSTOMER_CONFIG_KEY] = {}
    state[CUSTOMER_CONFIG_KEY][str(account_id)] = config


def split_accounts(df, account_col, amount_col, due_date_col,
                   remove_not_due=True, reference_date=None,
                   customer_configs=None):
    if reference_date is None:
        reference_date = datetime.date.today()
    ref_ts = pd.Timestamp(reference_date)

    # ── Strip SAP-inserted subtotal/summary rows FIRST
    # Real transactions always have a Document Number. Rows without one are
    # SAP-generated subtotals (account totals, grand totals) that must be removed
    # before any splitting or totalling happens.
    doc_num_col = next(
        (c for c in df.columns
         if c.lower() in ("document number", "belegnummer", "doc_number", "doc number")),
        None
    )
    if doc_num_col:
        has_doc = (
            df[doc_num_col].notna() &
            ~df[doc_num_col].astype(str).str.strip().isin(["", "nan", "0", "0.0"])
        )
        df = df[has_doc].copy()

    accounts = sorted([
        clean_id(a) for a in df[account_col].dropna().unique()
        if clean_id(a) is not None
    ])

    result = {}
    for acc in accounts:
        mask = df[account_col].apply(clean_id) == acc
        acc_df = df[mask].copy()

        # ── Remove invoices not yet due so total reflects only displayed rows
        if remove_not_due and due_date_col and due_date_col in acc_df.columns:
            due = pd.to_datetime(acc_df[due_date_col], errors="coerce")
            acc_df = acc_df[due.isna() | (due <= ref_ts)].copy()

        # ── Filter to current year only (year of reference_date)
        # Prevents multi-year exports from including historical years
        if due_date_col and due_date_col in acc_df.columns:
            due_yr = pd.to_datetime(acc_df[due_date_col], errors="coerce").dt.year
            acc_df = acc_df[due_yr.isna() | (due_yr == ref_ts.year)].copy()

        # ── Strip unwanted SAP columns
        cols_to_drop = [col for col in acc_df.columns if col in STRIP_COLS]
        if cols_to_drop:
            acc_df = acc_df.drop(columns=cols_to_drop)

        if customer_configs:
            cfg = customer_configs.get(str(acc), {})
            cols = cfg.get("columns")
            if cols:
                valid = [col for col in cols if col in acc_df.columns]
                extra = [col for col in acc_df.columns if col not in valid]
                acc_df = acc_df[valid + extra]

        result[acc] = acc_df
    return result


def _safe_tab(name, idx):
    for ch in r"\/*?:[]":
        name = name.replace(ch, "_")
    return (name[:28] if len(name) > 28 else name) or f"Acct_{idx+1}"


def _recalc_arrears_df(df, today):
    """Recalculate Arrears after net due date based on reference date."""
    import pandas as _pd2
    ndd_col = next((c for c in df.columns if "net due" in c.lower()), None)
    arr_col = next((c for c in df.columns if "arrears" in c.lower()), None)
    if not ndd_col or not arr_col:
        return df
    df = df.copy()
    ref_ts = _pd2.Timestamp(today)
    due    = _pd2.to_datetime(df[ndd_col], errors="coerce")
    mask   = due.notna()
    days   = (ref_ts - due[mask]).dt.days
    # Write as same dtype as the column (str or numeric)
    if hasattr(df[arr_col], 'dtype') and str(df[arr_col].dtype) == 'string':
        df.loc[mask, arr_col] = days.astype(str)
    else:
        try:
            df[arr_col] = _pd2.to_numeric(df[arr_col], errors='coerce')
            df.loc[mask, arr_col] = days
        except Exception:
            df.loc[mask, arr_col] = days.astype(str)
    return df
def build_split_workbook(account_data, amount_col, today=None, title_prefix="", templates=None, lang="en"):
    if today is None:
        today = datetime.date.today()
    today_str = pd.Timestamp(today).strftime("%d/%m/%Y")
    # Recalculate arrears for each account df based on reference date
    account_data = {acc: _recalc_arrears_df(df, today) for acc, df in account_data.items()}
    # Translate Document Type codes to human-readable descriptions
    account_data = {acc: translate_doc_types(df, lang) for acc, df in account_data.items()}

    wb = openpyxl.Workbook()
    _default_sheet = wb.active  # track default sheet to remove after adding real sheets

    for idx, (acc, acc_df) in enumerate(account_data.items()):
        tab_name = _safe_tab(str(acc), idx)

        # No placeholders in standard layout — render everything as plain sheets.
        if False:  # placeholder logic disabled for standard download
            pass
            # Try to write the template-based sheet into the combined workbook
            try:
                from poc_builder import build_poc_sheet, _load_poc_names
                import io as _io, openpyxl as _oxl
                tmpl_b = templates[str(acc)]
                # Detect POC template
                _twb  = _oxl.load_workbook(_io.BytesIO(tmpl_b))
                _tws  = _twb.active
                _maxr = min((_tws.max_row or 20), 20)
                is_poc = any(
                    str(_tws.cell(r, 1).value or '').strip().startswith('29')
                    for r in range(1, _maxr + 1)
                )
                if is_poc:
                    sheet_bytes = build_poc_sheet(acc_df, str(acc), tmpl_b, today=today)
                else:
                    sheet_bytes = build_template_sheet(str(acc), acc_df, tmpl_b,
                                                       amount_col, today=today)
                # Copy the first sheet from the result into our combined workbook
                src_wb = _oxl.load_workbook(_io.BytesIO(sheet_bytes))
                src_ws = src_wb.active
                dst_ws = wb.create_sheet(title=tab_name)
                for row in src_ws.iter_rows():
                    for cell in row:
                        dst = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            from copy import copy as _copy
                            dst.font      = _copy(cell.font)
                            dst.fill      = _copy(cell.fill)
                            dst.border    = _copy(cell.border)
                            dst.alignment = _copy(cell.alignment)
                            dst.number_format = cell.number_format
                for i in range(1, src_ws.max_column + 1):
                    ltr = __import__("openpyxl").utils.get_column_letter(i)
                    dst_ws.column_dimensions[ltr].width =                         src_ws.column_dimensions[ltr].width
                for r_idx in range(1, src_ws.max_row + 1):
                    dst_ws.row_dimensions[r_idx].height =                         src_ws.row_dimensions[r_idx].height
            except Exception as _te:
                # Log the error in the sheet itself for debugging
                ws_ph = wb.create_sheet(title=tab_name)
                mr(ws_ph, 1, 1, 5,
                   f"Account {acc} — template error: {str(_te)[:80]}",
                   bg="md_blue", fg="white", sz=10)
            continue

        ws = wb.create_sheet(title=tab_name)
        ncols = len(acc_df.columns)

        r2 = 1
        mr(ws, r2, 1, ncols, "Account: " + str(acc),
           bold=True, bg="dk_blue", fg="white", sz=13, ha="center")
        ws.row_dimensions[r2].height = 32
        r2 += 1
        mr(ws, r2, 1, ncols,
           today_str + "  \u00b7  " + f"{len(acc_df):,}" + " lines  \u00b7  Invoices not yet due removed",
           bg="md_blue", fg="white", sz=9, ha="center")
        ws.row_dimensions[r2].height = 16
        r2 += 2

        for ci, col_name in enumerate(acc_df.columns, 1):
            display_name = "Description" if "document type" in str(col_name).lower() else str(col_name)
            cell = ws.cell(row=r2, column=ci, value=display_name)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor=BG["md_blue"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[r2].height = 18
        r2 += 1

        amount_ci = None
        if amount_col and amount_col in acc_df.columns:
            amount_ci = list(acc_df.columns).index(amount_col) + 1

        # Identify date columns (exclude "Arrears" column which contains days, not dates)
        date_cols = set()
        for ci_chk, col_name in enumerate(acc_df.columns, 1):
            if "arrears" in col_name.lower():
                continue
            if any(kw in col_name.lower() for kw in ["date", "datum"]):
                date_cols.add(ci_chk)
            elif pd.api.types.is_datetime64_any_dtype(acc_df[col_name]):
                date_cols.add(ci_chk)

        for ri, (_, row) in enumerate(acc_df.iterrows(), r2):
            bg = "grey" if ri % 2 == 0 else "white"
            for ci, (col_name, val) in enumerate(row.items(), 1):
                is_amt  = (ci == amount_ci)
                is_date = (ci in date_cols)
                if is_amt:
                    cell_val = float(val) if pd.notna(val) else 0.0
                elif is_date:
                    try:
                        cell_val = pd.Timestamp(val).to_pydatetime() if pd.notna(val) else ""
                    except Exception:
                        cell_val = str(val) if pd.notna(val) else ""
                elif pd.isna(val):
                    cell_val = ""
                elif isinstance(val, float) and val == int(val):
                    cell_val = int(val)
                else:
                    cell_val = val
                cell = ws.cell(row=ri, column=ci, value=cell_val)
                fg_col = (
                    "FFC00000" if is_amt and isinstance(cell_val, (int, float)) and cell_val >= 0
                    else "FF375623" if is_amt and isinstance(cell_val, (int, float)) and cell_val < 0
                    else "000000"
                )
                cell.font = Font(name="Arial", size=9, color=fg_col)
                cell.fill = PatternFill("solid", fgColor=BG.get(bg, "FFFFFF"))
                cell.alignment = Alignment(
                    horizontal="right" if is_amt else "left", vertical="center"
                )
                if is_amt:
                    cell.number_format = "#,##0.00"
                elif is_date and isinstance(cell_val, datetime.datetime):
                    cell.number_format = "DD/MM/YYYY"
            ws.row_dimensions[ri].height = 13

        total_r = r2 + len(acc_df)
        c(ws, total_r, 1, "TOTAL", bold=True, bg="dk_blue", fg="white", sz=10)
        for ci in range(2, ncols + 1):
            c(ws, total_r, ci, None, bg="dk_blue")
        if amount_ci:
            total_val = acc_df[amount_col].sum()
            cell = ws.cell(row=total_r, column=amount_ci, value=total_val)
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BG["dk_blue"])
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[total_r].height = 18

        auto_widths(ws, acc_df, start_col=1)
        ws.freeze_panes = "A5"

    # Remove the default empty sheet openpyxl creates automatically
    if _default_sheet in wb.worksheets and len(wb.worksheets) > 1:
        wb.remove(_default_sheet)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_template_sheet(account_id: str, acc_df: pd.DataFrame,
                         template_bytes: bytes, amount_col: str,
                         today=None) -> bytes:
    """
    Build a single-account workbook using the customer's custom template.
    Returns the filled workbook as bytes.
    Falls back to the standard layout if template application fails.
    """
    from template_manager import apply_template
    today_str = pd.Timestamp(today or datetime.date.today()).strftime("%d/%m/%Y")

    try:
        result = apply_template(template_bytes, acc_df)
        # Post-process: update account number and title in any title rows
        import openpyxl, io as _io, re as _re
        wb = openpyxl.load_workbook(_io.BytesIO(result))
        ws = wb.active
        new_title = str(account_id)[:31]
        # Rename any conflicting sheets first
        for other_ws in list(wb.worksheets):
            if other_ws != ws and other_ws.title == new_title:
                other_ws.title = "_old_" + other_ws.title[:26]
        ws.title = new_title
        for r in range(1, min(6, ws.max_row + 1)):
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(r, ci)
                val  = str(cell.value or "")
                if not val:
                    continue
                # Replace old account number with new one
                new_val = _re.sub(r"\b3\d{7}\b", str(account_id), val)
                # Update line count in subtitle
                new_val = _re.sub(r"\d+ lines", f"{len(acc_df)} lines", new_val)
                new_val = _re.sub(r"\d+ lignes", f"{len(acc_df)} lignes", new_val)
                new_val = _re.sub(r"\d+ regels", f"{len(acc_df)} regels", new_val)
                # Update date
                new_val = _re.sub(r"\d{2}/\d{2}/\d{4}", today_str, new_val)
                if new_val != val:
                    cell.value = new_val
        out = _io.BytesIO()
        wb.save(out)
        return out.getvalue()
    except Exception:
        # Fallback: standard layout for this account
        data = build_split_workbook(
            {account_id: acc_df}, amount_col,
            today=today,
            title_prefix=f"Account {account_id} — ",
        )
        return data.getvalue()


def build_individual_sheet(acc_id: str, acc_df, amount_col: str, today=None) -> bytes:
    if today is not None:
        acc_df = _recalc_arrears_df(acc_df, today)
    """
    Build the correct individual workbook for one account.
    - If account is in CHUNKED_ACCOUNTS: apply chunking rules
    - Otherwise: standard layout
    Returns bytes.
    """
    from chunked_builder import build_chunked_sheet

    cfg = CHUNKED_ACCOUNTS.get(str(acc_id))
    if cfg:
        rule = {
            "chunk_size":     cfg.get("chunk_size", 40000),
            "show_account":   False,
            "total_position": cfg.get("total_position", "right"),
            "columns":        cfg.get("columns", []),
            "sort_by":        cfg.get("sort_by", ["Net due date"]),
        }
        return build_chunked_sheet(acc_df, str(acc_id), rule, today=today)
    else:
        wb_obj = build_split_workbook(
            {acc_id: acc_df}, amount_col,
            today=today, title_prefix=f"Account {acc_id} \u2014 "
        )
        return wb_obj.getvalue()
