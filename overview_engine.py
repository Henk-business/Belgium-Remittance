"""
Customer yearly overview engine.

GROUPING: Uses SAP blank-row separators (rows with no Account) to identify groups.
YEAR RULE: A group belongs to the year of the oldest positive RV invoice in it.
           If no positive RV exists, uses the oldest document date in the group.
SORT:      Within each year: groups sorted newest→oldest by the oldest doc date.
           Within each group: rows in original SAP order (preserved as-is).
G/L SPLIT: If multiple G/L accounts in a year, split into sub-sections with subtotals.
COLUMNS:   Same strip rules as Account Splitter PLUS keep Clearing date, Clearing Document,
           Payment Method, G/L Account. Document Type replaced with description.
LANGUAGE:  EN / NL / FR.
COLOURS:   Positive = RED (invoices), Negative = GREEN (credits/payments).
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import warnings
warnings.filterwarnings("ignore")

# ── COLOURS ───────────────────────────────────────────────────────────────────
DK_BLUE  = "FF1F3864"
MD_BLUE  = "FF2E75B6"
LT_BLUE  = "FFBDD7EE"
YELLOW   = "FFEE09"
WHITE    = "FFFFFFFF"
GREY     = "FFF2F2F2"
POS_FG   = "FFC00000"   # red   – invoices / positive
NEG_FG   = "FF375623"   # green – credits / payments / negative
BLACK_FG = "FF000000"

# ── STRIP COLUMNS (same as splitter, but KEEP the overview-specific ones) ─────
STRIP_COLS = {
    "Reason code","Clerk Abbreviation","Cleared/open items symbol",
    "Case ID","Status","Dunning Block","Disputed item","Payment Block",
    "Net due date symbol","Text","Dunning Level","Last Dunned",
    "Reversed with","Document Header Text","User Name","Special G/L ind.",
    "Billing Document","Reference Key 1",
    "Clearing date","Clearing Document",
    # KEPT: Payment Method, G/L Account
}

GL_LABELS = {
    "en": {"2400000": "Beer",  "2530009": "Rent"},
    "nl": {"2400000": "Bier",  "2530009": "Huur"},
    "fr": {"2400000": "Bière", "2530009": "Loyer"},
}

T = {
    "en": {
        "title_suffix":    "Customer Overview",
        "subtitle":        "Grouped by clearing document  ·  Year = oldest invoice in group  ·  Positive = invoices (red)  ·  Negative = credits / payments (green)",
        "year_banner":     "{yr}  ·  {n} groups  ·  Invoices: €{inv}  ·  Credits: €{cred}  ·  Net: €{net}",
        "year_total":      "{yr} — Total",
        "grand_total":     "Grand Total {a}–{b}",
        "net_balance":     "Net Balance",
        "no_transactions": "No transactions in {yr}",
        "gl_subtotal":     "{lbl} — Subtotal",
        "group_subtotal":  "Subtotal",
        "gl_other":        "Other",
        "desc_col":        "Description",
        "doc_types": {
            "RV+": "Invoice",          "RV-": "Credit note",
            "ZP":  "Payment",          "DZ":  "Payment",
            "RS+": "Re-invoice",       "RS-": "Bonus",
            "AB":  "Clearing",         "X_PAY": "Payout to customer",
        },
    },
    "nl": {
        "title_suffix":    "Klantoverzicht",
        "subtitle":        "Gegroepeerd per verrekeningsdocument  ·  Jaar = oudste factuur in groep  ·  Positief = facturen (rood)  ·  Negatief = creditnota's / betalingen (groen)",
        "year_banner":     "{yr}  ·  {n} groepen  ·  Facturen: €{inv}  ·  Creditnota's: €{cred}  ·  Netto: €{net}",
        "year_total":      "{yr} — Totaal",
        "grand_total":     "Eindtotaal {a}–{b}",
        "net_balance":     "Nettosaldo",
        "no_transactions": "Geen transacties in {yr}",
        "gl_subtotal":     "{lbl} — Subtotaal",
        "group_subtotal":  "Subtotaal",
        "gl_other":        "Overig",
        "desc_col":        "Omschrijving",
        "doc_types": {
            "RV+": "Factuur",          "RV-": "Creditnota",
            "ZP":  "Betaling",         "DZ":  "Betaling",
            "RS+": "Refactuur",        "RS-": "Bonus",
            "AB":  "Verrekening",      "X_PAY": "Uitbetaling aan klant",
        },
    },
    "fr": {
        "title_suffix":    "Aperçu client",
        "subtitle":        "Groupé par document de compensation  ·  Année = facture la plus ancienne  ·  Positif = factures (rouge)  ·  Négatif = avoirs / paiements (vert)",
        "year_banner":     "{yr}  ·  {n} groupes  ·  Factures: €{inv}  ·  Notes de crédit: €{cred}  ·  Net: €{net}",
        "year_total":      "{yr} — Total",
        "grand_total":     "Total général {a}–{b}",
        "net_balance":     "Solde net",
        "no_transactions": "Aucune transaction en {yr}",
        "gl_subtotal":     "{lbl} — Sous-total",
        "group_subtotal":  "Sous-total",
        "gl_other":        "Autre",
        "desc_col":        "Description",
        "doc_types": {
            "RV+": "Facture",          "RV-": "Note de crédit",
            "ZP":  "Paiement",         "DZ":  "Paiement",
            "RS+": "Re-facturation",   "RS-": "Bonus",
            "AB":  "Ajustement comptable",     "X_PAY": "Virement au client",
        },
    },
}


def _t(lang, key, **kw):
    val = T.get(lang, T["en"]).get(key, T["en"].get(key, key))
    try:    return val.format(**kw) if kw else val
    except: return val


def _desc(doc_type, amount, pay_method, lang):
    dt  = str(doc_type  or "").strip().upper()
    pm  = str(pay_method or "").strip().upper()
    amt = float(amount) if str(amount) not in ("", "nan") else 0.0
    d   = T.get(lang, T["en"])["doc_types"]
    if pm == "X":          return d["X_PAY"]
    if dt == "RV":         return d["RV+"] if amt >= 0 else d["RV-"]
    if dt in ("ZP","DZ"):  return d["ZP"]
    if dt == "RS":         return d["RS+"] if amt >= 0 else d["RS-"]
    if dt == "AB":         return d["AB"]
    return d.get(dt, "")


def _gl_lbl(gl, lang):
    g = str(gl or "").strip().split(".")[0]
    return GL_LABELS.get(lang, GL_LABELS["en"]).get(g,
           _t(lang, "gl_other") if g else "")


# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────
def _fill(rgb):   return PatternFill("solid", fgColor=rgb)
def _font(bold=False, color=BLACK_FG, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _align(ha="left"):
    return Alignment(horizontal=ha, vertical="center")
def _thin():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _w(ws, row, col, val=None, bold=False, bg=WHITE, fg=BLACK_FG,
       size=10, ha="left", fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _font(bold=bold, color=fg, size=size)
    c.fill = _fill(bg); c.alignment = _align(ha); c.border = _thin()
    if fmt: c.number_format = fmt
    return c

def _mw(ws, row, c1, c2, val=None, bold=False, bg=WHITE,
        fg=BLACK_FG, size=10, ha="center"):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    c = ws.cell(row=row, column=c1, value=val)
    c.font = _font(bold=bold, color=fg, size=size)
    c.fill = _fill(bg); c.alignment = _align(ha); c.border = _thin()
    for col in range(c1+1, c2+1):
        ws.cell(row=row, column=col).fill   = _fill(bg)
        ws.cell(row=row, column=col).border = _thin()
    return c


# ── PARSE ─────────────────────────────────────────────────────────────────────

def prepare_df(file_obj):
    """Read raw SAP export, stripping only trailing SAP grand-total rows."""
    df = pd.read_excel(file_obj, sheet_name=0, header=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    for col in df.columns:
        if any(kw in col.lower() for kw in ["date", "datum"]):
            if "arrears" in col.lower():
                continue
            df[col] = pd.to_datetime(df[col], errors="coerce")
    amt_col = next((c for c in df.columns if "amount" in c.lower()
                    or "bedrag" in c.lower() or "betrag" in c.lower()), None)
    if amt_col:
        df[amt_col] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)

    # Strip SAP-generated grand-total rows that appear at the TRAILING end of the file.
    # Mid-file blank rows (clearing group separators) must be PRESERVED — build_overview
    # uses them to detect the clearing-group structure and split current vs historical.
    # Only strip rows where BOTH Account AND Document Number are blank AND no real
    # account row appears after them (i.e. they are truly at the end).
    acc_col_pd = next((c for c in df.columns if c.lower() == "account"), None)
    doc_col_pd = next((c for c in df.columns if "document number" in c.lower()), None)
    if acc_col_pd and doc_col_pd:
        is_blank = (
            (df[acc_col_pd].isna() | df[acc_col_pd].astype(str).str.strip().isin(["", "nan", "None"])) &
            (df[doc_col_pd].isna() | df[doc_col_pd].astype(str).str.strip().isin(["", "nan", "None"]))
        )
        # Find the last row that has a real account number
        real_mask = ~(df[acc_col_pd].isna() | df[acc_col_pd].astype(str).str.strip().isin(["", "nan", "None"]))
        if real_mask.any():
            last_real = real_mask[real_mask].index[-1]
            # Only strip blank rows that come AFTER the last real row (trailing totals)
            trailing_blank = is_blank & (df.index > last_real)
            df = df[~trailing_blank].reset_index(drop=True)
        else:
            df = df[~is_blank].reset_index(drop=True)

    return df, amt_col


def _recalc_arrears(df: pd.DataFrame, reference_date) -> pd.DataFrame:
    """Recalculate arrears based on reference date: (ref - net_due_date).days"""
    if reference_date is None:
        return df
    ndd_col = next((c for c in df.columns if "net due" in c.lower()), None)
    arr_col = next((c for c in df.columns if "arrears" in c.lower()), None)
    if not ndd_col or not arr_col:
        return df
    df = df.copy()
    ref_ts = pd.Timestamp(reference_date)
    due    = pd.to_datetime(df[ndd_col], errors="coerce")
    mask   = due.notna()
    days   = (ref_ts - due[mask]).dt.days
    if str(df[arr_col].dtype) == 'string':
        df.loc[mask, arr_col] = days.astype(str)
    else:
        try:
            df[arr_col] = pd.to_numeric(df[arr_col], errors='coerce')
            df.loc[mask, arr_col] = days
        except Exception:
            df.loc[mask, arr_col] = days.astype(str)
    return df
def _parse_groups(df, amt_col):
    """
    Split df into groups using blank Account rows as separators.
    Each group is a list of row dicts (real SAP rows only, no blank rows).
    Returns list of groups, each group = list of row Series.
    """
    acc_col = next((c for c in df.columns
                    if c.lower() in ("account","customer","debitor","konto")), None)
    groups, current = [], []
    for _, row in df.iterrows():
        acc = str(row.get(acc_col, "") or "").strip() if acc_col else ""
        if acc and acc not in ("nan", "None", ""):
            current.append(row)
        else:
            if current:
                groups.append(current)
                current = []
    if current:
        groups.append(current)
    return groups


def _group_year(group, doc_date_col, doc_type_col, amt_col):
    """
    Year of a group = year of the oldest net due date among NON-clearing rows.
    AB/Clearing rows are excluded so one old clearing doesn't drag current
    invoices into a past year. Fallback: all rows, then doc date.
    """
    net_due_col = next((c for c in group[0].index
                        if "net due" in c.lower()
                        or "vervaldatum" in c.lower()), None) if group else None

    SKIP_TYPES = {"AB", "ZP", "DZ"}   # clearing/payment types to skip for year assignment

    due_dates_inv, due_dates_all = [], []
    doc_dates_inv, doc_dates_all = [], []
    for row in group:
        dt = str(row.get(doc_type_col, "") or "").strip().upper() if doc_type_col else ""
        is_clearing = dt in SKIP_TYPES
        if net_due_col:
            nd = row.get(net_due_col)
            if nd is not None and pd.notna(nd):
                due_dates_all.append(nd)
                if not is_clearing:
                    due_dates_inv.append(nd)
        if doc_date_col:
            dd = row.get(doc_date_col)
            if dd is not None and pd.notna(dd):
                doc_dates_all.append(dd)
                if not is_clearing:
                    doc_dates_inv.append(dd)

    # Prefer invoice dates; fall back to all dates
    dates = (due_dates_inv or due_dates_all or
             doc_dates_inv or doc_dates_all)
    if not dates:
        return None
    oldest = min(dates)
    return oldest.year if hasattr(oldest, "year") else int(str(oldest)[:4])


# ── BUILD ─────────────────────────────────────────────────────────────────────


def build_current_overview(df: pd.DataFrame, amt_col: str,
                            reference_date=None,
                            remove_not_due: bool = False,
                            remove_overdues: bool = False,
                            account_id: str = "",
                            customer_name: str = "",
                            lang: str = "en") -> BytesIO:
    """
    Current overview — shows ONLY the current open items section.
    These are the rows that appear before the first DZ/ZP (payment) row
    in the SAP export, i.e. the open/outstanding invoices.

    remove_not_due: removes rows where net due date > reference_date (not yet due)
    remove_overdues: removes rows where net due date < reference_date (already overdue)
"""
    import datetime as _dt
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ref_ts  = pd.Timestamp(reference_date) if reference_date else pd.Timestamp.now()

    ndd_col      = next((c for c in df.columns if "net due"       in c.lower()), None)
    arr_col      = next((c for c in df.columns if "arrears"       in c.lower()), None)
    doc_type_col = next((c for c in df.columns if "document type" in c.lower()), None)
    acc_col      = next((c for c in df.columns if c.lower() in ("account","konto","debitor")), None)
    pay_col      = next((c for c in df.columns if "payment method" in c.lower()), None)

    df = df.copy()
    if ndd_col: df[ndd_col] = pd.to_datetime(df[ndd_col], errors="coerce")
    if amt_col: df[amt_col] = pd.to_numeric(df[amt_col],  errors="coerce")
    if arr_col: df[arr_col] = pd.to_numeric(df[arr_col],  errors="coerce")

    # ── Determine export structure ────────────────────────────────────────────
    # Blank rows in SAP exports mean two different things:
    # (A) GL subtotal rows — in flat open-items exports with Beer+Rent GL codes.
    #     These are single rows with blank acc+doc appearing once or twice.
    # (B) Clearing group separators — in multi-year exports with full history.
    #     These appear many times, between each cleared group.
    #
    # Rule: if blank rows appear in the middle, count them.
    # If there are only 1-2 blank rows in the middle AND all real rows are open
    # (no clearing doc), treat as flat open-items (GL subtotals only) — no split.
    # If there are 3+ blank rows in the middle, it's a clearing-group export — split.
    doc_col = next((c for c in df.columns if "document number" in c.lower()), None)

    if acc_col:
        is_blank_acc = df[acc_col].isna() | df[acc_col].astype(str).str.strip().isin(["", "nan", "None"])
        is_blank_doc = (df[doc_col].isna() | df[doc_col].astype(str).str.strip().isin(["", "nan", "None"])) \
                       if doc_col else is_blank_acc
        is_blank_mask = is_blank_acc & is_blank_doc
        last_real_idx = df[~is_blank_acc].index.max() if (~is_blank_acc).any() else -1
        mid_blanks    = is_blank_mask & (df.index < last_real_idx)
        n_mid_blanks  = mid_blanks.sum()
        blanks_in_middle = n_mid_blanks > 0

        # If only 1 or 2 blank rows in the middle, this is very likely a flat
        # open-items export with GL section subtotals (Beer/Rent), NOT a
        # clearing-group history export. Don't split on the first ZP row.
        if blanks_in_middle and n_mid_blanks <= 2:
            blanks_in_middle = False
    else:
        blanks_in_middle = False

    if blanks_in_middle:
        # Multi-group export: the FIRST blank separator row marks the end of current open.
        # Everything before the first blank separator = current open items.
        # Everything from the first blank onwards = cleared historical groups.
        first_blank_idx = df[is_blank_mask].index.min() if is_blank_mask.any() else None
        if first_blank_idx is not None:
            df = df[df.index < first_blank_idx].copy()
    # else: no mid-file blank rows → entire file is current open items, use as-is

    # Strip blank/SAP-total rows — keep only real account rows
    if acc_col:
        is_real = df[acc_col].notna() & ~df[acc_col].astype(str).str.strip().isin(["", "nan", "None"])
        df = df[is_real].copy()

    # ── Remove zero-net clearing groups from current open ─────────────────────
    # Some exports contain balanced clearing pairs (AB type, net = 0) sitting
    # before the first DZ/ZP row, so they land in "current open" incorrectly.
    # Any group of rows that nets to exactly zero is already settled — remove it.
    # Rows with no reference key are standalone open items — always keep them.
    if amt_col and amt_col in df.columns:
        _amt_num = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)
        _ref_col = next((c for c in df.columns if "reference key 3" in c.lower()), None) or \
                   next((c for c in df.columns if "assignment" in c.lower()), None)
        if _ref_col and _ref_col in df.columns:
            _ref_vals = df[_ref_col].astype(str).str.strip()
            # Rows with no reference key are always kept (standalone open items)
            _no_ref = _ref_vals.isin(["", "nan", "None"]) | df[_ref_col].isna()
            # For rows that DO have a reference, keep only non-zero-net groups
            _group_nets = _amt_num[~_no_ref].groupby(_ref_vals[~_no_ref]).sum().abs()
            _open_refs  = _group_nets[_group_nets > 0.01].index
            _keep_mask  = _no_ref | _ref_vals.isin(_open_refs)
            df = df[_keep_mask].copy()

    # ── Recalculate arrears against reference_date ────────────────────────────
    df = _recalc_arrears(df, ref_ts.date())
    if arr_col: df[arr_col] = pd.to_numeric(df[arr_col], errors="coerce")

    # ── remove_not_due: hide rows where net due date is AFTER reference date ──
    # i.e. arrears < 0 (not yet overdue)
    if remove_not_due and ndd_col:
        due = df[ndd_col]
        df = df[due.isna() | (due <= ref_ts)].copy()

    # ── remove_overdues: hide rows where net due date is BEFORE reference date
    # i.e. arrears > 0 (already past due)
    if remove_overdues and ndd_col:
        due = df[ndd_col]
        df = df[due.isna() | (due >= ref_ts)].copy()


    # ── Group into clearing-doc groups using assignment/document grouping ──────
    # For the flat current overview we treat rows as one big group — no SAP
    # blank-row separators exist in the open-items section.
    # Sort newest net due date first.
    SKIP = {"AB", "ZP", "DZ"}
    if ndd_col:
        df = df.sort_values(ndd_col, ascending=False, na_position="last")

    # ── Colours ───────────────────────────────────────────────────────────────
    HDR_FILL  = "FF1F3864"
    ROW_WHITE = "FFFFFFFF"
    ROW_BLUE  = "FFF2F2F2"
    ROW_YELL  = "FFFFFF00"
    COL_POS   = "FFC00000"
    COL_NEG   = "FF375623"
    COL_WHT   = "FFFFFFFF"
    COL_BLK   = "FF000000"

    def _fill(rgb): return PatternFill("solid", fgColor=rgb)
    def _font(bold=False, color=COL_BLK, size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def _thin():
        s = Side(style="thin", color="DDDDDD")
        return Border(left=s, right=s, top=s, bottom=s)
    def _aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=False)

    STRIP = {
        "Reason code","Clerk Abbreviation","Cleared/open items symbol",
        "Disputed item","Payment Block","Net due date symbol",
        "Text","Clearing date","Clearing Document","Dunning Level",
        "Last Dunned","Reversed with","Document Header Text","User Name",
        "Special G/L ind.","Billing Document","Reference Key 1",
        "doc_number_str","ref","sap_class","is_open","header_text",
        "clearing_date","clearing_doc","text",
        # Always hidden in current overview
        "Case ID", "Status", "Dunning Block",
    }
    display_cols = [c for c in df.columns if c not in STRIP]
    ncols = len(display_cols)
    col_widths = {
        "Account":10,"Assignment":14,"Document Number":18,
        "Reference Key 3":14,"Document Date":13,"Net due date":13,
        "Document Type":26,"Amount in local currency":20,
        "Arrears after net due date":24,"Payment Method":13,
        "G/L Account":22,"Disputed item":13,
    }

    # ── G/L code → label mapping (language-aware) ─────────────────────────────
    _gl_map   = GL_LABELS.get(lang, GL_LABELS["en"])
    gl_col    = next((c for c in df.columns if "g/l" in c.lower() or "gl account" in c.lower()), None)

    # Determine which GL categories are present in the data
    _categories_present = set()
    if gl_col and gl_col in df.columns:
        for v in df[gl_col].dropna().astype(str):
            if v in _gl_map:
                _categories_present.add(v)

    _has_multi_cat = len(_categories_present) > 1
    _gl_ci = (display_cols.index(gl_col) + 1) if gl_col and gl_col in display_cols else None

    def _gl_label(raw_val):
        """Return G/L code with category label in brackets if known."""
        s = str(raw_val) if raw_val is not None else ""
        lbl = _gl_map.get(s)
        return f"{s} ({lbl})" if lbl else s

    # ── Subtotal helpers ──────────────────────────────────────────────────────
    _sub_lbl_tmpl = {"en": "{lbl} — Subtotal", "nl": "{lbl} — Subtotaal",
                     "fr": "{lbl} — Sous-total"}
    def _subtotal_label(gl_code):
        lbl = _gl_map.get(gl_code, gl_code)
        tmpl = _sub_lbl_tmpl.get(lang, _sub_lbl_tmpl["en"])
        return tmpl.format(lbl=lbl)

    def _write_subtotal_row(ws_obj, row_num, label, total, n_cols, a_ci):
        BAND = "FF2E75B6"
        for ci in range(1, n_cols + 1):
            cell = ws_obj.cell(row_num, ci)
            cell.fill   = _fill(BAND)
            cell.border = _thin()
        ws_obj.cell(row_num, 1).value     = label
        ws_obj.cell(row_num, 1).font      = _font(bold=True, color=COL_WHT, size=9)
        ws_obj.cell(row_num, 1).alignment = _aln("left")
        if a_ci:
            c = ws_obj.cell(row_num, a_ci)
            c.value         = total
            c.font          = _font(bold=True, color=COL_WHT, size=9)
            c.number_format = "#,##0.00"
            c.alignment     = _aln("right")
        ws_obj.row_dimensions[row_num].height = 16

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overview"
    for ci, col in enumerate(display_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, max(len(str(col))+2, 12))

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    _acc = account_id or (
        str(df[acc_col].dropna().iloc[0]).strip().split(".")[0]
        if acc_col and len(df[acc_col].dropna()) > 0 else ""
    )
    ref_yr = ref_ts.year
    _title_suffix = {"en": "Current Overview", "nl": "Huidig Overzicht", "fr": "Aperçu Actuel"}
    title_val = f"Account {_acc}  ·  {ref_yr}  {_title_suffix.get(lang, 'Current Overview')}"
    if customer_name:
        title_val = f"{customer_name}  ·  " + title_val
    for ci in range(1, ncols+1):
        ws.cell(1, ci).fill   = _fill(HDR_FILL)
        ws.cell(1, ci).border = _thin()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(1, 1).value     = title_val
    ws.cell(1, 1).font      = _font(bold=True, color=COL_WHT, size=14)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 34

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    BAND_FILL = "FF2E75B6"
    today_str = ref_ts.strftime("%d/%m/%Y")
    _sub = {"en": "Current open items  ·  Positive = invoices (red)  ·  Negative = credits / payments (green)",
            "nl": "Huidige openstaande posten  ·  Positief = facturen (rood)  ·  Negatief = creditnota's / betalingen (groen)",
            "fr": "Postes ouverts actuels  ·  Positif = factures (rouge)  ·  Négatif = avoirs / paiements (vert)"}
    subtitle_val = f"{_sub.get(lang, _sub['en'])}  ·  {today_str}"
    for ci in range(1, ncols+1):
        ws.cell(2, ci).fill   = _fill(BAND_FILL)
        ws.cell(2, ci).border = _thin()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(2, 1).value     = subtitle_val
    ws.cell(2, 1).font      = _font(bold=False, color=COL_WHT, size=9)
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[2].height = 16

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    for ci, col in enumerate(display_cols, 1):
        h = "Description" if doc_type_col and col == doc_type_col else col
        cell = ws.cell(3, ci, value=h)
        cell.font = _font(bold=True, color=COL_WHT, size=9)
        cell.fill = _fill(HDR_FILL)
        cell.alignment = _aln("center")
        cell.border = _thin()
    ws.row_dimensions[3].height = 15
    ws.freeze_panes = "A4"

    r = 4
    amt_ci = (display_cols.index(amt_col) + 1) if amt_col and amt_col in display_cols else None

    # ── Data rows — group by GL category if multiple present ──────────────────
    if _has_multi_cat and gl_col:
        # Sort rows into buckets by GL category
        _buckets = {}   # gl_code -> list of (idx, row)
        _other   = []   # rows with unrecognised GL codes
        for idx, row in df.iterrows():
            gl_val = str(row.get(gl_col, "") or "").strip()
            if gl_val in _categories_present:
                _buckets.setdefault(gl_val, []).append((idx, row))
            else:
                _other.append((idx, row))

        row_idx_global = 0
        _cat_totals = {}

        for gl_code in sorted(_categories_present):
            rows_in_cat = _buckets.get(gl_code, [])
            cat_total   = 0.0
            for idx, row in rows_in_cat:
                bg = ROW_WHITE if row_idx_global % 2 == 0 else ROW_BLUE
                for ci, col in enumerate(display_cols, 1):
                    val = row.get(col, "")
                    if doc_type_col and col == doc_type_col:
                        val = _desc(val, row.get(amt_col, 0), row.get(pay_col, "") if pay_col else "", lang)
                    elif gl_col and col == gl_col:
                        val = _gl_label(val)
                    elif isinstance(val, pd.Timestamp):
                        val = val.to_pydatetime()
                    elif not isinstance(val, (str, int, float, _dt.datetime, type(None))):
                        val = str(val)
                    elif isinstance(val, float):
                        if val != val: val = None
                        elif val == int(val): val = int(val)
                    is_amt = amt_col and col == amt_col
                    cell = ws.cell(r, ci, value=val if val != "" else None)
                    if is_amt and isinstance(val, (int, float)) and val is not None:
                        cat_total += float(val)
                        cell.font = _font(color=COL_POS if val > 0 else (COL_NEG if val < 0 else COL_BLK))
                        cell.number_format = "#,##0.00"
                        cell.alignment = _aln("right")
                    elif isinstance(val, _dt.datetime):
                        cell.font = _font()
                        cell.number_format = "DD/MM/YYYY"
                        cell.alignment = _aln("left")
                    else:
                        cell.font = _font()
                        cell.alignment = _aln("left")
                    cell.fill = _fill(bg)
                    cell.border = _thin()
                ws.row_dimensions[r].height = 13
                r += 1
                row_idx_global += 1

            _cat_totals[gl_code] = cat_total
            _write_subtotal_row(ws, r, _subtotal_label(gl_code), cat_total, ncols, amt_ci)
            r += 1

        # Any rows with other GL codes
        for idx, row in _other:
            bg = ROW_WHITE if row_idx_global % 2 == 0 else ROW_BLUE
            for ci, col in enumerate(display_cols, 1):
                val = row.get(col, "")
                if gl_col and col == gl_col:
                    val = _gl_label(val)
                elif isinstance(val, pd.Timestamp):
                    val = val.to_pydatetime()
                elif not isinstance(val, (str, int, float, _dt.datetime, type(None))):
                    val = str(val)
                elif isinstance(val, float):
                    if val != val: val = None
                    elif val == int(val): val = int(val)
                cell = ws.cell(r, ci, value=val if val != "" else None)
                cell.fill = _fill(bg)
                cell.border = _thin()
                cell.font = _font()
            ws.row_dimensions[r].height = 13
            r += 1
            row_idx_global += 1

        grand_total = sum(_cat_totals.values())
        if _other and amt_col in df.columns:
            grand_total += sum(float(row.get(amt_col, 0) or 0) for _, row in _other)

    else:
        # Single category or no GL column — flat display
        grand_total = 0.0
        for row_idx, (_, row) in enumerate(df.iterrows()):
            bg = ROW_WHITE if row_idx % 2 == 0 else ROW_BLUE
            for ci, col in enumerate(display_cols, 1):
                val = row.get(col, "")
                if doc_type_col and col == doc_type_col:
                    pm  = row.get(pay_col, "") if pay_col else ""
                    val = _desc(val, row.get(amt_col, 0), pm, lang)
                elif gl_col and col == gl_col:
                    val = _gl_label(val)
                elif isinstance(val, pd.Timestamp):
                    val = val.to_pydatetime()
                elif not isinstance(val, (str, int, float, _dt.datetime, type(None))):
                    val = str(val)
                elif isinstance(val, float):
                    if val != val:   val = None
                    elif val == int(val): val = int(val)
                is_amt = amt_col and col == amt_col
                cell = ws.cell(r, ci, value=val if val != "" else None)
                if is_amt and isinstance(val, (int, float)) and val is not None:
                    grand_total += float(val)
                    cell.font = _font(color=COL_POS if val > 0 else (COL_NEG if val < 0 else COL_BLK))
                    cell.number_format = "#,##0.00"
                    cell.alignment = _aln("right")
                elif isinstance(val, _dt.datetime):
                    cell.font = _font()
                    cell.number_format = "DD/MM/YYYY"
                    cell.alignment = _aln("left")
                else:
                    cell.font = _font()
                    cell.alignment = _aln("left")
                cell.fill = _fill(bg)
                cell.border = _thin()
            ws.row_dimensions[r].height = 13
            r += 1

    # ── Grand total row ───────────────────────────────────────────────────────
    _nb_lbl = {"en": "Net Balance", "nl": "Nettosaldo", "fr": "Solde net"}
    for ci in range(1, ncols + 1):
        cell = ws.cell(r, ci)
        cell.fill = _fill(HDR_FILL)
        cell.border = _thin()
        if ci == 1:
            cell.value = _nb_lbl.get(lang, "Net Balance")
            cell.font = _font(bold=True, color=COL_WHT, size=10)
            cell.alignment = _aln("left")
        elif ci == amt_ci:
            cell.value = grand_total
            cell.font = _font(bold=True, color=COL_WHT, size=11)
            cell.number_format = "#,##0.00"
            cell.alignment = _aln("right")
    ws.row_dimensions[r].height = 20

    out = BytesIO()
    wb.save(out); out.seek(0)
    return out



def build_overview(df: pd.DataFrame, amt_col: str,
                   year_from: int, year_to: int,
                   customer_name: str = "",
                   account_id:    str = "",
                   lang:          str = "en",
                   remove_overdues: bool = False) -> BytesIO:
    """
    Multi-year overview — preserves SAP blank-row group separators exactly.

    Structure:
      1. CURRENT OPEN ITEMS  — groups whose rows all precede the first
         historical clearing block (index boundary detected automatically).
         Plain white/light-blue alternating rows (no special background).
         Current-open total row (mid-blue).
      2. One section per calendar year, newest year first.
         Each year: dark-blue banner → mid-blue column headers → data rows
         → yellow zero-subtotal rows (payment grouping boundaries) →
         mid-blue year total.
      3. Net Balance grand total (dark blue).

    Colours match the AR Suite reference style:
      - Dark navy  #1F3864 — year banners, grand total
      - Mid blue   #2E75B6 — column headers, year totals, open total
      - White      #FFFFFF / Light grey #F2F2F2 — alternating data rows
      - Yellow     #FFFF00 — zero-netting group separator rows
      - Red        #C00000 — positive amounts (invoices)
      - Green      #375623 — negative amounts (credits / payments)
    """
    import datetime as _dt
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL  = "FF1F3864"   # dark blue  — year banner + grand total
    BAND_FILL = "FF2E75B6"   # mid blue   — column headers + year/open totals
    ROW_WHITE = "FFFFFFFF"   # plain white — data rows (incl. current open)
    ROW_BLUE  = "FFF2F2F2"   # light grey  — alternating data rows
    ROW_YELL  = "FFFFFF00"   # yellow      — zero-netting group separator
    COL_POS   = "FFC00000"   # red         — positive amounts
    COL_NEG   = "FF375623"   # green       — negative amounts
    COL_WHT   = "FFFFFFFF"
    COL_BLK   = "FF000000"

    def _fill(rgb): return PatternFill("solid", fgColor=rgb)
    def _font(bold=False, color=COL_BLK, size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def _thin():
        s = Side(style="thin", color="DDDDDD")
        return Border(left=s, right=s, top=s, bottom=s)
    def _aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=False)

    STRIP = {
        "Reason code","Clerk Abbreviation","Cleared/open items symbol",
        "Disputed item","Payment Block","Net due date symbol",
        "Text","Clearing date","Clearing Document","Dunning Level",
        "Last Dunned","Reversed with","Document Header Text","User Name",
        "Special G/L ind.","Billing Document","Reference Key 1",
        "doc_number_str","ref","sap_class","is_open","header_text",
        "clearing_date","clearing_doc","text",
        "Case ID", "Status", "Dunning Block",
    }
    display_cols = [c for c in df.columns if c not in STRIP and c is not None and str(c) != "None"]
    ncols  = len(display_cols)
    col_widths = {
        "Account":10,"Assignment":14,"Document Number":18,
        "Reference Key 3":14,"Document Date":13,"Net due date":13,
        "Document Type":26,"Amount in local currency":20,
        "Arrears after net due date":24,"Payment Method":13,
        "G/L Account":22,"Disputed item":13,
    }
    amt_ci = (display_cols.index(amt_col)+1) if amt_col and amt_col in display_cols else None
    pay_col      = next((c for c in df.columns if "payment method" in c.lower()), None)

    # G/L label mapping (language-aware)
    _gl_map_ov = GL_LABELS.get(lang, GL_LABELS["en"])
    gl_col_ov  = next((c for c in df.columns if "g/l" in c.lower() or "gl account" in c.lower()), None)

    def _gl_label_ov(raw_val):
        s = str(raw_val) if raw_val is not None else ""
        lbl = _gl_map_ov.get(s)
        return f"{s} ({lbl})" if lbl else s

    ndd_col      = next((c for c in df.columns if "net due"       in c.lower()), None)
    arr_col      = next((c for c in df.columns if "arrears"       in c.lower()), None)
    doc_type_col = next((c for c in df.columns if "document type" in c.lower()), None)
    doc_date_col = next((c for c in df.columns if c.lower() == "document date"), None)
    acc_col      = next((c for c in df.columns if c.lower() in ("account","konto","debitor")), None)

    if ndd_col:      df[ndd_col]      = pd.to_datetime(df[ndd_col],      errors="coerce")
    if doc_date_col: df[doc_date_col] = pd.to_datetime(df[doc_date_col], errors="coerce")
    if amt_col:      df[amt_col]      = pd.to_numeric( df[amt_col],      errors="coerce")
    if arr_col:      df[arr_col]      = pd.to_numeric( df[arr_col],      errors="coerce")

    # NOTE: arrears are NOT recalculated here. The SAP export already contains
    # the correct arrears values as of the export date. Recalculating against
    # today would change the numbers and cause a mismatch with the source export.
    # Only build_current_overview recalculates arrears (it's a live snapshot).
    # NOTE: remove_overdues for multi-year ONLY suppresses the Current Open
    # section — it does NOT filter rows from historical cleared years.
    # Historical year rows always show in full so the reconciliation is visible.
    # ── Parse SAP groups using blank Account rows as separators ──────────────
    # Each group = list of (original_index, row) tuples
    groups_raw, cur = [], []
    for idx, row in df.iterrows():
        acc = str(row.get(acc_col, "") or "").strip() if acc_col else ""
        if acc and acc not in ("nan", "None", ""):
            cur.append((idx, row))
        else:
            if cur:
                groups_raw.append(cur)
                cur = []
    if cur:
        groups_raw.append(cur)

    # ── Split: current-open groups vs historical ──────────────────────────────
    # Detect whether the export has a group-separator structure (blank rows in
    # the middle) or is a flat open-items export (blank rows only at the end).
    if acc_col:
        is_blank_all = df[acc_col].isna() | df[acc_col].astype(str).str.strip().isin(["", "nan", "None"])
        last_real = df[~is_blank_all].index.max() if (~is_blank_all).any() else -1
        blanks_in_middle = (is_blank_all & (df.index < last_real)).any()
    else:
        blanks_in_middle = False

    PAY_TYPES = {"DZ", "ZP"}

    if not blanks_in_middle:
        # Flat open-items export — all rows are current open, no historical groups
        current_open_groups = [[r for _, r in grp] for grp in groups_raw]
        historical_groups   = []
    else:
        # The FIRST blank separator row marks the boundary between current open
        # and historical cleared groups — same approach as build_current_overview.
        is_blank_all = df[acc_col].isna() | df[acc_col].astype(str).str.strip().isin(["", "nan", "None"]) \
                       if acc_col else pd.Series([False]*len(df), index=df.index)
        first_blank_idx = df[is_blank_all].index.min() if is_blank_all.any() else None

        if first_blank_idx is None:
            current_open_groups = [[r for _, r in grp] for grp in groups_raw]
            historical_groups   = []
        else:
            current_open_groups = []
            historical_groups   = []
            for grp in groups_raw:
                first_idx = grp[0][0]
                if first_idx < first_blank_idx:
                    current_open_groups.append([r for _, r in grp])
                else:
                    historical_groups.append([r for _, r in grp])

    # ── Bucket historical groups by year ──────────────────────────────────────
    SKIP_TYPES = {"AB", "ZP", "DZ"}

    def _group_year_local(grp):
        """Year of a group = year of the oldest NET DUE DATE among non-clearing rows.
        Falls back to doc date only if no net due dates exist.
        Falls back to clearing/payment rows' dates if all rows are clearing types."""
        ndd_dates = []
        doc_dates = []
        skip_ndd_dates = []
        skip_doc_dates = []
        for row in grp:
            dt = str(row.get(doc_type_col, "") or "").strip().upper()
            is_skip = dt in SKIP_TYPES
            if ndd_col:
                v = row.get(ndd_col)
                if v is not None and pd.notna(v):
                    try:
                        ts = pd.Timestamp(v)
                        (skip_ndd_dates if is_skip else ndd_dates).append(ts)
                    except Exception:
                        pass
            if doc_date_col:
                v = row.get(doc_date_col)
                if v is not None and pd.notna(v):
                    try:
                        ts = pd.Timestamp(v)
                        (skip_doc_dates if is_skip else doc_dates).append(ts)
                    except Exception:
                        pass
        dates = ndd_dates or doc_dates or skip_ndd_dates or skip_doc_dates
        if not dates:
            return None
        return min(dates).year

    year_groups: dict = {}
    for grp in historical_groups:
        yr = _group_year_local(grp)
        if yr is not None and year_from <= yr <= year_to:
            year_groups.setdefault(yr, []).append(grp)

    # ── Helper: write a single data row ──────────────────────────────────────
    def _write_data_row(ws, r, row, bg, amt_ci):
        for ci, col in enumerate(display_cols, 1):
            val = row.get(col, "")
            # Translate Document Type codes → human-readable description
            if doc_type_col and col == doc_type_col:
                pm  = row.get(pay_col, "") if pay_col else ""
                val = _desc(val, row.get(amt_col, 0), pm, lang)
            # Append Beer/Rent label to G/L Account column
            elif gl_col_ov and col == gl_col_ov:
                val = _gl_label_ov(val)
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            elif not isinstance(val, (str, int, float, _dt.datetime, type(None))):
                val = str(val)
            elif isinstance(val, float):
                if val != val: val = None
                elif val == int(val): val = int(val)
            is_amt = amt_col and col == amt_col
            cell = ws.cell(r, ci, value=val if val != "" else None)
            if is_amt and isinstance(val, (int, float)) and val is not None:
                cell.font = _font(color=COL_POS if val > 0 else (COL_NEG if val < 0 else COL_BLK))
                cell.number_format = "#,##0.00"
                cell.alignment = _aln("right")
            elif isinstance(val, _dt.datetime):
                cell.font = _font()
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = _aln("left")
            else:
                cell.font = _font()
                cell.alignment = _aln("left")
            cell.fill   = _fill(bg)
            cell.border = _thin()
        ws.row_dimensions[r].height = 13

    def _write_col_headers(ws, r):
        for ci, col in enumerate(display_cols, 1):
            # Rename Document Type column to translated description label
            h = _t(lang, "desc_col") if doc_type_col and col == doc_type_col else col
            cell = ws.cell(r, ci, value=h)
            cell.font      = _font(bold=True, color=COL_WHT, size=9)
            cell.fill      = _fill(BAND_FILL)
            cell.alignment = _aln("center")
            cell.border    = _thin()
        ws.row_dimensions[r].height = 15

    def _write_zero_subtotal(ws, r, amt_ci):
        for ci in range(1, ncols+1):
            cell = ws.cell(r, ci)
            cell.fill   = _fill(ROW_YELL)
            cell.border = _thin()
            if ci == amt_ci:
                cell.value         = 0
                cell.font          = _font(bold=True)
                cell.number_format = "#,##0.00"
                cell.alignment     = _aln("right")
        ws.row_dimensions[r].height = 8

    def _write_band_total(ws, r, label, total, amt_ci):
        for ci in range(1, ncols+1):
            cell = ws.cell(r, ci)
            cell.fill   = _fill(BAND_FILL)
            cell.border = _thin()
        ws.cell(r, 1).value     = label
        ws.cell(r, 1).font      = _font(bold=True, color=COL_WHT, size=10)
        ws.cell(r, 1).alignment = _aln("left")
        if amt_ci:
            c = ws.cell(r, amt_ci)
            c.value         = total
            c.font          = _font(bold=True, color=COL_WHT, size=10)
            c.number_format = "#,##0.00"
            c.alignment     = _aln("right")
        ws.row_dimensions[r].height = 18

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overview"
    for ci, col in enumerate(display_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, max(len(str(col))+2, 12))

    r = 1
    row_idx = 0
    grand_total = 0.0

    # ── Row 1: Title (dark blue, size 14, bold, centred, MERGED) ────────────
    acc_display = account_id or (
        str(df[acc_col].dropna().iloc[0]).strip().split(".")[0]
        if acc_col and len(df[acc_col].dropna()) > 0 else ""
    )
    yr_range = f"{year_from}–{year_to}" if year_from != year_to else str(year_from)
    title_val = f"Account {acc_display}  ·  {yr_range}  {_t(lang, 'title_suffix')}"
    # Fill all cells dark blue first
    for ci in range(1, ncols+1):
        ws.cell(r, ci).fill   = _fill(HDR_FILL)
        ws.cell(r, ci).border = _thin()
    # Merge across all columns so title is fully visible
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.cell(r, 1).value     = title_val
    ws.cell(r, 1).font      = _font(bold=True, color=COL_WHT, size=14)
    ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[r].height = 34
    r += 1

    # ── Row 2: Subtitle (mid blue, size 9, not bold, MERGED) ─────────────────
    today_str = _dt.date.today().strftime("%d/%m/%Y")
    subtitle_val = f"{_t(lang, 'subtitle')}  ·  {today_str}"
    for ci in range(1, ncols+1):
        ws.cell(r, ci).fill   = _fill(BAND_FILL)
        ws.cell(r, ci).border = _thin()
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.cell(r, 1).value     = subtitle_val
    ws.cell(r, 1).font      = _font(bold=False, color=COL_WHT, size=9)
    ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[r].height = 16
    r += 1

    # ── Row 3: blank gap ─────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 6
    r += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1 — CURRENT OPEN ITEMS
    # When remove_overdues is True, skip this section entirely — the open items
    # are by definition the outstanding/overdue items the user wants excluded.
    # ══════════════════════════════════════════════════════════════════════════
    if current_open_groups and not remove_overdues:
        # Dark-blue banner — MERGED across all columns
        for ci in range(1, ncols+1):
            ws.cell(r, ci).fill   = _fill(HDR_FILL)
            ws.cell(r, ci).border = _thin()
        _open_labels = {"en": "Current Open Items",
                        "nl": "Huidige Openstaande Posten",
                        "fr": "Postes Ouverts Actuels"}
        lbl = _open_labels.get(lang, "Current Open Items")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        ws.cell(r, 1).value     = lbl
        ws.cell(r, 1).font      = _font(bold=True, color=COL_WHT, size=12)
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.row_dimensions[r].height = 22
        r += 1

        # Column headers
        _write_col_headers(ws, r); r += 1

        open_total = 0.0
        for grp in current_open_groups:
            grp_total = sum(
                float(row.get(amt_col, 0) or 0)
                for row in grp if amt_col and pd.notna(row.get(amt_col))
            )
            open_total += grp_total
            for row in grp:
                bg = ROW_WHITE if row_idx % 2 == 0 else ROW_BLUE
                _write_data_row(ws, r, row, bg, amt_ci)
                r += 1; row_idx += 1
            if abs(grp_total) < 0.02:
                _write_zero_subtotal(ws, r, amt_ci)
                r += 1; row_idx += 1

        grand_total += open_total
        _open_total_lbl = {"en": "Current Open — Total",
                           "nl": "Huidige Openstaand — Totaal",
                           "fr": "Postes Ouverts — Total"}
        _write_band_total(ws, r, _open_total_lbl.get(lang, "Current Open — Total"), open_total, amt_ci)
        r += 1
        # Gap
        ws.row_dimensions[r].height = 8; r += 1

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2+ — ONE SECTION PER YEAR (newest first)
    # ══════════════════════════════════════════════════════════════════════════
    for yr in sorted(year_groups.keys(), reverse=True):
        groups = year_groups[yr]
        if not groups:
            continue

        yr_total = sum(
            float(row.get(amt_col, 0) or 0)
            for grp in groups for row in grp
            if amt_col and pd.notna(row.get(amt_col))
        )
        grand_total += yr_total

        # Dark-blue year banner — MERGED across all columns
        yr_inv  = sum(float(row.get(amt_col,0) or 0)
                      for grp in groups for row in grp
                      if amt_col and pd.notna(row.get(amt_col)) and float(row.get(amt_col,0) or 0) > 0)
        yr_cred = sum(float(row.get(amt_col,0) or 0)
                      for grp in groups for row in grp
                      if amt_col and pd.notna(row.get(amt_col)) and float(row.get(amt_col,0) or 0) < 0)
        banner_val = _t(lang, "year_banner",
                        yr=yr, n=len(groups),
                        inv=f"{yr_inv:,.2f}",
                        cred=f"{yr_cred:,.2f}",
                        net=f"{yr_total:,.2f}")
        for ci in range(1, ncols+1):
            ws.cell(r, ci).fill   = _fill(HDR_FILL)
            ws.cell(r, ci).border = _thin()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        ws.cell(r, 1).value     = banner_val
        ws.cell(r, 1).font      = _font(bold=True, color=COL_WHT, size=11)
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.row_dimensions[r].height = 22
        r += 1

        # Mid-blue column headers
        _write_col_headers(ws, r); r += 1

        # Data rows + yellow zero-subtotals
        row_idx = 0
        for grp in groups:
            grp_total = sum(
                float(row.get(amt_col, 0) or 0)
                for row in grp if amt_col and pd.notna(row.get(amt_col))
            )
            for row in grp:
                bg = ROW_WHITE if row_idx % 2 == 0 else ROW_BLUE
                _write_data_row(ws, r, row, bg, amt_ci)
                r += 1; row_idx += 1
            # Yellow separator row for every group that nets to zero
            if abs(grp_total) < 0.02:
                _write_zero_subtotal(ws, r, amt_ci)
                r += 1; row_idx += 1

        # Mid-blue year total
        yr_label = _t(lang, "year_total", yr=yr)
        _write_band_total(ws, r, yr_label, yr_total, amt_ci)
        r += 1

        # Small gap before next year
        ws.row_dimensions[r].height = 8; r += 1

    # ══════════════════════════════════════════════════════════════════════════
    # GRAND TOTAL — dark blue
    # ══════════════════════════════════════════════════════════════════════════
    for ci in range(1, ncols+1):
        cell = ws.cell(r, ci)
        cell.fill   = _fill(HDR_FILL)
        cell.border = _thin()
    ws.cell(r, 1).value     = _t(lang, "net_balance")
    ws.cell(r, 1).font      = _font(bold=True, color=COL_WHT, size=11)
    ws.cell(r, 1).alignment = _aln("left")
    if amt_ci:
        ws.cell(r, amt_ci).value         = grand_total
        ws.cell(r, amt_ci).font          = _font(bold=True, color=COL_WHT, size=12)
        ws.cell(r, amt_ci).number_format = "#,##0.00"
        ws.cell(r, amt_ci).alignment     = _aln("right")
    ws.row_dimensions[r].height = 22
    ws.freeze_panes = "A4"

    out = BytesIO()
    wb.save(out); out.seek(0)
    return out
