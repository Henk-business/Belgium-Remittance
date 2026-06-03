"""Remittance reconciliation engine. SAP is the source of truth."""
import pandas as pd
import openpyxl
import datetime
from io import BytesIO
import warnings
warnings.filterwarnings("ignore")

from common import (
    BG, FG, c, mr, col_w, hdr_row, fd, parse_sap
)


def _extract_pdf_cells(file_obj):
    """Extract text tokens from a PDF remittance, returning them as a list of rows."""
    import io
    data = file_obj.read() if hasattr(file_obj, "read") else file_obj
    rows = []
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                # Table extraction first (structured remittances)
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if row:
                                rows.append([str(c or "").strip() for c in row])
                else:
                    # Fall back to raw text lines
                    text = page.extract_text() or ""
                    for line in text.splitlines():
                        rows.append(line.split())
    except Exception:
        # pdfplumber failed — try pypdf as last resort
        try:
            import pypdf
            reader = pypdf.PdfReader(io.BytesIO(data))
            for page in reader.pages:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    rows.append(line.split())
        except Exception:
            pass
    return rows


def _parse_remittance(file_obj, sap_df):
    import io

    # Detect file type from name or content
    name = getattr(file_obj, "name", "") or ""
    is_pdf = name.lower().endswith(".pdf")

    # Peek at magic bytes if no name
    if not is_pdf and hasattr(file_obj, "read"):
        header = file_obj.read(5)
        file_obj.seek(0)
        is_pdf = header.startswith(b"%PDF")

    sap_refs = set(sap_df["ref"].unique()) - {"", "nan", "None"}
    sap_docs = set(sap_df["doc_number_str"].unique()) - {"", "nan", "None", "0"}

    found = {}

    if is_pdf:
        rows = _extract_pdf_cells(file_obj)
        for row in rows:
            for cell in row:
                cell = str(cell).strip()
                if not cell or cell.lower() in ("nan", "none", ""):
                    continue
                key = None
                if cell in sap_refs:
                    key = cell
                elif cell in sap_docs:
                    key = cell
                else:
                    for ref in sap_refs:
                        if len(str(ref)) >= 6 and str(ref) in cell:
                            key = ref
                            break
                if key and key not in found:
                    ctx = " | ".join(c for c in row if c.strip())
                    found[key] = {"sap_ref": key, "context": ctx}
    else:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        raw = pd.read_excel(file_obj, sheet_name=0, header=None, dtype=str)
        raw = raw.fillna("")
        for row_idx, row in raw.iterrows():
            for col_idx, cell_val in row.items():
                cell = str(cell_val).strip()
                if not cell or cell.lower() in ("nan", "none", ""):
                    continue
                key = None
                if cell in sap_refs:
                    key = cell
                elif cell in sap_docs:
                    key = cell
                else:
                    for ref in sap_refs:
                        if len(str(ref)) >= 6 and str(ref) in cell:
                            key = ref
                            break
                if key and key not in found:
                    ctx = " | ".join(
                        str(v).strip() for v in row.values
                        if str(v).strip() and str(v).strip().lower() not in ("nan", "none", "")
                    )
                    found[key] = {"sap_ref": key, "context": ctx}

    return list(found.values())


def run_reconciliation(sap_file, rem_file, payment_amount=None, customer_name=""):
    sap = parse_sap(sap_file)
    matches = _parse_remittance(rem_file, sap)

    rv_ru = sap[sap.get("doc_type", pd.Series(dtype=str)).str.upper().isin(["RV", "RU"])]
    rv_ru_open = rv_ru[rv_ru["is_open"]]
    rv_ru_cleared = rv_ru[~rv_ru["is_open"]]

    def lkp(df_in, key="ref"):
        d = {}
        for _, row in df_in.iterrows():
            k = row[key]
            if k and k != "nan":
                d.setdefault(k, []).append(row)
        return d

    open_ref = lkp(rv_ru_open, "ref")
    open_doc = lkp(rv_ru_open, "doc_number_str")
    clr_ref  = lkp(rv_ru_cleared, "ref")
    clr_doc  = lkp(rv_ru_cleared, "doc_number_str")

    matched_inv = []
    matched_cred = []
    already_cleared = []
    not_found = []
    matched_refs = set()

    for item in matches:
        ref = item["sap_ref"]
        rows = open_ref.get(ref) or open_doc.get(ref)
        if rows:
            matched_refs.add(ref)
            net = sum(r["amount"] for r in rows)
            entry = {
                **item,
                "sap_amount":   net,
                "sap_class":    "INVOICE" if net > 0 else "CREDIT_NOTE",
                "sap_doc_type": rows[0]["doc_type"],
                "sap_due_date": rows[0]["due_date"],
                "sap_doc_date": rows[0]["doc_date"],
                "sap_header":   str(rows[0].get("header_text", ""))
                                if pd.notna(rows[0].get("header_text")) else "",
            }
            (matched_inv if net > 0 else matched_cred).append(entry)
            continue
        cleared = clr_ref.get(ref) or clr_doc.get(ref)
        if cleared:
            matched_refs.add(ref)
            already_cleared.append({
                **item,
                "sap_amount":   sum(r["amount"] for r in cleared),
                "sap_class":    cleared[0]["sap_class"],
                "cleared_by":   str(cleared[0].get("clearing_doc", "")),
                "cleared_date": cleared[0].get("clearing_date"),
            })
            continue
        not_found.append(item)

    open_inv_all  = rv_ru_open[rv_ru_open["sap_class"] == "INVOICE"].copy()
    open_cred_all = rv_ru_open[rv_ru_open["sap_class"] == "CREDIT_NOTE"].copy()
    missing = open_inv_all[
        ~open_inv_all["ref"].isin(matched_refs) &
        ~open_inv_all["doc_number_str"].isin(matched_refs)
    ].copy()

    return {
        "matched_inv":      matched_inv,
        "matched_cred":     matched_cred,
        "already_cleared":  already_cleared,
        "not_found":        not_found,
        "missing":          missing,
        "open_inv_all":     open_inv_all,
        "open_cred_all":    open_cred_all,
        "payment_amount":   payment_amount,
        "customer_name":    customer_name,
        "t_inv":    sum(i["sap_amount"] for i in matched_inv),
        "t_cred":   sum(i["sap_amount"] for i in matched_cred),
        "t_missing": missing["amount"].sum() if len(missing) else 0,
        "t_open_cr": open_cred_all["amount"].sum() if len(open_cred_all) else 0,
    }


def build_recon_report(results):
    wb = openpyxl.Workbook()
    mi  = results["matched_inv"]
    mc  = results["matched_cred"]
    ac  = results["already_cleared"]
    nf  = results["not_found"]
    mfr = results["missing"]
    pmt = results["payment_amount"]
    cname = results["customer_name"] or "Customer"

    ws = wb.active
    ws.title = "Summary"
    col_w(ws, [4, 46, 4, 18, 4, 4])
    r = 1
    title = "REMITTANCE RECONCILIATION" + (f"  —  {cname}" if cname else "")
    if pmt:
        title += f"  ·  Payment €{pmt:,.2f}"
    mr(ws, r, 1, 6, title, bold=True, bg="dk_blue", fg="white", sz=13, ha="center")
    ws.row_dimensions[r].height = 34
    r += 1
    mr(ws, r, 1, 6, "SAP is the source of truth. Client signs and labels are ignored.",
       bg="md_blue", fg="white", sz=9, ha="center", italic=True)
    ws.row_dimensions[r].height = 16
    r += 2

    mr(ws, r, 1, 6, "RESULTS", bold=True, bg="dk_blue", fg="white", sz=11, ha="center")
    ws.row_dimensions[r].height = 22
    r += 1

    for desc, amt, bg, fg_ in [
        (f"Invoices matched — open in SAP  ({len(mi)} items)", results["t_inv"], "lt_green", "md_green"),
        (f"Credit notes matched  ({len(mc)} items)", results["t_cred"], "lt_green", "md_green"),
        ("", None, "white", "black"),
        (f"Already cleared — check for doubles  ({len(ac)} items)", None, "lt_red", "md_red"),
        (f"Not found in SAP  ({len(nf)} items)", None, "pink", "md_red"),
        ("", None, "white", "black"),
        (f"Open SAP invoices not on remittance  ({len(mfr)} items)", results["t_missing"], "lt_blue", "black"),
    ]:
        if not desc:
            mr(ws, r, 1, 6, None, bg="white")
            ws.row_dimensions[r].height = 6
            r += 1
            continue
        mr(ws, r, 1, 4, desc, bg=bg, fg=fg_, sz=10)
        c(ws, r, 5, amt, bg=bg, fg=fg_, fmt="#,##0.00", ha="right", sz=10)
        c(ws, r, 6, None, bg=bg)
        ws.row_dimensions[r].height = 20
        r += 1

    def _sheet(tab, title, subtitle, items, hdr_bg):
        ws2 = wb.create_sheet(tab)
        col_w(ws2, [4, 24, 36, 13, 13, 16, 4])
        r2 = 1
        mr(ws2, r2, 1, 7, title, bold=True, bg=hdr_bg, fg="white", sz=10)
        ws2.row_dimensions[r2].height = 22
        r2 += 1
        mr(ws2, r2, 1, 7, subtitle, bg="grey", fg="black", sz=9, italic=True, wrap=True)
        ws2.row_dimensions[r2].height = 16
        r2 += 1
        hdr_row(ws2, r2, ["#", "SAP Reference", "Context", "Invoice Date", "Due Date", "Amount (€)", ""])
        r2 += 1
        total = 0.0
        for idx, item in enumerate(items, 1):
            bg = "lt_green" if idx % 2 == 0 else "white"
            c(ws2, r2, 1, idx, bg=bg, sz=8, ha="center")
            c(ws2, r2, 2, item["sap_ref"], bg=bg, sz=9, bold=True)
            c(ws2, r2, 3, item.get("context", ""), bg=bg, sz=8)
            c(ws2, r2, 4, fd(item.get("sap_doc_date")), bg=bg, sz=9, ha="center")
            c(ws2, r2, 5, fd(item.get("sap_due_date")), bg=bg, sz=9, ha="center")
            c(ws2, r2, 6, item.get("sap_amount", 0), bg=bg, fmt="#,##0.00", ha="right", sz=9)
            c(ws2, r2, 7, None, bg=bg)
            total += item.get("sap_amount", 0) or 0
            ws2.row_dimensions[r2].height = 13
            r2 += 1
        mr(ws2, r2, 1, 5, "TOTAL", bold=True, bg=hdr_bg, fg="white", sz=10)
        c(ws2, r2, 6, total, bold=True, bg=hdr_bg, fg="white", fmt="#,##0.00", ha="right", sz=10)
        c(ws2, r2, 7, None, bg=hdr_bg)
        ws2.row_dimensions[r2].height = 16

    _sheet("Matched Invoices",
           f"INVOICES MATCHED — Open in SAP  ({len(mi)}  ·  €{results['t_inv']:,.2f})",
           "Open RV invoices found on remittance. SAP classification used.", mi, "md_green")
    _sheet("Matched Credits",
           f"CREDIT NOTES MATCHED  ({len(mc)})",
           "SAP classifies these as credit notes (negative RV or RU).", mc, "md_green")

    ws4 = wb.create_sheet("Already Cleared")
    col_w(ws4, [4, 24, 16, 14, 22, 4])
    r4 = 1
    mr(ws4, r4, 1, 6, f"ALREADY CLEARED — Potential Doubles  ({len(ac)} items)",
       bold=True, bg="md_red", fg="white", sz=10)
    ws4.row_dimensions[r4].height = 22
    r4 += 1
    mr(ws4, r4, 1, 6, "On remittance but already cleared in SAP. Verify before processing.",
       bg="lt_red", fg="black", sz=9, italic=True)
    ws4.row_dimensions[r4].height = 16
    r4 += 1
    hdr_row(ws4, r4, ["#", "SAP Reference", "SAP Class", "Cleared Date", "Clearing Doc", ""])
    r4 += 1
    for idx, item in enumerate(ac, 1):
        c(ws4, r4, 1, idx, bg="lt_red", sz=8, ha="center")
        c(ws4, r4, 2, item["sap_ref"], bg="lt_red", sz=9, bold=True)
        c(ws4, r4, 3, item.get("sap_class", ""), bg="lt_red", sz=9, ha="center")
        c(ws4, r4, 4, fd(item.get("cleared_date")), bg="lt_red", sz=9, ha="center")
        c(ws4, r4, 5, str(item.get("cleared_by", "")), bg="lt_red", sz=9)
        c(ws4, r4, 6, None, bg="lt_red")
        ws4.row_dimensions[r4].height = 14
        r4 += 1

    ws5 = wb.create_sheet("Not Found")
    col_w(ws5, [4, 24, 42, 4])
    r5 = 1
    mr(ws5, r5, 1, 4, f"NOT FOUND IN SAP  ({len(nf)} items)",
       bold=True, bg="purple", fg="white", sz=10)
    ws5.row_dimensions[r5].height = 22
    r5 += 1
    mr(ws5, r5, 1, 4, "On remittance but not found in SAP as any RV/RU document.",
       bg="lt_purple", fg="black", sz=9, italic=True)
    ws5.row_dimensions[r5].height = 16
    r5 += 1
    hdr_row(ws5, r5, ["#", "Value from Remittance", "Context", ""])
    r5 += 1
    for idx, item in enumerate(nf, 1):
        bg = "lt_purple" if idx % 2 == 0 else "white"
        c(ws5, r5, 1, idx, bg=bg, sz=8, ha="center")
        c(ws5, r5, 2, item["sap_ref"], bg=bg, sz=9, bold=True)
        c(ws5, r5, 3, item.get("context", ""), bg=bg, sz=8)
        c(ws5, r5, 4, None, bg=bg)
        ws5.row_dimensions[r5].height = 13
        r5 += 1

    if len(mfr) > 0:
        ws6 = wb.create_sheet("SAP Open Not on Remittance")
        col_w(ws6, [4, 24, 32, 13, 13, 15])
        r6 = 1
        mr(ws6, r6, 1, 6,
           f"OPEN IN SAP — NOT ON REMITTANCE  ({len(mfr)} items  ·  €{mfr['amount'].sum():,.2f})",
           bold=True, bg="md_blue", fg="white", sz=10)
        ws6.row_dimensions[r6].height = 22
        r6 += 1
        mr(ws6, r6, 1, 6, "Open SAP invoices not mentioned on the remittance.",
           bg="lt_blue", fg="black", sz=9, italic=True)
        ws6.row_dimensions[r6].height = 16
        r6 += 1
        hdr_row(ws6, r6, ["#", "SAP Reference", "Description", "Invoice Date", "Due Date", "Amount (€)"])
        r6 += 1
        for idx, (_, row) in enumerate(mfr.sort_values("due_date").iterrows(), 1):
            bg = "lt_blue" if idx % 2 == 0 else "white"
            c(ws6, r6, 1, idx, bg=bg, sz=8, ha="center")
            c(ws6, r6, 2, row["ref"], bg=bg, sz=9)
            ht = str(row.get("header_text", "")) if pd.notna(row.get("header_text")) else ""
            c(ws6, r6, 3, ht, bg=bg, sz=9)
            c(ws6, r6, 4, fd(row.get("doc_date")), bg=bg, sz=9, ha="center")
            c(ws6, r6, 5, fd(row.get("due_date")), bg=bg, sz=9, ha="center")
            c(ws6, r6, 6, row["amount"], bg=bg, fmt="#,##0.00", ha="right", sz=9)
            ws6.row_dimensions[r6].height = 13
            r6 += 1
        mr(ws6, r6, 1, 5, "TOTAL", bold=True, bg="md_blue", fg="white", sz=10)
        c(ws6, r6, 6, mfr["amount"].sum(), bold=True, bg="md_blue",
          fg="white", fmt="#,##0.00", ha="right", sz=10)
        ws6.row_dimensions[r6].height = 16

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_statement(results, today=None):
    """
    Customer statement: Section A (cleared, green), B (open due, red),
    NET DUE (yellow), Section C (due later, blue).
    """
    if today is None:
        today = datetime.date.today()
    today_ts  = pd.Timestamp(today)
    today_str = today_ts.strftime("%d/%m/%Y")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Account Statement"
    col_w(ws, [5, 22, 30, 13, 13, 16, 16, 5])

    cname = results["customer_name"] or "Customer"
    pmt   = results["payment_amount"]

    r = 1
    mr(ws, r, 1, 8, cname.upper() + "  \u00b7  ACCOUNT STATEMENT",
       bold=True, bg="dk_blue", fg="white", sz=15, ha="center")
    ws.row_dimensions[r].height = 38
    r += 1
    subtitle = "As at " + today_str
    if pmt:
        subtitle += "  \u00b7  After remittance payment of \u20ac" + f"{pmt:,.2f}"
    mr(ws, r, 1, 8, subtitle, bg="md_blue", fg="white", sz=9, ha="center", italic=True)
    ws.row_dimensions[r].height = 18
    r += 2

    cleared_items = results["matched_inv"] + results["matched_cred"]
    open_inv_all  = results["open_inv_all"]
    open_cred_all = results["open_cred_all"]
    cleared_total = sum(i["sap_amount"] for i in cleared_items)

    if "due_date" in open_inv_all.columns:
        open_due   = open_inv_all[open_inv_all["due_date"] <= today_ts].copy()
        open_later = open_inv_all[open_inv_all["due_date"] >  today_ts].copy()
    else:
        open_due   = open_inv_all.copy()
        open_later = pd.DataFrame()

    credits_total = open_cred_all["amount"].sum() if len(open_cred_all) else 0.0
    net_due = open_due["amount"].sum() + credits_total

    mr(ws, r, 1, 8, "ACCOUNT SUMMARY", bold=True, bg="dk_blue", fg="white", sz=11, ha="center")
    ws.row_dimensions[r].height = 22
    r += 1

    for desc, amt, bg, fg_, bold_ in [
        (f"Cleared by this payment  ({len(cleared_items)} items)", cleared_total, "lt_green", "md_green", False),
        ("", None, "white", "black", False),
        (f"Open invoices due by {today_str}  ({len(open_due)} items)",
         open_due["amount"].sum() if len(open_due) else 0, "lt_red", "md_red", False),
        ("Open credit notes available to offset", credits_total, "lt_green", "md_green", False),
        ("NET AMOUNT DUE", net_due, "yellow", "md_red", True),
        ("", None, "white", "black", False),
        (f"Due after {today_str}  ({len(open_later)} items) — for information only",
         open_later["amount"].sum() if len(open_later) else 0, "grey", "grey", False),
    ]:
        if not desc:
            mr(ws, r, 1, 8, None, bg="white")
            ws.row_dimensions[r].height = 6
            r += 1
            continue
        mr(ws, r, 1, 6, desc, bold=bold_, bg=bg, fg=fg_, sz=10)
        c(ws, r, 7, amt, bold=bold_, bg=bg, fg=fg_, fmt="#,##0.00", ha="right", sz=10)
        c(ws, r, 8, None, bg=bg)
        ws.row_dimensions[r].height = 24 if bold_ else 18
        r += 1
    r += 1

    def _inv_hdr(ws_, row_):
        hdr_row(ws_, row_, ["#", "SAP Reference", "Description",
                            "Invoice Date", "Due Date", "Amount (\u20ac)", "Status", ""])

    # Section A
    mr(ws, r, 1, 8,
       f"A.  CLEARED BY THIS PAYMENT  ({len(cleared_items)} items  \u00b7  \u20ac{cleared_total:,.2f})",
       bold=True, bg="md_green", fg="white", sz=10)
    ws.row_dimensions[r].height = 22
    r += 1
    mr(ws, r, 1, 8, "Fully settled by the remittance payment — no further action required.",
       bg="lt_green", fg="md_green", sz=9, italic=True)
    ws.row_dimensions[r].height = 16
    r += 1
    _inv_hdr(ws, r)
    r += 1
    for idx, item in enumerate(cleared_items, 1):
        bg = "lt_green" if idx % 2 == 0 else "white"
        c(ws, r, 1, idx, bg=bg, sz=8, ha="center")
        c(ws, r, 2, item["sap_ref"], bg=bg, sz=9)
        c(ws, r, 3, item.get("sap_header", ""), bg=bg, sz=9)
        c(ws, r, 4, fd(item.get("sap_doc_date")), bg=bg, sz=9, ha="center")
        c(ws, r, 5, fd(item.get("sap_due_date")), bg=bg, sz=9, ha="center")
        c(ws, r, 6, item.get("sap_amount", 0), bg=bg, fmt="#,##0.00", ha="right", sz=9)
        c(ws, r, 7, "Cleared \u2713", bg=bg, sz=8, ha="center", fg="md_green")
        c(ws, r, 8, None, bg=bg)
        ws.row_dimensions[r].height = 13
        r += 1
    mr(ws, r, 1, 5, "TOTAL CLEARED", bold=True, bg="md_green", fg="white", sz=10)
    c(ws, r, 6, cleared_total, bold=True, bg="md_green", fg="white", fmt="#,##0.00", ha="right", sz=10)
    c(ws, r, 7, None, bg="md_green")
    c(ws, r, 8, None, bg="md_green")
    ws.row_dimensions[r].height = 16
    r += 2

    # Section B
    gross_due = open_due["amount"].sum() if len(open_due) else 0
    mr(ws, r, 1, 8,
       f"B.  OPEN INVOICES DUE BY {today_str}  ({len(open_due)} items  \u00b7  gross \u20ac{gross_due:,.2f}  \u00b7  NET \u20ac{net_due:,.2f})",
       bold=True, bg="md_red", fg="white", sz=10)
    ws.row_dimensions[r].height = 22
    r += 1
    mr(ws, r, 1, 8,
       "These invoices must be paid. Pink rows are already overdue. Net is after applying available credits.",
       bg="lt_red", fg="dk_red", sz=9, italic=True, wrap=True)
    ws.row_dimensions[r].height = 16
    r += 1
    _inv_hdr(ws, r)
    r += 1
    for idx, (_, row_) in enumerate(open_due.sort_values("due_date").iterrows(), 1):
        overdue = pd.notna(row_.get("due_date")) and row_["due_date"] < today_ts
        bg      = "pink" if overdue else ("lt_red" if idx % 2 == 0 else "white")
        status  = "\u26a0 OVERDUE" if overdue else ("Due " + fd(row_.get("due_date")))
        c(ws, r, 1, idx, bg=bg, sz=8, ha="center")
        c(ws, r, 2, row_["ref"], bg=bg, sz=9, bold=overdue)
        ht = str(row_.get("header_text", "")) if pd.notna(row_.get("header_text")) else ""
        c(ws, r, 3, ht, bg=bg, sz=9)
        c(ws, r, 4, fd(row_.get("doc_date")), bg=bg, sz=9, ha="center")
        c(ws, r, 5, fd(row_.get("due_date")), bg=bg, sz=9, ha="center")
        c(ws, r, 6, row_["amount"], bg=bg, fmt="#,##0.00", ha="right", sz=9,
          bold=overdue, fg="md_red" if overdue else "black")
        c(ws, r, 7, status, bg=bg, sz=8, ha="center", bold=overdue,
          fg="md_red" if overdue else "black")
        c(ws, r, 8, None, bg=bg)
        ws.row_dimensions[r].height = 13
        r += 1

    mr(ws, r, 1, 5, "GROSS TOTAL DUE", bold=True, bg="md_red", fg="white", sz=10)
    c(ws, r, 6, gross_due, bold=True, bg="md_red", fg="white", fmt="#,##0.00", ha="right", sz=10)
    c(ws, r, 7, None, bg="md_red")
    c(ws, r, 8, None, bg="md_red")
    ws.row_dimensions[r].height = 16
    r += 1

    if credits_total:
        mr(ws, r, 1, 5, "  Less: credit notes available", bold=True, bg="lt_green", fg="md_green", sz=10)
        c(ws, r, 6, credits_total, bold=True, bg="lt_green", fg="md_green", fmt="#,##0.00", ha="right", sz=10)
        c(ws, r, 7, None, bg="lt_green")
        c(ws, r, 8, None, bg="lt_green")
        ws.row_dimensions[r].height = 16
        r += 1

    mr(ws, r, 1, 5, "NET AMOUNT DUE", bold=True, bg="yellow", fg="md_red", sz=13, ha="center")
    c(ws, r, 6, net_due, bold=True, bg="yellow", fg="md_red", fmt="#,##0.00", ha="right", sz=14)
    c(ws, r, 7, None, bg="yellow")
    c(ws, r, 8, None, bg="yellow")
    ws.row_dimensions[r].height = 30
    r += 2

    # Section C
    if len(open_later) > 0:
        later_total = open_later["amount"].sum()
        mr(ws, r, 1, 8,
           f"C.  DUE AFTER {today_str}  ({len(open_later)} items  \u00b7  \u20ac{later_total:,.2f})  \u2014  For information only",
           bold=True, bg="md_blue", fg="white", sz=10)
        ws.row_dimensions[r].height = 22
        r += 1
        _inv_hdr(ws, r)
        r += 1
        for idx, (_, row_) in enumerate(open_later.sort_values("due_date").iterrows(), 1):
            bg = "lt_blue" if idx % 2 == 0 else "white"
            c(ws, r, 1, idx, bg=bg, sz=8, ha="center")
            c(ws, r, 2, row_["ref"], bg=bg, sz=9)
            ht = str(row_.get("header_text", "")) if pd.notna(row_.get("header_text")) else ""
            c(ws, r, 3, ht, bg=bg, sz=9)
            c(ws, r, 4, fd(row_.get("doc_date")), bg=bg, sz=9, ha="center")
            c(ws, r, 5, fd(row_.get("due_date")), bg=bg, sz=9, ha="center")
            c(ws, r, 6, row_["amount"], bg=bg, fmt="#,##0.00", ha="right", sz=9)
            c(ws, r, 7, "Due " + fd(row_.get("due_date")), bg=bg, sz=8, ha="center")
            c(ws, r, 8, None, bg=bg)
            ws.row_dimensions[r].height = 13
            r += 1
        mr(ws, r, 1, 5, "TOTAL DUE AFTER TODAY", bold=True, bg="md_blue", fg="white", sz=10)
        c(ws, r, 6, later_total, bold=True, bg="md_blue",
          fg="white", fmt="#,##0.00", ha="right", sz=10)
        c(ws, r, 7, None, bg="md_blue")
        c(ws, r, 8, None, bg="md_blue")
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A5"
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ── AMOUNT-ONLY MATCHING ──────────────────────────────────────────────────────

def find_amount_combinations(sap_file, payment_amount: float,
                              tolerance: float = 0.05,
                              max_results: int = 5):
    """
    Given a payment amount with no remittance, find combinations of open SAP
    invoices whose amounts sum to the payment (within tolerance).

    Strategy:
    1. Exact single-invoice match first.
    2. Two-invoice combinations.
    3. Greedy subset-sum for larger combinations.
    4. All returned sorted by confidence (closeness to payment amount).

    Returns list of dicts:
        {invoices: [rows], total: float, diff: float, confidence: str}
    """
    sap = parse_sap(sap_file)

    # ── Quick check: full settlement? ────────────────────────────────────────
    # If all open items sum to the payment, report immediately
    all_open_check = sap[sap["is_open"] & sap["amount"].notna() & (sap["amount"] != 0)].copy()
    full_total = round(all_open_check["amount"].sum(), 2)
    if abs(full_total - round(float(payment_amount), 2)) <= tolerance:
        n_inv  = (all_open_check["amount"] > 0).sum()
        n_cred = (all_open_check["amount"] < 0).sum()
        label  = f"{n_inv} invoice(s)"
        if n_cred: label += f" + {n_cred} credit note(s)"
        return [{
            "invoices":   all_open_check.to_dict("records"),
            "total":      full_total,
            "diff":       abs(full_total - round(float(payment_amount), 2)),
            "confidence": "Exact — full settlement of all open items",
            "n":          len(all_open_check),
            "label":      label + " (full settlement)",
        }], sap

    # ── Check settlement by due date ─────────────────────────────────────────
    # NOTE: due-date cutoff strategies removed — the amount-only matching tool
    # must consider ALL open items regardless of due date. The user does not
    # want future-dated invoices excluded from the search.

    # Use ALL open items — invoices (positive) AND credit notes (negative)
    # A payment might be: several invoices minus some credit notes = payment amount
    all_open = sap[
        sap["is_open"] &
        sap["amount"].notna() &
        (sap["amount"] != 0)
    ].copy()

    # ── Sort by oldest due date first ─────────────────────────────────────────
    # Strategy: match oldest invoices first, progressively moving to newer ones.
    # This mirrors real-world payment behaviour and gives the most logical output.
    all_open_dated = all_open.copy()
    if "due_date" in all_open_dated.columns:
        all_open_dated["_sort_due"] = pd.to_datetime(
            all_open_dated["due_date"], errors="coerce"
        ).fillna(pd.Timestamp("2099-12-31"))
    else:
        all_open_dated["_sort_due"] = pd.Timestamp("2099-12-31")

    # Invoices: oldest due date first; credits: oldest first too
    inv_sorted  = all_open_dated[all_open_dated["amount"] > 0].sort_values("_sort_due")
    cred_sorted = all_open_dated[all_open_dated["amount"] < 0].sort_values("_sort_due")

    inv_idx  = inv_sorted.index.tolist()
    cred_idx = cred_sorted.index.tolist()

    amounts  = all_open["amount"].round(2).tolist()
    idx_list = all_open.index.tolist()
    target   = round(float(payment_amount), 2)
    results  = []
    seen     = set()

    def _add(indices, rows, total):
        key = frozenset(indices)
        if key in seen:
            return
        seen.add(key)
        diff = abs(total - target)
        if diff <= tolerance + 0.01:
            n_inv  = sum(1 for i in indices if i in set(inv_idx))
            n_cred = sum(1 for i in indices if i in set(cred_idx))
            label  = f"{n_inv} invoice(s)"
            if n_cred:
                label += f" + {n_cred} credit note(s)"
            results.append({
                "invoices":   rows.to_dict("records"),
                "total":      round(total, 2),
                "diff":       round(diff, 2),
                "confidence": "Exact" if diff < 0.01 else f"±€{diff:.2f}",
                "n":          len(indices),
                "label":      label,
            })

    amt_map = {idx: round(all_open.at[idx, "amount"], 2) for idx in idx_list}

    # 1. Exact single match (invoice or credit)
    for idx in idx_list:
        amt = amt_map[idx]
        if abs(amt - target) <= tolerance:
            _add([idx], all_open.loc[[idx]], amt)

    # 2. Two-item combinations — invoices only, then invoice+credit
    from itertools import combinations
    pair_limit = min(len(inv_idx), 200)
    # invoice + invoice
    for i, j in combinations(inv_idx[:pair_limit], 2):
        total = round(amt_map[i] + amt_map[j], 2)
        if abs(total - target) <= tolerance:
            _add([i, j], all_open.loc[[i, j]], total)
    # invoice + credit note
    cred_limit = min(len(cred_idx), 50)
    for inv_i in inv_idx[:pair_limit]:
        for cred_i in cred_idx[:cred_limit]:
            total = round(amt_map[inv_i] + amt_map[cred_i], 2)
            if abs(total - target) <= tolerance:
                _add([inv_i, cred_i], all_open.loc[[inv_i, cred_i]], total)

    # 3. Greedy subset-sum — oldest invoices first, progressively adding newer
    # Run multiple passes: each pass starts from invoice[0], invoice[1], etc.
    # so we explore starting from different oldest-invoice anchors.
    # Credits are always sorted oldest-first and added when they help close the gap.
    for start_offset in range(min(len(inv_idx), 30)):
        # Build candidate list: invoices from start_offset onwards (oldest first),
        # then append all credits so they can offset overshoots
        inv_candidates  = inv_idx[start_offset:]
        ordered_indices = inv_candidates + cred_idx  # credits at end

        remaining  = target
        chosen_idx = []
        for idx in ordered_indices:
            amt = round(all_open.at[idx, "amount"], 2)
            if amt > 0 and amt <= remaining + tolerance:
                chosen_idx.append(idx)
                remaining = round(remaining - amt, 2)
                if abs(remaining) <= tolerance:
                    break
            elif amt < 0 and remaining < -tolerance:
                # Overshot — a credit brings us back in range
                chosen_idx.append(idx)
                remaining = round(remaining - amt, 2)
                if abs(remaining) <= tolerance:
                    break
        if chosen_idx and abs(remaining) <= tolerance + 0.01:
            total = round(sum(all_open.loc[chosen_idx, "amount"]), 2)
            _add(chosen_idx, all_open.loc[chosen_idx], total)
        if len(results) >= max_results:
            break

    # Sort: exact matches first, then by diff asc, then prefer sets with oldest invoices
    def _oldest_due(result):
        rows = result.get("invoices", [])
        dues = []
        for r in rows:
            d = r.get("due_date")
            if d and pd.notna(d):
                try:
                    dues.append(pd.Timestamp(d))
                except Exception:
                    pass
        return min(dues) if dues else pd.Timestamp("2099-12-31")

    results.sort(key=lambda x: (x["diff"], x["n"], _oldest_due(x)))
    return results[:max_results], sap


def build_amount_match_report(matches, payment_amount: float,
                               customer_name: str = "",
                               today=None) -> BytesIO:
    """Build an Excel showing the best invoice combinations for the payment."""
    if today is None:
        today = datetime.date.today()
    today_str = pd.Timestamp(today).strftime("%d/%m/%Y")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DK = "FF1F3864"; MD = "FF2E75B6"; WHITE = "FFFFFFFF"
    GREY = "FFF2F2F2"; BLK = "FF000000"
    RED = "FFC00000"; GRN = "FF375623"

    def fill(rgb): return PatternFill("solid", fgColor=rgb)
    def font(bold=False, color=BLK, size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def aln(ha="left"):
        return Alignment(horizontal=ha, vertical="center")
    def thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s,right=s,top=s,bottom=s)
    def mw(ws, row, c1, c2, val, bold=False, bg=WHITE, fg=BLK, sz=10, ha="left"):
        ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
        cell = ws.cell(row=row, column=c1, value=val)
        cell.font = font(bold=bold, color=fg, size=sz)
        cell.fill = fill(bg); cell.alignment = aln(ha); cell.border = thin()
        for ci in range(c1+1, c2+1):
            ws.cell(row,ci).fill = fill(bg); ws.cell(row,ci).border = thin()

    wb = openpyxl.Workbook()
    first = True

    for mi, match in enumerate(matches, 1):
        ws = wb.active if first else wb.create_sheet()
        first = False
        label = f"Option {mi} ({match['confidence']})"
        ws.title = label[:31]

        # Column widths
        cols   = ["Account","Assignment","Document Number","Reference Key 3",
                  "Document Date","Net due date","Document Type",
                  "Amount in local currency","Arrears after net due date"]
        widths = [10,12,17,13,13,13,15,22,24]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ncols = len(cols)

        # Title
        mw(ws,1,1,ncols,
           f"{customer_name or 'Customer'}  —  Payment €{payment_amount:,.2f}  —  {today_str}",
           bold=True, bg=DK, fg=WHITE, sz=13)
        ws.row_dimensions[1].height = 32

        # Subtitle
        mw(ws,2,1,ncols,
           f"{label}  ·  {match['n']} invoice(s)  ·  Total €{match['total']:,.2f}  ·  Difference €{match['diff']:,.2f}",
           bold=False, bg=MD, fg=WHITE, sz=9)
        ws.row_dimensions[2].height = 16
        ws.row_dimensions[3].height = 6

        # Headers
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(4, ci, value=h)
            cell.font = font(bold=True, color=WHITE, size=9)
            cell.fill = fill(MD); cell.alignment = aln("center"); cell.border = thin()
        ws.row_dimensions[4].height = 15

        # Data rows
        key_map = {
            "Account":                     "account",
            "Assignment":                  "assignment",
            "Document Number":             "doc_number_str",
            "Reference Key 3":             "Reference Key 3",
            "Document Date":               "doc_date",
            "Net due date":                "due_date",
            "Document Type":               "doc_type",
            "Amount in local currency":    "amount",
            "Arrears after net due date":  "Arrears after net due date",
        }
        for ri, inv_row in enumerate(match["invoices"]):
            r = 5 + ri
            row_fill = WHITE if ri % 2 == 0 else GREY
            for ci, col in enumerate(cols, 1):
                key = key_map.get(col, col.lower())
                val = inv_row.get(key, "")
                is_amt = (col == "Amount in local currency")
                is_date = col in ("Document Date","Net due date")
                if is_amt:
                    try: val = float(val)
                    except: val = 0.0
                    fg = RED if val >= 0 else GRN
                elif is_date:
                    try: val = pd.Timestamp(val).to_pydatetime()
                    except: pass
                    fg = BLK
                else:
                    fg = BLK
                cell = ws.cell(r, ci, value=val)
                cell.font = font(color=fg, size=9)
                cell.fill = fill(row_fill)
                cell.alignment = aln("right" if is_amt else "left")
                cell.border = thin()
                if is_amt: cell.number_format = "#,##0.00"
                elif is_date and isinstance(val, datetime.datetime):
                    cell.number_format = "DD/MM/YYYY"
            ws.row_dimensions[r].height = 13

        # Total row
        r_tot = 5 + len(match["invoices"])
        for ci in range(1, ncols+1):
            cell = ws.cell(r_tot, ci)
            cell.fill = fill(DK); cell.border = thin()
            if ci == 1:
                cell.value = "TOTAL"
                cell.font = font(bold=True, color=WHITE, size=10)
                cell.alignment = aln("left")
            elif ci == 8:  # Amount in local currency col
                cell.value = match["total"]
                cell.font = font(bold=True, color=WHITE, size=10)
                cell.alignment = aln("right")
                cell.number_format = "#,##0.00"
            else:
                cell.font = font(bold=True, color=WHITE, size=10)
        ws.row_dimensions[r_tot].height = 16
        ws.freeze_panes = "A5"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_invoice_credit_report(sap_file, customer_name: str = "") -> tuple:
    """
    Match each invoice to the credit notes that best offset it.
    - Oldest invoices matched first
    - Exact (net = 0) matches prioritised
    - Near matches within €100 tolerance
    - Returns (BytesIO Excel, summary dict)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from itertools import combinations as _comb

    TOL = 100.00  # max acceptable net difference per match

    sap = parse_sap(sap_file)
    open_df = sap[sap["is_open"] & sap["amount"].notna() & (sap["amount"] != 0)].copy()
    open_df["due_date"] = pd.to_datetime(open_df["due_date"], errors="coerce")
    open_df["amount"]   = pd.to_numeric(open_df["amount"],   errors="coerce")

    invoices = (open_df[open_df["amount"] > 0]
                .sort_values("due_date", na_position="last")
                .copy().reset_index(drop=True))
    credits  = (open_df[open_df["amount"] < 0]
                .sort_values("due_date", na_position="last")
                .copy().reset_index(drop=True))

    def _row(r):
        return {
            "doc":    str(r.get("doc_number_str", "") or ""),
            "due":    r.get("due_date"),
            "amt":    round(float(r["amount"]), 2),
            "type":   str(r.get("doc_type", "") or ""),
            "ref":    str(r.get("ref", "") or ""),
        }

    inv_pool  = [_row(r) for _, r in invoices.iterrows()]
    cred_pool = [_row(r) for _, r in credits.iterrows()]

    used_inv  = set()
    used_cred = set()
    matches   = []   # list of {invoice, credits, net, diff, exact}

    def _best_match(inv_amt, available):
        """Find the subset of available credits with smallest |inv_amt + sum(credits)|."""
        best_diff  = float('inf')
        best_combo = None

        # Single credit
        for c in available:
            net = round(inv_amt + c["amt"], 2)
            if abs(net) < best_diff:
                best_diff = abs(net)
                best_combo = [c]
            if best_diff == 0: return best_combo, best_diff

        # Greedy: oldest credits first, keep adding until we cover invoice
        chosen = []
        running = inv_amt
        for c in available:   # already sorted oldest-first
            if running + c["amt"] > 0.01:
                chosen.append(c)
                running = round(running + c["amt"], 2)
            if abs(running) <= TOL:
                break
        if chosen and abs(running) < best_diff:
            best_diff = abs(running)
            best_combo = chosen

        # Try improving: swap one credit for a better-fitting one
        if best_combo and best_diff > 0.01:
            current_net = round(inv_amt + sum(c["amt"] for c in best_combo), 2)
            used_in_best = {c["doc"] for c in best_combo}
            for extra in available:
                if extra["doc"] in used_in_best:
                    continue
                new_net = round(current_net + extra["amt"], 2)
                if abs(new_net) < best_diff:
                    best_diff = abs(new_net)
                    best_combo = best_combo + [extra]
                    current_net = new_net

        return best_combo, best_diff

    for inv in inv_pool:
        if inv["doc"] in used_inv:
            continue
        # Only pass credits not yet used in any previous match
        av = [c for c in cred_pool if c["doc"] not in used_cred]
        if not av:
            break
        combo, diff = _best_match(inv["amt"], av)
        if combo and diff <= TOL:
            # Final safety check: ensure none of the chosen credits slipped into used_cred
            # (defensive guard against any edge case in _best_match)
            combo = [c for c in combo if c["doc"] not in used_cred]
            if not combo:
                continue
            diff = abs(round(inv["amt"] + sum(c["amt"] for c in combo), 2))
            if diff > TOL:
                continue
            net = round(inv["amt"] + sum(c["amt"] for c in combo), 2)
            matches.append({
                "invoice": inv,
                "credits": combo,
                "net":     net,
                "diff":    diff,
                "exact":   diff < 0.01,
            })
            used_inv.add(inv["doc"])
            for c in combo:
                used_cred.add(c["doc"])

    unmatched_inv  = [i for i in inv_pool  if i["doc"] not in used_inv]
    unmatched_cred = [c for c in cred_pool if c["doc"] not in used_cred]

    # ── Final deduplication sweep ─────────────────────────────────────────────
    # Guarantee no credit appears in more than one match — if any slipped through,
    # remove it from the later match (keeping the earlier one intact)
    seen_cred_global = set()
    clean_matches = []
    for m in matches:
        clean_credits = [c for c in m["credits"] if c["doc"] not in seen_cred_global]
        for c in clean_credits:
            seen_cred_global.add(c["doc"])
        if clean_credits:
            new_net  = round(m["invoice"]["amt"] + sum(c["amt"] for c in clean_credits), 2)
            new_diff = abs(new_net)
            if new_diff <= TOL:
                clean_matches.append({**m, "credits": clean_credits,
                                      "net": new_net, "diff": new_diff,
                                      "exact": new_diff < 0.01})
    matches = clean_matches

    summary = {
        "exact":         sum(1 for m in matches if m["exact"]),
        "near":          sum(1 for m in matches if not m["exact"]),
        "unmatched_inv": len(unmatched_inv),
        "unmatched_cred":len(unmatched_cred),
    }

    # ── Build Excel ────────────────────────────────────────────────────────────
    def _fill(c): return PatternFill("solid", fgColor=c)
    def _fnt(bold=False, color="000000", size=9): return Font(bold=bold, color=color, size=size)
    def _aln(h="left"): return Alignment(horizontal=h, vertical="center")
    def _thin():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    DARK="1F3864"; MID="2E75B6"; WHT="FFFFFF"; GREY="F4F4F4"
    RED="C00000"; GREEN="00B050"; GOLD="FFC72C"; GOLDD="7F5F00"

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    today_str = pd.Timestamp("today").strftime("%d/%m/%Y")

    def _mrow(ws, r, c1, c2, val, bold=False, bg=DARK, fg=WHT, sz=10, ha="center", h=16):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(r, c1, value=val)
        cell.font = _fnt(bold=bold, color=fg, size=sz)
        cell.fill = _fill(bg)
        cell.alignment = _aln(ha)
        ws.row_dimensions[r].height = h

    def _hrow(ws, r, labels):
        for c, h in enumerate(labels, 1):
            cell = ws.cell(r, c, value=h)
            cell.font = _fnt(bold=True, color=WHT, size=9)
            cell.fill = _fill(DARK); cell.alignment = _aln("center"); cell.border = _thin()
        ws.row_dimensions[r].height = 14

    def _drow(ws, r, vals, bg=None, bold=False, fg="000000"):
        bg = bg or (GREY if r % 2 == 0 else WHT)
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, value=val)
            cell.font = _fnt(bold=bold, color=fg, size=9)
            cell.fill = _fill(bg); cell.border = _thin()
            if isinstance(val, float):
                cell.number_format = "€ #,##0.00"; cell.alignment = _aln("right")
            else:
                cell.alignment = _aln("left")
        ws.row_dimensions[r].height = 13

    def _fd(d):
        if d is None or (hasattr(d, '__class__') and d.__class__.__name__ == 'NaTType'): return ""
        try: return pd.Timestamp(d).strftime("%d/%m/%Y")
        except: return str(d)[:10]

    # Sheet 1: Summary
    ws1 = wb.create_sheet("Summary")
    for i, w in enumerate([30, 14, 14, 14, 14, 14], 1): ws1.column_dimensions[get_column_letter(i)].width = w
    r = 1
    title = f"Invoice / Credit Matching — {customer_name or 'Account'}  ·  {today_str}"
    _mrow(ws1, r, 1, 6, title, bold=True, sz=13, h=28); r+=1
    _mrow(ws1, r, 1, 6, "Oldest invoices matched first · Exact matches first · Max difference €100", sz=9, bg=MID, h=16); r+=2
    for label, val, bg, fg in [
        ("Total invoices on account",         len(inv_pool),           WHT,      "000000"),
        ("Total credit notes on account",     len(cred_pool),          WHT,      "000000"),
        ("",None,WHT,"000000"),
        ("✓ Exact matches (net = €0.00)",     summary["exact"],        "E2EFDA", GREEN),
        ("~ Near matches (net ≤ €100)",       summary["near"],         "FFF9E6", "E07000"),
        ("Unmatched invoices",                summary["unmatched_inv"],"FFF2F2", RED),
        ("Remaining unmatched credits",       summary["unmatched_cred"],WHT,     "000000"),
    ]:
        if label == "":
            ws1.row_dimensions[r].height = 5; r+=1; continue
        _drow(ws1, r, [label, "", val if val is not None else "", "", "", ""], bg=bg, fg=fg)
        ws1.cell(r,1).alignment = _aln("left"); r+=1
    r+=1
    _mrow(ws1, r, 1, 6, 'See tabs: "Matches" | "Unmatched Invoices" | "Unmatched Credits"', bg=MID, sz=9, h=16)

    # Sheet 2: All matches
    ws2 = wb.create_sheet("Matches")
    for i, w in enumerate([10, 14, 12, 12, 14, 14, 14], 1): ws2.column_dimensions[get_column_letter(i)].width = w
    r = 1
    _mrow(ws2, r, 1, 7, f"All Matches — {len(matches)} invoice(s) matched  ·  {today_str}", bold=True, sz=12, h=26); r+=1
    _mrow(ws2, r, 1, 7, "Each invoice matched to offsetting credit notes · oldest first · max €100 difference", sz=9, bg=MID, h=16); r+=1
    _hrow(ws2, r, ["Match", "Doc Number", "Due Date", "Type", "Role", "Amount", "Net / Diff"]); r+=1

    for i, m in enumerate(matches, 1):
        inv = m["invoice"]
        tag = "✓ EXACT" if m["exact"] else f"~€{m['diff']:.2f}"
        inv_bg = "1F6B3B" if m["exact"] else MID
        # Invoice row
        for ci, val in enumerate([f"#{i} {tag}", inv["doc"], _fd(inv["due"]), inv["type"], "INVOICE", float(inv["amt"]), ""], 1):
            cell = ws2.cell(r, ci, value=val)
            cell.font = _fnt(bold=True, color=WHT, size=9); cell.fill = _fill(inv_bg)
            cell.border = _thin()
            if isinstance(val, float): cell.number_format = "€ #,##0.00"; cell.alignment = _aln("right")
            else: cell.alignment = _aln("left")
        ws2.row_dimensions[r].height = 14; r+=1
        # Credit rows
        for c in m["credits"]:
            _drow(ws2, r, ["", c["doc"], _fd(c["due"]), c["type"], "CREDIT", float(c["amt"]), ""], bg="F2FFF2")
            ws2.cell(r, 6).font = _fnt(color=GREEN, size=9); r+=1
        # Net row
        net_bg = "E2EFDA" if m["exact"] else GOLD
        net_fg = GREEN if m["exact"] else GOLDD
        cred_total = sum(c["amt"] for c in m["credits"])
        for ci, val in enumerate(["", "", "", "NET", "", float(cred_total), float(m["diff"])], 1):
            cell = ws2.cell(r, ci, value=val)
            cell.font = _fnt(bold=True, color=net_fg, size=9); cell.fill = _fill(net_bg)
            cell.border = _thin()
            if isinstance(val, float): cell.number_format = "€ #,##0.00"; cell.alignment = _aln("right")
            else: cell.alignment = _aln("center")
        ws2.row_dimensions[r].height = 14; r+=2

    # Sheet 3: Unmatched invoices
    ws3 = wb.create_sheet("Unmatched Invoices")
    for i, w in enumerate([14, 12, 14, 22], 1): ws3.column_dimensions[get_column_letter(i)].width = w
    r = 1
    _mrow(ws3, r, 1, 4, f"Unmatched Invoices — {len(unmatched_inv)}  ·  No credits available within €100", bold=True, sz=12, bg=RED, h=26); r+=1
    _hrow(ws3, r, ["Doc Number", "Due Date", "Amount", "Note"]); r+=1
    for inv in unmatched_inv:
        _drow(ws3, r, [inv["doc"], _fd(inv["due"]), float(inv["amt"]), "No matching credits remain"], bg="FFF2F2")
        ws3.cell(r, 3).font = _fnt(color=RED, size=9); r+=1
    if unmatched_inv:
        _drow(ws3, r, ["TOTAL", "", float(sum(i["amt"] for i in unmatched_inv)), ""], bg="FFD7D7", bold=True, fg=RED)
        ws3.cell(r, 3).font = _fnt(bold=True, color=RED, size=9)

    # Sheet 4: Unmatched credits
    ws4 = wb.create_sheet("Unmatched Credits")
    for i, w in enumerate([14, 12, 14, 22], 1): ws4.column_dimensions[get_column_letter(i)].width = w
    r = 1
    _mrow(ws4, r, 1, 4, f"Unmatched Credits — {len(unmatched_cred)}  ·  Remaining open after matching", bold=True, sz=12, bg=MID, h=26); r+=1
    _hrow(ws4, r, ["Doc Number", "Due Date", "Amount", "Note"]); r+=1
    for crd in unmatched_cred:
        _drow(ws4, r, [crd["doc"], _fd(crd["due"]), float(crd["amt"]), "No matching invoice"], bg="F2FFF2")
        ws4.cell(r, 3).font = _fnt(color=GREEN, size=9); r+=1
    if unmatched_cred:
        _drow(ws4, r, ["TOTAL", "", float(sum(c["amt"] for c in unmatched_cred)), ""], bg="E2EFDA", bold=True, fg=GREEN)
        ws4.cell(r, 3).font = _fnt(bold=True, color=GREEN, size=9)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out, summary
