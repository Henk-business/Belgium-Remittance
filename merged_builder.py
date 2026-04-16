"""
Merged account builder.

When a customer has multiple SAP accounts that should be delivered as one
combined Excel file, this builder:
  1. Reads the structure from a reference template (column order, widths,
     header style, row heights).
  2. Produces a workbook with one sheet per account + a Summary sheet,
     matching the template exactly.
  3. Updates all titles, dates, line counts, and amounts with fresh data.

Account groups are stored in GitHub as JSON under
  templates/group_<primary_account>.json
  e.g. {"accounts": ["30172457", "30521289"], "label": "North and South Beverages"}
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO
import datetime
import re
import warnings
warnings.filterwarnings("ignore")

# ── COLOURS (match standard splitter output) ──────────────────────────────────
DK_BLUE  = "FF1F3864"
MD_BLUE  = "FF2E75B6"
WHITE    = "FFFFFFFF"
GREY     = "FFF2F2F2"
POS_FG   = "FFC00000"   # red  = positive (invoices)
NEG_FG   = "FF375623"   # green = negative (credits)
BLACK_FG = "FF000000"


def _fill(rgb): return PatternFill("solid", fgColor=rgb)
def _font(bold=False, color=BLACK_FG, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _align(ha="left"):
    return Alignment(horizontal=ha, vertical="center")
def _thin():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _w(ws, row, col, val=None, bold=False, fill=WHITE, fg=BLACK_FG,
       size=10, ha="left", fmt=None, border=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = _font(bold=bold, color=fg, size=size)
    cell.fill      = _fill(fill)
    cell.alignment = _align(ha)
    if fmt:    cell.number_format = fmt
    if border: cell.border = _thin()
    return cell

def _mw(ws, row, c1, c2, val=None, bold=False, fill=WHITE,
        fg=BLACK_FG, size=10, ha="left"):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    cell = ws.cell(row=row, column=c1, value=val)
    cell.font      = _font(bold=bold, color=fg, size=size)
    cell.fill      = _fill(fill)
    cell.alignment = _align(ha)
    cell.border    = _thin()
    for c in range(c1+1, c2+1):
        ws.cell(row=row, column=c).fill   = _fill(fill)
        ws.cell(row=row, column=c).border = _thin()
    return cell


def _read_template_structure(template_bytes: bytes) -> dict:
    """
    Extract structure info from a multi-sheet template:
    - Column headers (from first account sheet)
    - Column widths (from first account sheet)
    - Row heights
    - Header row number
    - Summary sheet column widths
    """
    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    info = {
        "account_sheets": [],
        "has_summary":    False,
        "summary_cols":   {},
        "data_cols":      [],
        "col_widths":     {},
        "header_row":     4,
        "ncols":          9,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name.lower() == "summary":
            info["has_summary"] = True
            for i in range(1, ws.max_column + 1):
                ltr = get_column_letter(i)
                info["summary_cols"][i] = ws.column_dimensions[ltr].width or 15
        else:
            info["account_sheets"].append(sheet_name)
            if not info["data_cols"]:
                # Find header row
                for r in range(1, min(10, ws.max_row + 1)):
                    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                    non_empty = [v for v in row_vals if v and str(v).strip()]
                    if len(non_empty) >= 3:
                        info["header_row"] = r
                        info["data_cols"]  = [v for v in row_vals if v]
                        info["ncols"]      = len(row_vals)
                        break
                # Column widths
                for i in range(1, ws.max_column + 1):
                    ltr = get_column_letter(i)
                    info["col_widths"][i] = ws.column_dimensions[ltr].width or 12

    return info


def _write_account_sheet(wb, ws, acc_id: str, acc_df: pd.DataFrame,
                          tmpl_info: dict, today_str: str,
                          amount_col: str):
    """Write one account sheet matching the template layout."""
    header_row = tmpl_info["header_row"]
    data_cols  = tmpl_info["data_cols"]
    ncols      = tmpl_info["ncols"]
    n_lines    = len(acc_df)

    # Set column widths
    for i, w in tmpl_info["col_widths"].items():
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Rows above header ─────────────────────────────────────────────────────
    # Row 1: Account title (merged, dark blue)
    _mw(ws, 1, 1, ncols, f"Account: {acc_id}",
        bold=True, fill=DK_BLUE, fg=WHITE, size=13, ha="left")
    ws.row_dimensions[1].height = 32

    # Row 2: Subtitle (merged, medium blue)
    subtitle = f"{today_str}  ·  {n_lines} lines  ·  Invoices not yet due removed"
    _mw(ws, 2, 1, ncols, subtitle,
        bold=False, fill=MD_BLUE, fg=WHITE, size=9, ha="left")
    ws.row_dimensions[2].height = 16

    # Row 3: blank gap
    ws.row_dimensions[3].height = 6

    # ── Column headers ────────────────────────────────────────────────────────
    for ci, col_name in enumerate(data_cols, 1):
        cell = ws.cell(row=header_row, column=ci, value=col_name)
        cell.font      = _font(bold=True, color=WHITE, size=9)
        cell.fill      = _fill(MD_BLUE)
        cell.alignment = _align("center")
        cell.border    = _thin()
    ws.row_dimensions[header_row].height = 18

    # ── Build column map: data_col name → sap_df column ───────────────────────
    def _norm(s): return re.sub(r"[^a-z0-9]", " ", str(s).lower()).split()
    col_map = {}
    for ci, tmpl_hdr in enumerate(data_cols, 1):
        t_words = set(_norm(tmpl_hdr))
        best, best_s = None, 0
        for sap_col in acc_df.columns:
            s_words = set(_norm(sap_col))
            overlap = len(t_words & s_words) / max(len(t_words | s_words), 1)
            if overlap > best_s and overlap >= 0.3:
                best_s, best = overlap, sap_col
        if best:
            col_map[ci] = best

    amt_ci = next((ci for ci, col in col_map.items()
                   if amount_col and col == amount_col), None)
    date_cols_ci = {ci for ci, col in col_map.items()
                    if any(kw in col.lower() for kw in ["date","datum"])
                    or (col in acc_df.columns and
                        pd.api.types.is_datetime64_any_dtype(acc_df[col]))}

    # ── Data rows ─────────────────────────────────────────────────────────────
    for di, (_, row_data) in enumerate(acc_df.iterrows()):
        r        = header_row + 1 + di
        row_fill = WHITE if di % 2 == 0 else GREY
        for ci in range(1, ncols + 1):
            sap_col = col_map.get(ci)
            is_amt  = (ci == amt_ci)
            is_date = (ci in date_cols_ci)

            if sap_col and sap_col in row_data.index:
                raw = row_data[sap_col]
                if is_amt:
                    val = float(raw) if pd.notna(raw) else 0.0
                    fg  = "FFC00000" if val >= 0 else "FF375623"
                elif is_date:
                    try:
                        val = pd.Timestamp(raw).to_pydatetime() if pd.notna(raw) else ""
                    except Exception:
                        val = str(raw) if pd.notna(raw) else ""
                    fg = BLACK_FG
                elif pd.isna(raw):
                    val, fg = "", BLACK_FG
                elif isinstance(raw, float) and raw == int(raw):
                    val, fg = int(raw), BLACK_FG
                else:
                    val, fg = raw, BLACK_FG
            else:
                val, fg = "", BLACK_FG

            cell = ws.cell(row=r, column=ci, value=val)
            cell.font      = _font(color=fg, size=9)
            cell.fill      = _fill(row_fill)
            cell.alignment = _align("right" if is_amt else "left")
            cell.border    = _thin()
            if is_amt:
                cell.number_format = "#,##0.00"
            elif is_date and isinstance(val, datetime.datetime):
                cell.number_format = "DD/MM/YYYY"

        ws.row_dimensions[r].height = 13

    acc_total = acc_df[amount_col].sum() if amount_col and amount_col in acc_df.columns else 0

    # ── Total row at the bottom ───────────────────────────────────────────────
    total_row = header_row + 1 + len(acc_df)
    for ci in range(1, ncols + 1):
        cell = ws.cell(row=total_row, column=ci)
        cell.fill   = _fill(DK_BLUE)
        cell.border = _thin()
        if ci == 1:
            cell.value     = "TOTAL"
            cell.font      = _font(bold=True, color=WHITE, size=10)
            cell.alignment = _align("left")
        elif ci == amt_ci:
            cell.value          = acc_total
            cell.font           = _font(bold=True, color=WHITE, size=10)
            cell.alignment      = _align("right")
            cell.number_format  = "#,##0.00"
        else:
            cell.font = _font(bold=True, color=WHITE, size=10)
    ws.row_dimensions[total_row].height = 16

    return acc_total


def build_merged_workbook(account_dfs: dict, template_bytes: bytes,
                           amount_col: str, today=None,
                           group_label: str = "",
                           lang: str = "en") -> bytes:
    """
    Build a combined workbook with one sheet per account + Summary sheet.

    account_dfs: {account_id: DataFrame}
    template_bytes: the reference template to match
    """
    if today is None:
        today = datetime.date.today()
    today_str = pd.Timestamp(today).strftime("%d/%m/%Y")

    if not template_bytes:
        raise ValueError("No template provided — upload a template for the primary account first.")
    tmpl_info = _read_template_structure(template_bytes)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    account_totals = {}

    # ── One sheet per account ─────────────────────────────────────────────────
    for acc_id, acc_df in account_dfs.items():
        ws = wb.create_sheet(title=str(acc_id)[:31])
        total = _write_account_sheet(
            wb, ws, str(acc_id), acc_df, tmpl_info,
            today_str, amount_col
        )
        account_totals[str(acc_id)] = {"total": total, "lines": len(acc_df)}


    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def build_flat_workbook(account_dfs: dict, amount_col: str,
                         today=None, group_label: str = "",
                         lang: str = "en") -> bytes:
    """
    Combine all accounts into ONE flat sheet — no per-account sheets, no summary.
    All rows merged together, sorted by net due date desc, single TOTAL row.
    """
    import datetime as _dt
    if today is None:
        today = _dt.date.today()
    today_str = pd.Timestamp(today).strftime("%d/%m/%Y")

    from splitter_engine import translate_doc_types
    # Concatenate all account dfs with translation applied
    all_dfs = []
    for acc_id, acc_df in account_dfs.items():
        all_dfs.append(translate_doc_types(acc_df.copy(), lang))
    combined = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()

    if len(combined) == 0:
        raise ValueError("No data to combine")

    # Sort by net due date desc
    ndd = next((c for c in combined.columns if "net due" in c.lower()), None)
    if ndd:
        combined = combined.sort_values(ndd, ascending=False)

    amt_ci_raw = next((c for c in combined.columns if "amount" in c.lower()), None)
    total = combined[amt_ci_raw].sum() if amt_ci_raw else 0
    n     = len(combined)
    acc_ids = list(account_dfs.keys())
    title = group_label or " + ".join(acc_ids)
    ncols = len(combined.columns)

    date_col_names = {c for c in combined.columns
                      if any(k in c.lower() for k in ["date","datum"])
                      or pd.api.types.is_datetime64_any_dtype(combined[c])}
    amt_ci = (list(combined.columns).index(amt_ci_raw) + 1) if amt_ci_raw else None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (title[:31])
    ws.freeze_panes = "A5"

    # Column widths
    for ci, col in enumerate(combined.columns, 1):
        if col == amt_ci_raw:
            ws.column_dimensions[get_column_letter(ci)].width = 20
        elif col in date_col_names:
            ws.column_dimensions[get_column_letter(ci)].width = 13
        else:
            w = max(len(col)+2, 10)
            ws.column_dimensions[get_column_letter(ci)].width = min(w, 28)

    # Row 1: title
    _mw(ws,1,1,ncols,f"{title}  —  {today_str}",
        bold=True,fill=DK_BLUE,fg=WHITE,size=13)
    ws.row_dimensions[1].height = 32

    # Row 2: subtitle
    _mw(ws,2,1,ncols,
        f"{today_str}  ·  {n} lines  ·  Invoices not yet due removed",
        bold=False,fill=MD_BLUE,fg=WHITE,size=9)
    ws.row_dimensions[2].height = 16

    # Row 3: blank gap
    ws.row_dimensions[3].height = 6

    # Row 4: column headers
    for ci, col in enumerate(combined.columns, 1):
        cell = ws.cell(row=4, column=ci, value=col)
        cell.font      = _font(bold=True, color=WHITE, size=9)
        cell.fill      = _fill(MD_BLUE)
        cell.alignment = _align("center")
        cell.border    = _thin()
    ws.row_dimensions[4].height = 18

    # Data rows
    for ri, (_, row_data) in enumerate(combined.iterrows()):
        r        = 5 + ri
        row_fill = WHITE if ri % 2 == 0 else GREY
        for ci, col in enumerate(combined.columns, 1):
            val     = row_data[col]
            is_amt  = (ci == amt_ci)
            is_date = col in date_col_names
            if is_amt:
                cell_val = float(val) if pd.notna(val) else 0.0
                fg = "FFC00000" if cell_val >= 0 else "FF375623"
            elif is_date:
                try:
                    cell_val = pd.Timestamp(val).to_pydatetime() if pd.notna(val) else ""
                except Exception:
                    cell_val = ""
                fg = BLACK_FG
            elif pd.isna(val):
                cell_val, fg = "", BLACK_FG
            elif isinstance(val, float) and val == int(val):
                cell_val, fg = int(val), BLACK_FG
            else:
                cell_val, fg = val, BLACK_FG
            cell = ws.cell(r, ci, value=cell_val)
            cell.font      = _font(color=fg, size=9)
            cell.fill      = _fill(row_fill)
            cell.alignment = _align("right" if is_amt else "left")
            cell.border    = _thin()
            if is_amt:
                cell.number_format = "#,##0.00"
            elif is_date and isinstance(cell_val, datetime.datetime):
                cell.number_format = "DD/MM/YYYY"
        ws.row_dimensions[r].height = 13

    # Total row
    r_total = 5 + len(combined)
    for ci in range(1, ncols+1):
        cell = ws.cell(r_total, ci)
        cell.fill = _fill(DK_BLUE); cell.border = _thin()
        if ci == 1:
            cell.value = "TOTAL"
            cell.font  = _font(bold=True, color=WHITE, size=10)
            cell.alignment = _align("left")
        elif ci == amt_ci:
            cell.value = total
            cell.font  = _font(bold=True, color=WHITE, size=10)
            cell.alignment = _align("right")
            cell.number_format = "#,##0.00"
        else:
            cell.font = _font(bold=True, color=WHITE, size=10)
    ws.row_dimensions[r_total].height = 16

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
