"""
POC-grouped Excel builder for NEGOBOISSONS-style accounts.

Matches the customer's template exactly:
  - Rows 1-4: customer header block (merged C1:F1, C2:F2, C3:F3, C4:F4) with pink fill
  - Row 8 col K-L merged: "Grand total" label (yellow), col M: value (yellow)
  - Each POC group starts at next available row:
      header row: col A = POC name (pink bold sz12), cols B-I = column headers (grey DDDDDD)
      first data row: pink fill bold
      remaining data rows: plain white
      subtotal row: col I only, bold
      2 blank separator rows
  - Colour: positive amounts RED, negative amounts GREEN (Belgian convention)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import warnings
warnings.filterwarnings("ignore")

# ── COLOURS (match template exactly) ──────────────────────────────────────────
PINK      = "FDB9FD"   # header / first data row fill
GREY_HDR  = "DDDDDD"   # column header fill
YELLOW    = "FFEE09"   # grand total fill
WHITE     = "FFFFFF"
# Belgian convention: positive (invoice) = red, negative (credit) = green
POS_FG    = "FFC00000"   # red for invoices (positive)
NEG_FG    = "FF375623"   # green for credits (negative)
BLACK_FG  = "000000"

# Template column widths (cols A-M)
COL_WIDTHS = [22, 11, 16.5, 14.8, 14, 11.5, 14.2, 22, 13, 13, 13, 13, 20.3]

# Data columns to show (in order)
DATA_COLS = [
    "Account", "Assignment", "Document Number", "Reference Key 3",
    "Document Date", "Net due date", "Document Type", "Amount in local currency",
]


def _f(rgb):
    return PatternFill("solid", fgColor=rgb)


def _font(bold=False, color=BLACK_FG, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)


def _align(ha="left", va="center", wrap=False):
    return Alignment(horizontal=ha, vertical=va, wrap_text=wrap)


def _w(ws, row, col, val=None, bold=False, fill=WHITE, fg=BLACK_FG,
       size=10, ha="left", fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = _font(bold=bold, color=fg, size=size)
    cell.fill      = _f(fill)
    cell.alignment = _align(ha)
    if fmt:
        cell.number_format = fmt
    return cell


def _merge_write(ws, row, c1, c2, val=None, bold=False, fill=WHITE,
                 fg=BLACK_FG, size=10, ha="left"):
    ws.merge_cells(
        f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}"
    )
    cell = ws.cell(row=row, column=c1, value=val)
    cell.font      = _font(bold=bold, color=fg, size=size)
    cell.fill      = _f(fill)
    cell.alignment = _align(ha, wrap=False)
    for c in range(c1 + 1, c2 + 1):
        ws.cell(row=row, column=c).fill = _f(fill)
    return cell


def _load_poc_names(template_bytes: bytes) -> dict:
    """Extract POC number -> POC name from the template."""
    names = {}
    try:
        wb = openpyxl.load_workbook(BytesIO(template_bytes))
        ws = wb.active
        max_r = ws.max_row or 50
        for r in range(1, max_r + 1):
            col_b = ws.cell(r, 2).value
            col_a = ws.cell(r, 1).value
            # Header row: col B = 'Account', col A = POC name
            if col_b == "Account" and col_a and not str(col_a).startswith("29"):
                name = str(col_a).strip().rstrip("\xa0").strip()
                # POC number is in the next row col A
                poc_val = ws.cell(r + 1, 1).value
                if poc_val and str(poc_val).startswith("29"):
                    names[str(poc_val).strip()] = name
    except Exception:
        pass
    return names


def build_poc_sheet(acc_df: pd.DataFrame, account_id: str,
                    template_bytes: bytes = None, today=None) -> bytes:
    """Build a POC-grouped Excel sheet matching the NEGOBOISSONS template."""
    if today is None:
        today = datetime.date.today()

    # ── Load POC names from template ──────────────────────────────────────────
    poc_names = _load_poc_names(template_bytes) if template_bytes else {}

    # ── Prepare data ──────────────────────────────────────────────────────────
    amount_col = next(
        (c for c in acc_df.columns if "amount" in c.lower()), None
    )
    ref_col = next(
        (c for c in acc_df.columns
         if "reference key 3" in c.lower() or
         ("ref" in c.lower() and "key" in c.lower() and "3" in c)), None
    )

    acc_df = acc_df.copy()
    if amount_col:
        acc_df[amount_col] = pd.to_numeric(
            acc_df[amount_col], errors="coerce").fillna(0)

    # ── Identify date columns (1-based positions) ─────────────────────────────
    date_col_names = set()
    for col in acc_df.columns:
        if any(kw in col.lower() for kw in ["date", "datum"]):
            date_col_names.add(col)
        elif pd.api.types.is_datetime64_any_dtype(acc_df[col]):
            date_col_names.add(col)

    # ── Group rows by POC (Reference Key 3 starting with "29") ───────────────
    if ref_col and ref_col in acc_df.columns:
        acc_df["_ref3"] = acc_df[ref_col].astype(str).str.strip()
    else:
        acc_df["_ref3"] = ""
        for col in acc_df.columns:
            s = acc_df[col].astype(str).str.strip()
            if s.str.startswith("29").sum() > len(acc_df) * 0.3:
                acc_df["_ref3"] = s
                break

    poc_groups  = {}
    no_poc_rows = []
    for _, row in acc_df.iterrows():
        ref = str(row["_ref3"]).strip()
        if ref.startswith("29") and len(ref) >= 7:
            poc_groups.setdefault(ref, []).append(row)
        else:
            no_poc_rows.append(row)

    sorted_pocs = sorted(poc_groups.keys())

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(account_id)[:31]

    # Set column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Customer header block (rows 1-4, matching template) ───────────────────
    # Rows 1-2: account number and name — merged C:F, pink fill, bold size 14
    ws.row_dimensions[1].height = 19
    ws.row_dimensions[2].height = 19
    ws.row_dimensions[3].height = 13
    ws.row_dimensions[4].height = 13.5

    _merge_write(ws, 1, 3, 6, account_id, bold=True, fill=PINK, size=14)
    _merge_write(ws, 2, 3, 6, "NEGOBOISSONS", bold=True, fill=PINK, size=14)
    # Rows 3-4: address (plain, small) — no content needed in generated file
    # Just set the fills for blank rows
    for r in (3, 4):
        for c in range(1, 14):
            ws.cell(r, c).fill = _f(WHITE)

    # Blank row 5
    ws.row_dimensions[5].height = 8

    # ── Grand total placeholder at row 8 (K-L merged = label, M = value) ─────
    # Will be filled after we know the total
    ws.row_dimensions[8].height = 24
    ws.merge_cells("K8:L8")
    cell_gt_label = ws.cell(8, 11, value="Grand total")
    cell_gt_label.font      = _font(bold=True, size=15)
    cell_gt_label.fill      = _f(YELLOW)
    cell_gt_label.alignment = _align("center")
    ws.cell(8, 12).fill = _f(YELLOW)  # fill merged part

    # We'll set the value after calculating grand_total
    cell_gt_value = ws.cell(8, 13)
    cell_gt_value.fill      = _f(YELLOW)
    cell_gt_value.alignment = _align("right")
    cell_gt_value.font      = _font(bold=True, size=15)
    cell_gt_value.number_format = "#,##0.00"

    # ── Determine which SAP columns to show ───────────────────────────────────
    show_cols = [c for c in DATA_COLS if c in acc_df.columns]
    if not show_cols:
        show_cols = [c for c in acc_df.columns if not c.startswith("_")]

    # Amount column position (1-based within show_cols + 1 for col A)
    # Col A = POC name/number, then show_cols start at col B
    amount_pos = None
    if amount_col and amount_col in show_cols:
        amount_pos = show_cols.index(amount_col) + 2  # +2 because col A is POC

    # ── Write POC groups ──────────────────────────────────────────────────────
    r = 6   # start at row 6 (matching template)
    grand_total = 0.0

    for poc in sorted_pocs:
        rows    = poc_groups[poc]
        poc_name = poc_names.get(poc, "")

        # ── Header row: col A = POC name, cols B onwards = column names ──────
        ws.row_dimensions[r].height = 16

        # Col A: POC name (pink, bold, size 12)
        _w(ws, r, 1, poc_name or poc, bold=True, fill=PINK, size=12)

        # Cols B onwards: column headers (grey fill)
        for ci, col_name in enumerate(show_cols, 2):
            _w(ws, r, ci, col_name, bold=False, fill=GREY_HDR, size=10, ha="center")

        r += 1

        # ── Data rows ─────────────────────────────────────────────────────────
        poc_total = 0.0
        for row_idx, data_row in enumerate(rows):
            is_first  = (row_idx == 0)
            row_fill  = PINK if is_first else WHITE
            bold_row  = is_first
            ws.row_dimensions[r].height = 16

            # Col A: POC number on first row only
            _w(ws, r, 1,
               int(poc) if is_first else None,
               bold=bold_row, fill=row_fill, size=10)

            for ci, col_name in enumerate(show_cols, 2):
                val      = data_row.get(col_name, "")
                is_amt   = (col_name == amount_col)
                is_date  = (col_name in date_col_names)

                if is_amt:
                    cell_val = float(val) if pd.notna(val) else 0.0
                    poc_total += cell_val
                    # Belgian convention: positive=red (invoice), negative=green (credit)
                    fg = POS_FG if cell_val >= 0 else NEG_FG
                elif is_date:
                    try:
                        cell_val = (pd.Timestamp(val).to_pydatetime()
                                    if pd.notna(val) else "")
                    except Exception:
                        cell_val = str(val) if pd.notna(val) else ""
                    fg = BLACK_FG
                elif pd.isna(val):
                    cell_val = ""
                    fg = BLACK_FG
                elif isinstance(val, float) and val == int(val):
                    cell_val = int(val)
                    fg = BLACK_FG
                else:
                    cell_val = val
                    fg = BLACK_FG

                cell = ws.cell(row=r, column=ci, value=cell_val)
                cell.font      = _font(bold=bold_row, color=fg, size=10)
                cell.fill      = _f(row_fill)
                cell.alignment = _align("right" if is_amt else "left")
                if is_amt:
                    cell.number_format = "#,##0.00"
                elif is_date and isinstance(cell_val, datetime.datetime):
                    cell.number_format = "DD/MM/YYYY"

            r += 1

        # ── Subtotal row: only col I (amount), all others blank ───────────────
        ws.row_dimensions[r].height = 16
        for ci in range(1, len(show_cols) + 2):
            ws.cell(r, ci).fill = _f(WHITE)

        if amount_pos:
            fg_sub = POS_FG if poc_total >= 0 else NEG_FG
            cell_sub = ws.cell(row=r, column=amount_pos, value=poc_total)
            cell_sub.font          = _font(bold=True, color=fg_sub, size=10)
            cell_sub.fill          = _f(WHITE)
            cell_sub.alignment     = _align("right")
            cell_sub.number_format = "#,##0.00"

        grand_total += poc_total
        r += 1

        # ── 2 blank separator rows ─────────────────────────────────────────────
        for _ in range(2):
            ws.row_dimensions[r].height = 8
            r += 1

    # ── Catch-all for rows with no POC ────────────────────────────────────────
    if no_poc_rows:
        ws.row_dimensions[r].height = 16
        _w(ws, r, 1, "OTHER", bold=True, fill=PINK, size=12)
        for ci, col_name in enumerate(show_cols, 2):
            _w(ws, r, ci, col_name, bold=False, fill=GREY_HDR, size=10, ha="center")
        r += 1

        other_total = 0.0
        for row_idx, data_row in enumerate(no_poc_rows):
            ws.row_dimensions[r].height = 16
            ws.cell(r, 1).fill = _f(WHITE)
            for ci, col_name in enumerate(show_cols, 2):
                val    = data_row.get(col_name, "")
                is_amt = (col_name == amount_col)
                if is_amt:
                    cell_val = float(val) if pd.notna(val) else 0.0
                    other_total += cell_val
                    fg = POS_FG if cell_val >= 0 else NEG_FG
                elif pd.isna(val):
                    cell_val = ""
                    fg = BLACK_FG
                else:
                    cell_val = val
                    fg = BLACK_FG
                cell = ws.cell(r, ci, value=cell_val)
                cell.font  = _font(color=fg, size=10)
                cell.fill  = _f(WHITE)
                if is_amt:
                    cell.number_format = "#,##0.00"
                    cell.alignment = _align("right")
            r += 1

        # Other subtotal
        if amount_pos:
            fg_sub = POS_FG if other_total >= 0 else NEG_FG
            cell_sub = ws.cell(row=r, column=amount_pos, value=other_total)
            cell_sub.font          = _font(bold=True, color=fg_sub, size=10)
            cell_sub.fill          = _f(WHITE)
            cell_sub.alignment     = _align("right")
            cell_sub.number_format = "#,##0.00"
        grand_total += other_total
        r += 1

    # ── Fill grand total ──────────────────────────────────────────────────────
    gt_fg = POS_FG if grand_total >= 0 else NEG_FG
    cell_gt_value.value = grand_total
    cell_gt_value.font  = _font(bold=True, color=gt_fg, size=15)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
