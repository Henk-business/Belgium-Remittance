"""
Chunked Excel builder.

Applies per-customer rules to produce a structured output:
  - Custom column selection and order
  - Groups rows into chunks where cumulative abs(amount) ≈ chunk_size
  - Each chunk: header row + data rows + subtotal row
  - Grand total either at bottom or in a yellow box to the right
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import warnings
warnings.filterwarnings("ignore")

# ── COLOURS ───────────────────────────────────────────────────────────────────
DK_BLUE  = "1F3864"
MD_BLUE  = "2E75B6"
LT_BLUE  = "D6E4F0"
YELLOW   = "FFFF00"
WHITE    = "FFFFFF"
GREY     = "F2F2F2"
GREEN_FG = "FF375623"
RED_FG   = "FFC00000"


def _recalc_arrears_df(df, today):
    """Recalculate Arrears after net due date based on reference date."""
    import pandas as _pd2
    ndd_col = next((c for c in df.columns if "net due" in c.lower()), None)
    arr_col = next((c for c in df.columns if "arrears" in c.lower()), None)
    if not ndd_col or not arr_col:
        return df
    df = df.copy()
    ref_ts = _pd2.Timestamp(today)
    due    = _pd2.to_datetime(df[ndd_col], errors="coerce")
    mask   = due.notna()
    days   = (ref_ts - due[mask]).dt.days
    # Write as same dtype as the column (str or numeric)
    if hasattr(df[arr_col], 'dtype') and str(df[arr_col].dtype) == 'string':
        df.loc[mask, arr_col] = days.astype(str)
    else:
        try:
            df[arr_col] = _pd2.to_numeric(df[arr_col], errors='coerce')
            df.loc[mask, arr_col] = days
        except Exception:
            df.loc[mask, arr_col] = days.astype(str)
    return df
def _thin():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val=None, bold=False, bg=WHITE, fg="000000",
          sz=9, ha="left", fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=ha, vertical="center")
    if fmt:
        c.number_format = fmt
    if border:
        c.border = _thin()
    return c

def _hdr_row(ws, row, headers, ncols):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color=WHITE, size=9)
        c.fill      = PatternFill("solid", fgColor=DK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _thin()
    ws.row_dimensions[row].height = 15


def _chunk_rows(df, amount_col, chunk_size):
    """
    Bin-pack rows into chunks as close to chunk_size (net) as possible.

    Algorithm (first-fit decreasing):
    1. Sort rows by abs(amount) descending so large items are placed first.
    2. Each row goes into the bin where abs(new_net) is closest to chunk_size,
       provided it stays within 1.3x chunk_size.
    3. If no bin fits, open a new bin.
    4. Merge any bin whose abs net < 15k into its best neighbour.
    5. Re-sort each chunk's rows by due date for display.
    6. Return chunks sorted by first due date.

    Single rows larger than chunk_size become their own chunk per spec.
    """
    if chunk_size <= 0 or amount_col not in df.columns:
        return [df]

    rows = []
    for idx, row in df.iterrows():
        try:
            amt = float(row[amount_col]) if pd.notna(row[amount_col]) else 0.0
        except (TypeError, ValueError):
            amt = 0.0
        rows.append((idx, amt))

    # Sort largest absolute value first
    rows.sort(key=lambda x: abs(x[1]), reverse=True)

    bins = []   # each entry: [net_amount, [original_indices]]

    for orig_idx, amt in rows:
        best_bin  = None
        best_dist = float('inf')

        for i, (net, _) in enumerate(bins):
            new_net = net + amt
            # Only place here if it doesn't overfill beyond 1.3x
            if abs(new_net) <= chunk_size * 1.5:
                dist = abs(abs(new_net) - chunk_size)
                if dist < best_dist:
                    best_dist = dist
                    best_bin  = i

        if best_bin is not None:
            bins[best_bin][0] += amt
            bins[best_bin][1].append(orig_idx)
        else:
            bins.append([amt, [orig_idx]])

    # Merge tiny bins (abs net < 15k) into closest neighbour
    min_keep = chunk_size * 0.60   # 24k when chunk=40k — merge anything under this
    changed = True
    while changed and len(bins) > 1:
        changed = False
        for i, (net, indices) in enumerate(bins):
            if abs(net) < min_keep:
                best_j, best_d = None, float('inf')
                for j, (netj, _) in enumerate(bins):
                    if j != i:
                        d = abs(abs(netj + net) - chunk_size)
                        if d < best_d:
                            best_d, best_j = d, j
                if best_j is not None:
                    bins[best_j][0] += net
                    bins[best_j][1].extend(indices)
                    bins.pop(i)
                    changed = True
                    break

    # Rebuild DataFrames
    result = []
    for _, (_, indices) in enumerate(bins):
        chunk_df = df.loc[indices].copy()
        # Sort within chunk by due date
        if 'due_date' in chunk_df.columns or 'Net due date' in chunk_df.columns:
            sort_col = 'due_date' if 'due_date' in chunk_df.columns else 'Net due date'
            chunk_df = chunk_df.sort_values(sort_col)
        result.append(chunk_df.reset_index(drop=True))

    # Sort chunks by their first row's due date
    def first_date(cdf):
        for col in ('due_date', 'Net due date'):
            if col in cdf.columns:
                v = cdf[col].iloc[0]
                if pd.notna(v):
                    return pd.Timestamp(v)
        return pd.Timestamp('2099-01-01')

    result.sort(key=first_date)
    return result if result else [df]


def build_chunked_sheet(acc_df: pd.DataFrame, account_id: str,
                        rule: dict, today=None) -> bytes:
    """
    Build a single-account Excel workbook applying the customer's rules.
    """
    if today is None:
        today = datetime.date.today()
    acc_df = _recalc_arrears_df(acc_df, today)
    today_str = pd.Timestamp(today).strftime("%d/%m/%Y")

    # ── Apply column selection ────────────────────────────────────────────────
    rule_cols    = rule.get("columns", [])
    show_account = rule.get("show_account", True)
    chunk_size   = float(rule.get("chunk_size", 0))
    total_pos    = rule.get("total_position", "bottom")
    sort_by      = rule.get("sort_by", ["Net due date"])

    # Determine final column list
    if rule_cols:
        # Use rule columns — only keep ones that actually exist
        cols = [c for c in rule_cols if c in acc_df.columns]
        # Add any remaining columns not in rule (at the end)
        extra = [c for c in acc_df.columns if c not in cols]
        cols = cols + extra
    else:
        cols = list(acc_df.columns)

    if not show_account:
        cols = [c for c in cols if c.lower() not in ("account", "klant", "debiteurnummer")]

    df = acc_df[cols].copy() if cols else acc_df.copy()

    # ── Find amount column ────────────────────────────────────────────────────
    amount_col = next(
        (c for c in df.columns
         if "amount" in c.lower() or "bedrag" in c.lower() or "betrag" in c.lower()),
        None
    )
    if amount_col:
        df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0)

    # ── Sort ──────────────────────────────────────────────────────────────────
    for sort_col in sort_by:
        if sort_col in df.columns:
            df = df.sort_values(sort_col).reset_index(drop=True)
            break

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(account_id)[:31]

    ncols      = len(df.columns)
    headers    = list(df.columns)
    amount_ci  = (headers.index(amount_col) + 1) if amount_col else None

    # Auto column widths (set before writing)
    for ci, col_name in enumerate(headers, 1):
        max_len = max(len(str(col_name)),
                      df[col_name].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 8), 32)

    # ── Identify date columns ─────────────────────────────────────────────────
    date_cols = set()
    for ci, col_name in enumerate(headers, 1):
        if any(kw in col_name.lower() for kw in ["date", "datum"]):
            date_cols.add(ci)
        elif pd.api.types.is_datetime64_any_dtype(df[col_name]):
            date_cols.add(ci)

    # ── Split into chunks ─────────────────────────────────────────────────────
    chunks = _chunk_rows(df, amount_col, chunk_size)

    r          = 1
    grand_total = 0.0

    for chunk_idx, chunk_df in enumerate(chunks):
        # Header row
        _hdr_row(ws, r, headers, ncols)
        ws.row_dimensions[r].height = 15
        r += 1

        # Data rows
        chunk_total = 0.0
        for ri_local, (_, row_data) in enumerate(chunk_df.iterrows()):
            row_bg = GREY if ri_local % 2 == 0 else WHITE
            for ci, (col_name, val) in enumerate(row_data.items(), 1):
                is_amt  = (ci == amount_ci)
                is_date = (ci in date_cols)

                if is_amt:
                    cell_val = float(val) if pd.notna(val) else 0.0
                    chunk_total += cell_val
                elif is_date:
                    try:
                        cell_val = pd.Timestamp(val).to_pydatetime() if pd.notna(val) else ""
                    except Exception:
                        cell_val = str(val) if pd.notna(val) else ""
                elif pd.isna(val):
                    cell_val = ""
                elif isinstance(val, float) and val == int(val):
                    cell_val = int(val)
                else:
                    cell_val = val

                c = ws.cell(row=r, column=ci, value=cell_val)
                fg = (RED_FG   if is_amt and isinstance(cell_val,(int,float)) and cell_val >= 0
                      else GREEN_FG if is_amt and isinstance(cell_val,(int,float)) and cell_val < 0
                      else "000000")
                c.font      = Font(name="Arial", size=9, color=fg)
                c.fill      = PatternFill("solid", fgColor=row_bg)
                c.alignment = Alignment(horizontal="right" if is_amt else "left",
                                        vertical="center")
                c.border    = _thin()
                if is_amt:
                    c.number_format = "#,##0.00"
                elif is_date and isinstance(cell_val, datetime.datetime):
                    c.number_format = "DD/MM/YYYY"
            ws.row_dimensions[r].height = 13
            r += 1

        grand_total += chunk_total

        # Subtotal row for this chunk
        for ci in range(1, ncols + 1):
            if ci == 1:
                _cell(ws, r, ci, None, bg=YELLOW, bold=True, border=True)
            elif ci == amount_ci:
                _cell(ws, r, ci, chunk_total, bg=YELLOW, bold=True,
                      fmt="#,##0.00", ha="right", sz=10,
                      fg=RED_FG if chunk_total >= 0 else GREEN_FG, border=True)
                # Currency symbol to the left of amount
                _cell(ws, r, amount_ci - 1, "\u20ac" if amount_ci > 1 else "",
                      bg=YELLOW, bold=True, sz=10, border=True) if amount_ci > 1 else None
            else:
                _cell(ws, r, ci, None, bg=YELLOW, border=True)
        ws.row_dimensions[r].height = 18
        r += 1

        # Blank separator row between chunks
        if chunk_idx < len(chunks) - 1:
            ws.row_dimensions[r].height = 8
            r += 1

    # ── Grand total ────────────────────────────────────────────────────────────
    if total_pos == "right":
        # Yellow box to the right — place in a wide merged cell
        total_col = ncols + 2
        ws.column_dimensions[get_column_letter(total_col)].width = 18
        ws.column_dimensions[get_column_letter(total_col + 1)].width = 4

        # Find a good row for the total box (roughly the middle)
        total_row = max(3, r // 2)
        ws.merge_cells(
            f"{get_column_letter(total_col)}{total_row}:"
            f"{get_column_letter(total_col + 1)}{total_row + 2}"
        )
        c = ws.cell(row=total_row, column=total_col)
        c.value     = f"\u20ac  {grand_total:,.2f}"
        c.font      = Font(name="Arial", bold=True, size=18,
                           color=RED_FG if grand_total >= 0 else GREEN_FG)
        c.fill      = PatternFill("solid", fgColor=YELLOW)
        c.alignment = Alignment(horizontal="center", vertical="center")

        # "TOTAL" label above
        label_row = total_row - 2
        if label_row >= 1:
            ws.merge_cells(
                f"{get_column_letter(total_col)}{label_row}:"
                f"{get_column_letter(total_col + 1)}{label_row + 1}"
            )
            cl = ws.cell(row=label_row, column=total_col)
            cl.value     = "TOTAL"
            cl.font      = Font(name="Arial", bold=True, size=14, color=WHITE)
            cl.fill      = PatternFill("solid", fgColor=DK_BLUE)
            cl.alignment = Alignment(horizontal="center", vertical="center")
    else:
        # Standard bottom total row
        r += 1
        for ci in range(1, ncols + 1):
            if ci == 1:
                _cell(ws, r, ci, "TOTAL", bg=DK_BLUE, bold=True, fg=WHITE, sz=10, border=True)
            elif ci == amount_ci:
                _cell(ws, r, ci, grand_total, bg=DK_BLUE, bold=True, fg=WHITE,
                      fmt="#,##0.00", ha="right", sz=10, border=True)
            else:
                _cell(ws, r, ci, None, bg=DK_BLUE, border=True)
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
