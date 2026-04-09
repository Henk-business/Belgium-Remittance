"""Shared utilities: SAP parsing, Excel styling, email templates."""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import warnings
warnings.filterwarnings("ignore")

# ── COLOURS ───────────────────────────────────────────────────────────────────
BG = {
    "dk_blue": "1F3864", "md_blue": "2E75B6", "lt_blue": "D6E4F0",
    "md_green": "375623", "lt_green": "E2EFDA",
    "md_red": "C00000",  "lt_red": "FFE2E2", "pink": "FFD7D7",
    "yellow": "FFF2CC",  "orange": "FCE4D6",
    "grey": "F2F2F2",    "mid_grey": "CBD5E1", "white": "FFFFFF",
    "purple": "4A235A",  "lt_purple": "F5EEF8",
    "dk_grey": "334155",
}
FG = {
    "white": "FFFFFF", "black": "000000", "md_red": "C00000",
    "md_green": "375623", "dk_red": "7B0000", "md_blue": "2E75B6",
    "grey": "595959", "purple": "4A235A",
}


def _thin():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def c(ws, row, col, val=None, bold=False, bg="white", fg="black",
      sz=10, ha="left", wrap=False, fmt=None, border=False, italic=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(name="Arial", bold=bold, color=FG.get(fg, fg), size=sz, italic=italic)
    cell.fill = PatternFill("solid", fgColor=BG.get(bg, bg))
    cell.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = _thin()
    return cell


def mr(ws, row, c1, c2, val=None, **kw):
    ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
    c(ws, row, c1, val, **kw)
    bg = kw.get("bg", "white")
    for col in range(c1 + 1, c2 + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BG.get(bg, "FFFFFF"))


def col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def hdr_row(ws, row, labels, bg="md_blue"):
    for ci, lbl in enumerate(labels, 1):
        cell = ws.cell(row=row, column=ci, value=lbl)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=BG[bg])
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 15


def fd(ts):
    try:
        return pd.Timestamp(ts).strftime("%d/%m/%Y") if pd.notna(ts) else ""
    except Exception:
        return ""


def auto_widths(ws, df, start_col=1):
    for ci, col_name in enumerate(df.columns, start_col):
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) else 0,
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 8), 42)


# ── SAP COLUMN MAP ─────────────────────────────────────────────────────────────
SAP_COL_MAP = {
    "Assignment": "assignment", "Zuordnung": "assignment",
    "Document Number": "doc_number", "Belegnummer": "doc_number",
    "Document Type": "doc_type", "Boekingssoort": "doc_type", "Belegtyp": "doc_type",
    "Document Date": "doc_date", "Boekingsdatum": "doc_date", "Belegdatum": "doc_date",
    "Net due date": "due_date", "Netto-vervaldatum": "due_date",
    "Amount in local currency": "amount", "Bedrag in lokale valuta": "amount",
    "Betrag in Hauswährung": "amount", "Amount in document currency": "amount",
    "Clearing Document": "clearing_doc", "Verrekeningsdocument": "clearing_doc",
    "Clearing date": "clearing_date", "Verrekeningsdatum": "clearing_date",
    "Text": "text", "Tekst": "text",
    "Document Header Text": "header_text", "Documentkoptekst": "header_text",
    "Account": "account", "Customer": "account", "Debtor": "account",
    "Debiteurnummer": "account", "Konto": "account",
}


def parse_sap(file_obj):
    """Parse any SAP customer line-item export to a normalised DataFrame."""
    df = pd.read_excel(file_obj, sheet_name=0, header=0, dtype=str)
    df.columns = [str(col).strip() for col in df.columns]
    rename = {k: v for k, v in SAP_COL_MAP.items() if k in df.columns}
    df = df.rename(columns=rename)

    for col in ["doc_date", "due_date", "clearing_date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    if "amount" in df.columns:
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)

    df["doc_number_str"] = (
        df.get("doc_number", pd.Series(dtype=str))
        .astype(str).str.strip().str.split(".").str[0]
    )
    df["ref"] = df.get("assignment", pd.Series(dtype=str)).astype(str).str.strip()

    def classify(row):
        dt = str(row.get("doc_type", "")).strip().upper()
        amt = row.get("amount", 0)
        if dt == "RV":
            return "CREDIT_NOTE" if amt < 0 else "INVOICE"
        if dt == "RU":
            return "CREDIT_NOTE"
        if dt in ("DZ", "ZP"):
            return "PAYMENT"
        if dt == "AB":
            return "CLEARING_RESIDUAL"
        return "OTHER"

    df["sap_class"] = df.apply(classify, axis=1)
    df["is_open"] = (
        df.get("clearing_doc", pd.Series(dtype=str)).isna()
        | (df.get("clearing_doc", pd.Series(dtype=str)).astype(str).str.strip() == "")
    )
    return df


def clean_id(val):
    if pd.isna(val):
        return None
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s if s and s.lower() not in ("nan", "none", "") else None


# ── EMAIL TEMPLATES ────────────────────────────────────────────────────────────
LANG_LABELS = {"en": "🇬🇧 English", "nl": "🇳🇱 Dutch", "fr": "🇫🇷 French"}

EMAIL_TEMPLATES = {
    "remittance": {
        "en": {
            "subject": "Account Statement — Outstanding Balance After Remittance",
            "body": (
                "Dear {customer_name},\n\n"
                "Thank you for your recent payment of {payment_amount}.\n\n"
                "Please find attached a statement of your account showing what has been "
                "matched to your remittance and what remains outstanding.\n\n"
                "Could you please arrange payment of the outstanding balance at your earliest convenience?\n\n"
                "If you have any questions regarding this statement, please do not hesitate to contact us.\n\n"
                "Kind regards,\n{sender_name}\n{company_name}"
            ),
        },
        "nl": {
            "subject": "Rekeningoverzicht — Openstaand saldo na betaling",
            "body": (
                "Geachte {customer_name},\n\n"
                "Hartelijk dank voor uw recente betaling van {payment_amount}.\n\n"
                "Bijgevoegd vindt u een overzicht van uw rekening met de verrekende facturen "
                "en het openstaande saldo.\n\n"
                "Gelieve het openstaande saldo zo spoedig mogelijk te voldoen.\n\n"
                "Voor vragen over dit overzicht kunt u altijd contact met ons opnemen.\n\n"
                "Met vriendelijke groeten,\n{sender_name}\n{company_name}"
            ),
        },
        "fr": {
            "subject": "Relevé de compte — Solde restant après remise",
            "body": (
                "Cher/Chère {customer_name},\n\n"
                "Nous vous remercions de votre récent paiement de {payment_amount}.\n\n"
                "Veuillez trouver ci-joint un relevé de votre compte indiquant les factures "
                "imputées à votre remise et le solde restant dû.\n\n"
                "Nous vous prions de bien vouloir régler le solde restant dans les meilleurs délais.\n\n"
                "Pour toute question concernant ce relevé, n'hésitez pas à nous contacter.\n\n"
                "Cordialement,\n{sender_name}\n{company_name}"
            ),
        },
    },
    "account": {
        "en": {
            "subject": "Account Statement — {account_id} — {date}",
            "body": (
                "Dear {customer_name},\n\n"
                "Please find attached your account statement as at {date}.\n\n"
                "The attached overview shows all open invoices currently on your account. "
                "Could you please arrange payment of the outstanding balance of {total_amount} "
                "either by direct debit or manual bank transfer?\n\n"
                "If you have any questions, please do not hesitate to contact us.\n\n"
                "Kind regards,\n{sender_name}\n{company_name}"
            ),
        },
        "nl": {
            "subject": "Rekeningoverzicht — {account_id} — {date}",
            "body": (
                "Geachte {customer_name},\n\n"
                "Bijgevoegd vindt u uw rekeningoverzicht per {date}.\n\n"
                "Het bijgevoegde overzicht toont alle openstaande facturen op uw rekening. "
                "Gelieve het openstaande saldo van {total_amount} zo spoedig mogelijk te voldoen "
                "via automatische incasso of handmatige overboeking.\n\n"
                "Voor vragen kunt u altijd contact met ons opnemen.\n\n"
                "Met vriendelijke groeten,\n{sender_name}\n{company_name}"
            ),
        },
        "fr": {
            "subject": "Relevé de compte — {account_id} — {date}",
            "body": (
                "Cher/Chère {customer_name},\n\n"
                "Veuillez trouver ci-joint votre relevé de compte au {date}.\n\n"
                "Le document ci-joint présente toutes les factures ouvertes sur votre compte. "
                "Nous vous prions de bien vouloir régler le solde de {total_amount} "
                "par prélèvement automatique ou virement bancaire manuel.\n\n"
                "Pour toute question, n'hésitez pas à nous contacter.\n\n"
                "Cordialement,\n{sender_name}\n{company_name}"
            ),
        },
    },
}


def get_email(template_key, lang, **kwargs):
    t = EMAIL_TEMPLATES[template_key][lang]
    return t["subject"].format(**kwargs), t["body"].format(**kwargs)


def mailto_link(to, subject, body):
    import urllib.parse
    params = urllib.parse.urlencode({"subject": subject, "body": body})
    return f"mailto:{to}?{params}"
