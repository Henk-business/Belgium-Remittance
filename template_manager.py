"""
Customer template manager.

Templates are stored in Streamlit session_state as base64-encoded xlsx files,
keyed by account number. Users can export all templates to a JSON backup file
and re-import them — this is how they survive across sessions on Streamlit Cloud.

Two template modes are supported automatically:
  - COLUMN MAP: template is a plain table with custom headers in row 1.
    The tool maps SAP columns to the template headers by name matching,
    fills data into the template's column order.
  - FULL LAYOUT: template has merged cells, logos, custom header blocks etc.
    The tool finds the first data row (first row after the last merged/styled
    header row), copies the template exactly, and fills data rows in from there,
    preserving all template formatting.
"""

import base64
import json
import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy
import warnings
warnings.filterwarnings("ignore")

TEMPLATE_STATE_KEY = "customer_templates"


# ── STORAGE ───────────────────────────────────────────────────────────────────

def _encode(file_bytes: bytes) -> str:
    return base64.b64encode(file_bytes).decode("utf-8")

def _decode(b64: str) -> bytes:
    return base64.b64decode(b64.encode("utf-8"))


def save_template(state, account_id: str, file_bytes: bytes):
    """Store a template for an account (persists in session_state)."""
    if TEMPLATE_STATE_KEY not in state:
        state[TEMPLATE_STATE_KEY] = {}
    state[TEMPLATE_STATE_KEY][str(account_id)] = _encode(file_bytes)


def delete_template(state, account_id: str):
    templates = state.get(TEMPLATE_STATE_KEY, {})
    if str(account_id) in templates:
        del state[TEMPLATE_STATE_KEY][str(account_id)]


def get_template(state, account_id: str) -> bytes | None:
    """Return template bytes for account_id, or None if not saved."""
    templates = state.get(TEMPLATE_STATE_KEY, {})
    b64 = templates.get(str(account_id))
    return _decode(b64) if b64 else None


def list_templates(state) -> list[str]:
    return list(state.get(TEMPLATE_STATE_KEY, {}).keys())


def export_templates_json(state) -> bytes:
    """Export all templates as a JSON file the user can save as a backup."""
    return json.dumps(state.get(TEMPLATE_STATE_KEY, {}), indent=2).encode("utf-8")


def import_templates_json(state, json_bytes: bytes):
    """Import templates from a previously exported JSON backup file."""
    data = json.loads(json_bytes.decode("utf-8"))
    if TEMPLATE_STATE_KEY not in state:
        state[TEMPLATE_STATE_KEY] = {}
    state[TEMPLATE_STATE_KEY].update(data)
    return len(data)


# ── TEMPLATE ANALYSIS ─────────────────────────────────────────────────────────

def _is_plain_table(wb: openpyxl.Workbook) -> tuple[bool, int, list[str]]:
    """
    Detect whether the template is a plain table (headers in one row)
    or a full custom layout (merged cells / styled header block).

    Returns (is_plain, header_row_index_1based, column_headers).
    """
    ws = wb.active

    # Count merged cell regions
    merged_regions = list(ws.merged_cells.ranges)
    has_merges = len(merged_regions) > 0

    # Scan first 20 rows to find a row that looks like column headers
    # (mostly text, no merged cells spanning it, reasonable number of values)
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_vals = [
            str(ws.cell(row=row_idx, column=ci).value or "").strip()
            for ci in range(1, ws.max_column + 1)
        ]
        # A header row: at least 2 non-empty text cells, mostly strings
        non_empty = [v for v in row_vals if v and v.lower() != "none"]
        if len(non_empty) >= 2:
            # Check none of the values look purely numeric
            text_vals = [v for v in non_empty if not _is_number(v)]
            if len(text_vals) >= 2:
                # This looks like a header row
                headers = [v for v in row_vals if v]
                # Treat as plain table if:
                # - no merges at all, OR
                # - header is in first 2 rows, OR
                # - all headers look like SAP column names (data header, not a title)
                sap_keywords = {"account","assignment","document","amount","date",
                                "reference","type","method","arrears","payment",
                                "clearing","balance","currency","gl","g/l"}
                looks_like_sap = sum(
                    1 for h in text_vals
                    if any(kw in h.lower() for kw in sap_keywords)
                ) >= min(2, len(text_vals))
                is_plain = (not has_merges) or (row_idx <= 2) or looks_like_sap
                return is_plain, row_idx, headers

    return True, 1, []


def _is_number(s: str) -> bool:
    try:
        float(s.replace(",", "").replace("€", "").replace("$", "").strip())
        return True
    except ValueError:
        return False


def _find_data_start_row(ws) -> int:
    """
    In a custom-layout template, find the first row where data should be written.
    This is the row AFTER the last row that contains merged cells or styled headers.
    """
    merged_rows = set()
    for region in ws.merged_cells.ranges:
        for row in range(region.min_row, region.max_row + 1):
            merged_rows.add(row)

    # Find the deepest row that is either merged or looks like a header
    last_header_row = 0
    for row_idx in range(1, ws.max_row + 1):
        if row_idx in merged_rows:
            last_header_row = row_idx
            continue
        # Check if this row has any fill/bold formatting (styled header)
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=ci)
            if cell.font and cell.font.bold:
                last_header_row = row_idx
                break
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb not in ("00000000", "FFFFFFFF", "FF000000"):
                last_header_row = row_idx
                break

    return last_header_row + 1


def _build_sap_to_template_map(template_headers: list[str], sap_df: pd.DataFrame) -> dict:
    """
    Map SAP column names to template column positions.
    Uses fuzzy matching: normalise both sides to lowercase words and find overlaps.
    Returns {template_col_idx_1based: sap_col_name}
    """
    def normalise(s):
        return re.sub(r"[^a-z0-9]", " ", str(s).lower()).split()

    sap_cols = list(sap_df.columns)
    mapping = {}  # template_col_position -> sap_col_name

    for ti, t_hdr in enumerate(template_headers, 1):
        if not t_hdr:
            continue
        t_words = set(normalise(t_hdr))
        best_score = 0
        best_sap = None
        for sap_col in sap_cols:
            s_words = set(normalise(sap_col))
            if not s_words:
                continue
            # Jaccard-style overlap
            overlap = len(t_words & s_words)
            score = overlap / max(len(t_words | s_words), 1)
            if score > best_score and score >= 0.3:  # at least 30% word overlap
                best_score = score
                best_sap = sap_col
        if best_sap:
            mapping[ti] = best_sap

    return mapping


# ── APPLY TEMPLATE ────────────────────────────────────────────────────────────

def apply_template(template_bytes: bytes, sap_df: pd.DataFrame) -> bytes:
    """
    Apply a customer template to a SAP DataFrame.
    Automatically detects plain table vs full custom layout.
    Returns the filled workbook as bytes.
    """
    tmpl_wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = tmpl_wb.active

    is_plain, header_row, template_headers = _is_plain_table(tmpl_wb)

    if is_plain:
        return _apply_plain_template(ws, tmpl_wb, header_row, template_headers, sap_df)
    else:
        return _apply_custom_template(ws, tmpl_wb, header_row, template_headers, sap_df)


def _apply_plain_template(ws, wb, header_row: int, template_headers: list[str],
                           sap_df: pd.DataFrame) -> bytes:
    """
    Plain table template: preserve everything above and including the header row,
    clear data below headers, fill in SAP data mapped to template columns.
    Also updates any title/subtitle rows above the header with fresh values.
    """
    import datetime as _dt
    col_map = _build_sap_to_template_map(template_headers, sap_df)

    # Detect key fields for updating title rows
    acc_col = next((c for c in sap_df.columns if c.lower() in
                    ("account","customer","debitor","konto")), None)
    amt_col = next((c for c in sap_df.columns if "amount" in c.lower()
                    or "bedrag" in c.lower()), None)
    account_id = str(sap_df[acc_col].dropna().iloc[0]).split(".")[0]                  if acc_col and len(sap_df) > 0 else ""
    today_str  = _dt.date.today().strftime("%d/%m/%Y")
    n_lines    = len(sap_df)

    # Update title rows above header_row (keep formatting, just refresh values)
    for row_idx in range(1, header_row):
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=ci)
            val  = str(cell.value or "")
            # Update account number reference
            if account_id and any(c.isdigit() for c in val) and len(val) > 5:
                # Looks like it contains the old account number — update it
                if any(str(account_id) not in val for _ in [1]):
                    pass  # keep as-is, titles stay the same across accounts
            # Update date references
            if "·" in val and "/" in val:
                # Subtitle row like "15/04/2026  ·  70 lines  ·  ..."
                import re
                new_val = re.sub(r"\d{2}/\d{2}/\d{4}", today_str, val)
                new_val = re.sub(r"\d+ lines", f"{n_lines} lines", new_val)
                new_val = re.sub(r"\d+ regels", f"{n_lines} regels", new_val)
                cell.value = new_val

    # Capture row style from first existing data row for new rows
    style_row = header_row + 1
    style_cells = {}
    if style_row <= ws.max_row:
        for ci in range(1, ws.max_column + 1):
            src = ws.cell(row=style_row, column=ci)
            from copy import copy as _copy
            style_cells[ci] = {
                "font":          _copy(src.font)      if src.font else None,
                "fill":          _copy(src.fill)      if src.fill else None,
                "border":        _copy(src.border)    if src.border else None,
                "alignment":     _copy(src.alignment) if src.alignment else None,
                "number_format": src.number_format,
            }
        alt_style = {}  # alternating row style (row after first data row)
        if style_row + 1 <= ws.max_row:
            for ci in range(1, ws.max_column + 1):
                src = ws.cell(row=style_row + 1, column=ci)
                from copy import copy as _copy
                alt_style[ci] = {
                    "font":          _copy(src.font)      if src.font else None,
                    "fill":          _copy(src.fill)      if src.fill else None,
                    "border":        _copy(src.border)    if src.border else None,
                    "alignment":     _copy(src.alignment) if src.alignment else None,
                    "number_format": src.number_format,
                }
        else:
            alt_style = style_cells

    # Clear existing data rows
    for row_idx in range(header_row + 1, ws.max_row + 2):
        for ci in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=ci).value = None

    # Write fresh data rows with template styling
    for di, (_, data_row) in enumerate(sap_df.iterrows()):
        write_row = header_row + 1 + di
        s = style_cells if di % 2 == 0 else alt_style
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(row=write_row, column=ci)
            if s:
                from copy import copy as _copy
                if s.get(ci, {}).get("font"):      cell.font      = _copy(s[ci]["font"])
                if s.get(ci, {}).get("fill"):      cell.fill      = _copy(s[ci]["fill"])
                if s.get(ci, {}).get("border"):    cell.border    = _copy(s[ci]["border"])
                if s.get(ci, {}).get("alignment"): cell.alignment = _copy(s[ci]["alignment"])
            sap_col = col_map.get(ci)
            if sap_col:
                val = _clean_val(data_row.get(sap_col, ""))
                cell.value = val
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
                elif isinstance(val, __import__("datetime").datetime):
                    cell.number_format = "DD/MM/YYYY"
        ws.row_dimensions[write_row].height = (
            ws.row_dimensions[style_row].height or 14
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _apply_custom_template(ws, wb, header_row: int, template_headers: list[str],
                            sap_df: pd.DataFrame) -> bytes:
    """
    Full custom layout template: preserve ALL existing content and formatting
    above the data area. Find the first data row, copy row styles from any
    existing data row (row after headers), then fill in SAP data.
    """
    data_start = _find_data_start_row(ws)
    col_map = _build_sap_to_template_map(template_headers, sap_df)

    # Capture the style of the first data row to use as a template for new rows
    style_row_idx = data_start
    style_cells = {}
    for ci in range(1, ws.max_column + 1):
        src = ws.cell(row=style_row_idx, column=ci)
        style_cells[ci] = {
            "font":      copy(src.font)      if src.font else None,
            "fill":      copy(src.fill)      if src.fill else None,
            "border":    copy(src.border)    if src.border else None,
            "alignment": copy(src.alignment) if src.alignment else None,
            "number_format": src.number_format,
        }

    # Clear from data_start downward
    for row_idx in range(data_start, ws.max_row + 2):
        for ci in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=ci).value = None

    # Write SAP data rows
    for di, (_, data_row) in enumerate(sap_df.iterrows()):
        write_row = data_start + di
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(row=write_row, column=ci)
            # Restore style from template row
            s = style_cells.get(ci, {})
            if s.get("font"):      cell.font      = copy(s["font"])
            if s.get("fill"):      cell.fill      = copy(s["fill"])
            if s.get("border"):    cell.border    = copy(s["border"])
            if s.get("alignment"): cell.alignment = copy(s["alignment"])

            # Fill value if this column is mapped
            sap_col = col_map.get(ci)
            if sap_col:
                val = _clean_val(data_row.get(sap_col, ""))
                cell.value = val
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
                elif isinstance(val, __import__("datetime").datetime):
                    cell.number_format = "DD/MM/YYYY"

        ws.row_dimensions[write_row].height = (
            ws.row_dimensions[style_row_idx].height or 14
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _clean_val(val):
    """Convert a pandas value to something Excel-safe."""
    import datetime
    if pd.isna(val):
        return ""
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime()
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val


# ── PREVIEW ───────────────────────────────────────────────────────────────────

def template_preview(template_bytes: bytes) -> dict:
    """Return metadata about a template for display in the UI."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    is_plain, header_row, headers = _is_plain_table(wb)
    data_start = header_row + 1 if is_plain else _find_data_start_row(ws)

    return {
        "sheet_name":   ws.title,
        "is_plain":     is_plain,
        "layout_type":  "Plain table" if is_plain else "Custom layout",
        "header_row":   header_row,
        "data_start":   data_start,
        "headers":      headers,
        "max_col":      ws.max_column,
        "max_row":      ws.max_row,
    }
