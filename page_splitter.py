import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import traceback

from splitter_engine import (
    split_accounts, build_split_workbook, build_template_sheet,
    build_individual_sheet, CHUNKED_ACCOUNTS, translate_doc_types,
)
from template_manager import template_preview
from template_manager import (
    save_template as _sess_save,
    delete_template as _sess_del,
    get_template as _sess_get,
    list_templates as _sess_list,
    export_templates_json as _sess_export,
    import_templates_json as _sess_import,
)
from github_storage import (
    github_configured, list_github_templates,
    get_template_cached, save_github_template, delete_github_template,
    invalidate_cache,
    save_account_group, load_account_group, list_account_groups, delete_account_group,
)
from common import clean_id, get_email, LANG_LABELS, mailto_link
from chunked_builder import build_chunked_sheet
from poc_builder import build_poc_sheet
from merged_builder import build_merged_workbook, build_flat_workbook
from customer_rules import (
    load_rule_github as _load_rule_direct,
    get_rule_cached, save_rule_github, delete_rule_github,
    invalidate_rule_cache, merge_rule, DEFAULT_RULE,
    _gh_ok as rules_gh_ok,
)

ACCOUNT_COLS  = ["Account", "Customer", "Debtor", "Klant",
                 "Debiteurnummer", "Debiteur", "Konto", "Kundennummer"]
AMOUNT_COLS   = ["Amount in local currency", "Bedrag in lokale valuta",
                 "Betrag in Hauswährung", "Amount in document currency"]
DUE_DATE_COLS = ["Net due date", "Netto-vervaldatum", "Nettofälligkeitsdatum"]


def _find(df, candidates):
    for n in candidates:
        if n in df.columns:
            return n
    for col in df.columns:
        for cand in candidates:
            if cand.lower() in col.lower():
                return col
    return None


def show():
    st.markdown("## 📂 Account Splitter")
    st.caption(
        "Split a multi-account SAP export into one sheet per customer. "
        "Removes invoices not yet due and applies customer templates automatically."
    )



    # ── UPLOAD ────────────────────────────────────────────────────────────────
    st.markdown("### 1 · Upload SAP export")
    st.markdown("**SAP Multi-Account Export** — FBL5N or any open items report (.xlsx)")
    uploaded = st.file_uploader(
        "SAP export", type=["xlsx", "xls"],
        label_visibility="collapsed", key="spl_file",
    )

    if not uploaded:
        st.info("Export from SAP (FBL5N) with your customer account range, save as .xlsx, and upload here.")
        _template_manager()
        return

    # ── PARSE FILE ────────────────────────────────────────────────────────────
    try:
        df_raw = pd.read_excel(uploaded, sheet_name=0, header=0, dtype=str)
        df_raw.columns = [str(col).strip() for col in df_raw.columns]
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return

    account_col  = _find(df_raw, ACCOUNT_COLS)
    amount_col   = _find(df_raw, AMOUNT_COLS)
    due_date_col = _find(df_raw, DUE_DATE_COLS)

    if not account_col:
        st.error("Could not detect an account column. Use the override below.")

    # ── SETTINGS ──────────────────────────────────────────────────────────────
    st.markdown("### 2 · Confirm settings")

    with st.expander("Column detection — click to override if needed"):
        col_opts = df_raw.columns.tolist()
        account_col = st.selectbox(
            "Account column", col_opts,
            index=col_opts.index(account_col) if account_col in col_opts else 0,
            key="spl_acc_col",
        )
        amount_col = st.selectbox(
            "Amount column", ["(none)"] + col_opts,
            index=col_opts.index(amount_col) + 1 if amount_col in col_opts else 0,
            key="spl_amt_col",
        )
        due_date_col = st.selectbox(
            "Due date column", ["(none)"] + col_opts,
            index=col_opts.index(due_date_col) + 1 if due_date_col in col_opts else 0,
            key="spl_due_col",
        )
        if amount_col == "(none)":
            amount_col = None
        if due_date_col == "(none)":
            due_date_col = None

    accounts = sorted([
        clean_id(a) for a in df_raw[account_col].dropna().unique()
        if clean_id(a) is not None
    ])

    if accounts:
        st.markdown(f"**{len(accounts)} accounts detected:** " + "  ".join(f"`{a}`" for a in accounts))
    else:
        st.warning("No accounts found in the selected column.")

    c1, c2 = st.columns(2)
    with c1:
        remove_not_due = st.checkbox("Remove invoices not yet due", value=True, key="spl_remove")
    with c2:
        ref_date = st.date_input("Reference date", value=datetime.date.today(), key="spl_refdate")

    # ── GENERATE ──────────────────────────────────────────────────────────────
    st.markdown("### 3 · Generate")
    gen_col, _ = st.columns([1, 2])
    with gen_col:
        generate = st.button(
            "▶  Split into separate sheets",
            use_container_width=True,
            type="primary",
            key="spl_run",
        )

    if generate:
        if not accounts:
            st.error("No accounts detected — check your column selection.")
        else:
            with st.spinner(f"Splitting {len(accounts)} accounts…"):
                try:
                    if due_date_col:
                        df_raw[due_date_col] = pd.to_datetime(df_raw[due_date_col], errors="coerce")
                    if amount_col:
                        df_raw[amount_col] = pd.to_numeric(df_raw[amount_col], errors="coerce").fillna(0)

                    account_data = split_accounts(
                        df_raw, account_col, amount_col, due_date_col,
                        remove_not_due=remove_not_due,
                        reference_date=ref_date,
                    )
                    # Build the combined workbook — standard layout for all accounts.
                    wb_bytes = build_split_workbook(account_data, amount_col, today=ref_date)
                    st.session_state["spl_result"]       = wb_bytes
                    st.session_state["spl_account_data"] = account_data
                    st.session_state["spl_ref_date"]     = ref_date
                    st.session_state["spl_amount_col"]   = amount_col
                    # Pre-fetch which accounts have templates so we can show a notice
                    has_templates = []
                    for acc_check in account_data.keys():
                        if github_configured():
                            tb = get_template_cached(str(acc_check))
                        else:
                            tb = _sess_get(st.session_state, str(acc_check))
                        if tb:
                            has_templates.append(str(acc_check))
                    st.session_state["spl_has_templates"] = has_templates
                except Exception as e:
                    st.error(f"Error: {e}")
                    with st.expander("Detail"):
                        st.code(traceback.format_exc())

    if "spl_result" not in st.session_state:
        _template_manager()
        return

    # ── RESULTS ───────────────────────────────────────────────────────────────
    account_data = st.session_state["spl_account_data"]
    ref_date     = st.session_state["spl_ref_date"]
    amount_col   = st.session_state["spl_amount_col"]
    today_str    = pd.Timestamp(ref_date).strftime("%d/%m/%Y")
    safe_date    = str(ref_date).replace("-", "")

    st.markdown("---")
    st.success(f"Done — {len(account_data)} account sheets generated")

    summary_rows = []
    for acc, acc_df in account_data.items():
        total = acc_df[amount_col].sum() if amount_col and amount_col in acc_df.columns else None
        summary_rows.append({
            "Account": acc, "Lines": len(acc_df),
            "Total (€)": f"{total:,.2f}" if total is not None else "—",
        })
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    # ── SECTION 1: DOWNLOADS ────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Downloads")
    st.caption("Custom rules and templates are applied automatically per account.")

    # Language selector — must come BEFORE the download button so the translated
    # workbook is built with the correct language when the user clicks download.
    dl_l, dl_r = st.columns([1, 3])
    with dl_l:
        dl_lang = st.selectbox(
            "Document language",
            ["en", "nl", "fr"],
            format_func=lambda x: {"en": "🇬🇧 English", "nl": "🇳🇱 Dutch", "fr": "🇫🇷 French"}[x],
            key="spl_dl_lang",
            help="Translates Document Type column: Invoice, Credit note, Payment, etc.",
        )

    # Build the combined workbook with the selected language (happens on every render,
    # cheap enough for typical account counts).
    _translated_all = build_split_workbook(
        account_data, amount_col, today=ref_date, lang=dl_lang
    )

    st.download_button(
        "⬇  Download all accounts (standard layout)",
        data=_translated_all.getvalue(),
        file_name=f"Accounts_{safe_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, key="spl_dl_all",
    )
    # ── Build account group mapping ──────────────────────────────────────────
    # Groups: {primary_id: {accounts:[...], label:..., tmpl_bytes:...}}
    group_map    = {}  # acc_id -> primary_id (which group it belongs to)
    group_defs   = {}  # primary_id -> group definition
    if github_configured():
        for grp in list_account_groups():
            accs  = grp.get("accounts", [])
            # Only activate a group if ALL its accounts are in this split
            if all(str(a) in [str(k) for k in account_data.keys()] for a in accs):
                primary = accs[0]
                tmpl    = get_template_cached(primary) or                           next((get_template_cached(a) for a in accs
                                if get_template_cached(a)), None)
                group_defs[primary] = {**grp, "tmpl_bytes": tmpl}
                for a in accs:
                    group_map[str(a)] = primary

    # Track which groups have already been written
    written_groups = set()

    st.markdown("**Individual downloads (with custom rules/templates applied):**")

    for acc, acc_df_sel in account_data.items():
        total_sel = acc_df_sel[amount_col].sum() if amount_col and amount_col in acc_df_sel.columns else 0

        # ── ACCOUNT GROUP: build combined file once for all accounts in group ──
        primary = group_map.get(str(acc))
        if primary:
            if primary in written_groups:
                continue  # already written this group
            grp_def  = group_defs[primary]
            grp_accs = grp_def["accounts"]
            grp_dfs  = {str(a): account_data[a].copy()
                        for a in grp_accs if a in account_data}
            tmpl_b   = grp_def.get("tmpl_bytes")
            label    = grp_def.get("label", "")
            # Always compute these before try so they're always defined
            grp_name  = label or " + ".join(str(a) for a in grp_accs)
            safe_grp  = grp_name.replace(" ", "_")[:30]
            grp_total = sum(df[amount_col].sum() for df in grp_dfs.values()
                            if amount_col and amount_col in df.columns)

            grp_bytes = None
            grp_label = ""
            is_flat   = grp_def.get("flat", False)
            if is_flat:
                try:
                    grp_bytes = build_flat_workbook(
                        grp_dfs, amount_col,
                        today=ref_date, group_label=label, lang=dl_lang,
                    )
                    grp_label = f"✓ flat merged ({len(grp_accs)} accounts)"
                except Exception as e:
                    grp_label = f"merge error: {e}"
            elif not tmpl_b:
                grp_label = "merge error: no template found — upload a template for the primary account first"
            else:
                try:
                    grp_bytes = build_merged_workbook(
                        grp_dfs, tmpl_b, amount_col,
                        today=ref_date, group_label=label, lang=dl_lang,
                    )
                    grp_label = f"✓ merged ({len(grp_accs)} accounts)"
                except Exception as e:
                    grp_label = f"merge error: {e}"

            written_groups.add(primary)
            dl_c, info_c = st.columns([2, 3])
            with dl_c:
                if grp_bytes:
                    st.download_button(
                        f"⬇  {grp_name}",
                        data=grp_bytes,
                        file_name=f"{safe_grp}_{safe_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"spl_dl_grp_{primary}",
                        use_container_width=True,
                    )
                else:
                    st.error(grp_label)
            with info_c:
                st.caption(
                    f"{', '.join(str(a) for a in grp_accs)}  ·  "
                    f"€{grp_total:,.2f}  ·  {grp_label}"
                )
            continue  # skip individual download for this account

        # Apply document type translation
        acc_df_sel = translate_doc_types(acc_df_sel, dl_lang)

        # Priority: 1) chunked rule  2) POC template  3) plain template  4) standard
        # Load template and rule first so we pick the right builder
        tmpl_bytes = None
        try:
            if github_configured():
                tmpl_bytes = get_template_cached(str(acc))
            else:
                tmpl_bytes = _sess_get(st.session_state, str(acc))
        except Exception:
            pass

        rule = st.session_state.get(f"rule_{acc}", None)
        if rule is None:
            try:
                if rules_gh_ok():
                    loaded = _load_rule_direct(str(acc))
                    if loaded:
                        rule = loaded
                        st.session_state[f"rule_{acc}"] = rule
            except Exception:
                pass

        try:
            if rule and rule.get("chunk_size", 0) > 0:
                # Chunked account (e.g. 40k batches)
                acc_wb_bytes = build_chunked_sheet(acc_df_sel, str(acc), rule, today=ref_date, lang=dl_lang)
                layout_label = f"✓ chunked €{rule.get('chunk_size',0):,.0f}"

            elif str(acc) in CHUNKED_ACCOUNTS:
                # Hardcoded chunked account (legacy)
                acc_wb_bytes = build_individual_sheet(str(acc), acc_df_sel, amount_col, today=ref_date)
                layout_label = f"✓ chunked €{CHUNKED_ACCOUNTS[str(acc)]['chunk_size']:,.0f}"

            elif tmpl_bytes:
                # Detect POC by checking the TEMPLATE structure, not the SAP data
                # (SAP data always has 29xxxxx in Ref Key 3, but only NEGO-style
                # templates have 29xxxxx POC numbers in column A)
                try:
                    import openpyxl as _oxl, io as _io2
                    _twb = _oxl.load_workbook(_io2.BytesIO(tmpl_bytes))
                    _tws = _twb.active
                    _maxr = min((_tws.max_row or 20), 20)
                    is_poc = any(
                        str(_tws.cell(r, 1).value or '').strip().startswith('29')
                        for r in range(1, _maxr + 1)
                    )
                except Exception:
                    is_poc = False
                if is_poc:
                    acc_wb_bytes = build_poc_sheet(acc_df_sel, str(acc), tmpl_bytes, today=ref_date, lang=dl_lang)
                    layout_label = '✓ POC grouped'
                else:
                    acc_wb_bytes = build_template_sheet(
                        str(acc), acc_df_sel, tmpl_bytes, amount_col, today=ref_date)
                    layout_label = '✓ custom template'

            else:
                # Standard layout
                acc_wb_bytes = build_split_workbook(
                    {acc: acc_df_sel}, amount_col, today=ref_date,
                    title_prefix=f"Account {acc} — ", lang=dl_lang).getvalue()
                layout_label = "standard layout"

        except Exception as e:
            acc_wb_bytes = build_split_workbook(
                {acc: acc_df_sel}, amount_col, today=ref_date, lang=dl_lang).getvalue()
            layout_label = f"error: {e}"

        dl_c, info_c = st.columns([2, 3])
        with dl_c:
            st.download_button(
                f"⬇  Account {acc}",
                data=acc_wb_bytes,
                file_name=f"Account_{acc}_{safe_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key=f"spl_dl_{acc}",
            )
        with info_c:
            st.caption(f"{len(acc_df_sel)} lines  ·  €{total_sel:,.2f}  ·  {layout_label}")

    # ── SECTION 2: EMAILS ───────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📧 Payment reminder emails")

    e1, e2, e3 = st.columns(3)
    with e1:
        lang    = st.selectbox("Language", list(LANG_LABELS.keys()), format_func=lambda x: LANG_LABELS[x], key="spl_lang")
    with e2:
        sender  = st.text_input("Your name",  key="spl_sender",  placeholder="Your Name",
                               value=st.session_state.get("_persist_sender",""))
    with e3:
        company = st.text_input("Company",    key="spl_company", placeholder="Your Company",
                               value=st.session_state.get("_persist_company",""))

    # Persist sender/company
    if sender: st.session_state["_persist_sender"]  = sender
    if company: st.session_state["_persist_company"] = company

    for acc_idx, (acc, acc_df_sel) in enumerate(account_data.items()):
        total_sel = acc_df_sel[amount_col].sum() if amount_col and amount_col in acc_df_sel.columns else 0
        subject, body = get_email(
            "account", lang,
            customer_name=f"Account {acc}", account_id=str(acc), date=today_str,
            total_amount=f"€{total_sel:,.2f}",
            sender_name=sender or "[Your Name]", company_name=company or "[Your Company]",
        )
        with st.expander(f"Account {acc}  ·  €{total_sel:,.2f}", expanded=(acc_idx == 0)):
            st.text_input("Subject", value=subject, key=f"spl_subj_{acc}")
            st.text_area("Body",    value=body,    height=200, key=f"spl_body_{acc}")
            to_email = st.text_input("Customer email", key=f"spl_to_{acc}", placeholder="customer@example.com")
            if to_email:
                mailto = mailto_link(to_email, subject, body)
                st.markdown(
                    f'<a href="{mailto}" style="display:block;text-align:center;'
                    f'background:linear-gradient(135deg,#0f2942,#1d4ed8);color:white;'
                    f'font-weight:600;padding:12px;border-radius:8px;'
                    f'text-decoration:none;font-size:13px;margin-top:4px;">'
                    f'📧 Open in Email Client</a>',
                    unsafe_allow_html=True,
                )
            else:
                st.caption("Enter email above for mailto link.")

    # ── TEMPLATE MANAGER ──────────────────────────────────────────────────────
    st.markdown("---")
    _template_manager()


def _is_admin() -> bool:
    """Return True if the current session has authenticated as admin."""
    return st.session_state.get("_admin_authenticated", False)


def _template_manager():
    """Clean GitHub-backed template + rules manager."""

    # ── Admin authentication ──────────────────────────────────────────────────
    # Delete / replace actions are restricted to the admin.
    # The admin password is stored in st.secrets["admin"]["password"].
    # If no password is configured, admin mode is disabled for everyone.
    _admin_pw = st.secrets.get("admin", {}).get("password", "")
    if _admin_pw:
        if not _is_admin():
            with st.expander("🔐 Admin login (required to delete or replace templates)"):
                pw_input = st.text_input(
                    "Admin password", type="password", key="tmgr_pw_input"
                )
                if st.button("Unlock", key="tmgr_pw_btn"):
                    if pw_input == _admin_pw:
                        st.session_state["_admin_authenticated"] = True
                        st.rerun()
                    else:
                        st.error("Incorrect password.")

    st.markdown("""
    <style>
    .tmpl-card {
        background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 12px;
        padding: 16px 20px; margin-bottom: 10px;
    }
    .tmpl-card .acc  { font-size:15px; font-weight:700; color:#0f172a; }
    .tmpl-card .meta { font-size:12px; color:#64748b; margin-top:2px; }
    .tmpl-tag {
        display:inline-block; background:#dbeafe; color:#1e40af;
        font-size:11px; font-weight:600; padding:2px 8px;
        border-radius:20px; margin-right:4px;
    }
    .tmpl-tag.poc  { background:#fce7f3; color:#9d174d; }
    .tmpl-tag.rule { background:#d1fae5; color:#065f46; }
    </style>
    """, unsafe_allow_html=True)

    if not github_configured():
        _template_manager_session()
        return

    with st.spinner("Loading from GitHub…"):
        saved = list_github_templates()

    # ── SAVED TEMPLATES ──────────────────────────────────────────────────────
    if saved:
        for tmpl_info in saved:
            acc_id     = tmpl_info["account_id"]
            size_kb    = tmpl_info["size"] / 1024
            tmpl_bytes = get_template_cached(acc_id)

            # Detect type
            layout_type = "Custom"
            tag_class   = ""
            if tmpl_bytes:
                try:
                    info = template_preview(tmpl_bytes)
                    if info.get("layout_type") == "Plain table":
                        layout_type = "Plain table"
                    else:
                        layout_type = "Custom layout"
                    # Check if POC template
                    import openpyxl, io as _io
                    wb_check = openpyxl.load_workbook(_io.BytesIO(tmpl_bytes))
                    ws_check = wb_check.active
                    has_poc = any(
                        str(ws_check.cell(r,1).value or "").startswith("29")
                        for r in range(1, min(20, (ws_check.max_row or 20)+1))
                    )
                    if has_poc:
                        layout_type = "POC grouped"
                        tag_class   = "poc"
                except Exception:
                    pass

            # Check if this account also has a rule
            from github_storage import _repo as _gr_check
            has_rule = bool(get_rule_cached(acc_id, _gr_check()) if rules_gh_ok() else None)

            tags_html = f'<span class="tmpl-tag {tag_class}">{layout_type}</span>'
            if has_rule:
                tags_html += '<span class="tmpl-tag rule">✓ chunked</span>'

            col_main, col_btns = st.columns([4, 2])
            with col_main:
                st.markdown(f"""
                <div class="tmpl-card">
                    <div class="acc">Account {acc_id}</div>
                    <div class="meta">{tags_html}&nbsp;&nbsp;{size_kb:.1f} KB</div>
                </div>
                """, unsafe_allow_html=True)
            with col_btns:
                st.write("")
                b1, b2, b3 = st.columns(3)
                with b1:
                    if tmpl_bytes:
                        st.download_button(
                            "⬇", data=tmpl_bytes,
                            file_name=f"template_{acc_id}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"gh_dl_{acc_id}", use_container_width=True,
                            help="Download template",
                        )
                with b2:
                    if _is_admin() and st.button("🔄", key=f"gh_replace_{acc_id}",
                                 use_container_width=True, help="Replace template"):
                        st.session_state[f"replacing_{acc_id}"] = True
                with b3:
                    if _is_admin() and st.button("🗑", key=f"gh_del_{acc_id}",
                                 use_container_width=True, help="Delete template"):
                        with st.spinner("Deleting…"):
                            ok, msg = delete_github_template(acc_id)
                        if ok:
                            invalidate_cache()
                            st.rerun()
                        else:
                            st.error(msg)

            if _is_admin() and st.session_state.get(f"replacing_{acc_id}"):
                rep_file = st.file_uploader(
                    f"Upload new template for {acc_id}",
                    type=["xlsx","xls"], key=f"gh_rep_file_{acc_id}",
                )
                if rep_file:
                    with st.spinner("Uploading…"):
                        ok, msg = save_github_template(acc_id, rep_file.read())
                    if ok:
                        invalidate_cache()
                        del st.session_state[f"replacing_{acc_id}"]
                        st.rerun()
                    else:
                        st.error(msg)
    else:
        st.info("No templates saved yet. Upload one below to get started.")

    st.write("")

    # ── UPLOAD NEW ────────────────────────────────────────────────────────────
    with st.expander("➕  Upload a new customer template"):
        c1, c2 = st.columns([1, 2])
        with c1:
            acc_input = st.text_input(
                "Account number", key="gh_tmpl_acc", placeholder="e.g. 30104552",
            )
        with c2:
            new_file = st.file_uploader(
                "Template file (.xlsx)", type=["xlsx","xls"], key="gh_tmpl_file",
            )

        if new_file and acc_input:
            raw = new_file.read()
            try:
                info = template_preview(raw)
                m1, m2, m3 = st.columns(3)
                m1.metric("Layout",     info["layout_type"])
                m2.metric("Columns",    info["max_col"])
                m3.metric("Header row", info["header_row"])
                if info["headers"]:
                    st.caption("Columns: " + "  ·  ".join(info["headers"][:8]) +
                               ("…" if len(info["headers"]) > 8 else ""))
            except Exception:
                pass

            if _is_admin() and st.button(
                f"💾  Save template for account {acc_input}",
                key="gh_tmpl_save", type="primary", use_container_width=True,
            ):
                with st.spinner("Saving to GitHub…"):
                    ok, msg = save_github_template(acc_input, raw)
                if ok:
                    invalidate_cache()
                    st.success(f"Template saved for account {acc_input}.")
                    st.rerun()
                else:
                    st.error(msg)
        elif new_file and not acc_input:
            st.warning("Enter the account number first.")

    # ── CUSTOM RULES ──────────────────────────────────────────────────────────
    with st.expander("⚙️  Chunking rules per account"):
        st.caption(
            "Set a chunk size to group invoices into payment batches "
            "(e.g. €40,000 each). Rules are saved permanently to GitHub."
        )

        rule_acc = st.text_input(
            "Account number", key="rule_acc_input", placeholder="e.g. 30111788",
        )

        if rule_acc:
            from github_storage import _repo as _gr2
            existing = get_rule_cached(rule_acc, _gr2()) if rules_gh_ok() else None
            base = existing or DEFAULT_RULE.copy()

            r1, r2 = st.columns(2)
            with r1:
                chunk_size = st.number_input(
                    "Chunk size (€)", min_value=0.0,
                    value=float(base.get("chunk_size", 0)),
                    step=1000.0, format="%.0f", key="rule_chunk",
                    help="0 = no chunking",
                )
                show_account = st.checkbox(
                    "Include Account column",
                    value=base.get("show_account", True), key="rule_show_acc",
                )
            with r2:
                total_position = st.radio(
                    "Grand total", ["bottom", "right"],
                    index=0 if base.get("total_position","bottom") == "bottom" else 1,
                    key="rule_total_pos",
                )
                sort_col = st.text_input(
                    "Sort by column",
                    value=(base.get("sort_by") or ["Net due date"])[0],
                    key="rule_sort",
                )

            cols_text = st.text_area(
                "Column order (one per line, leave blank for default)",
                value="\n".join(base.get("columns", [])),
                height=110, key="rule_cols",
            )

            rule_obj = {
                "chunk_size":     chunk_size,
                "show_account":   show_account,
                "total_position": total_position,
                "columns":        [c.strip() for c in cols_text.strip().splitlines() if c.strip()],
                "sort_by":        [sort_col] if sort_col.strip() else ["Net due date"],
            }

            rb1, rb2, rb3 = st.columns(3)
            with rb1:
                if _is_admin() and st.button("💾  Save rule", key="rule_save",
                             type="primary", use_container_width=True):
                    st.session_state[f"rule_{rule_acc}"] = rule_obj
                    try:
                        ok, msg = save_rule_github(rule_acc, rule_obj)
                        if ok:
                            invalidate_rule_cache()
                            st.success(f"Rule saved for account {rule_acc}.")
                        else:
                            st.warning("Saved to session. GitHub: " + msg)
                    except Exception as e:
                        st.warning(f"Saved to session only. ({e})")
                    st.rerun()
            with rb2:
                if st.button("🔍  Verify", key="rule_verify",
                             use_container_width=True):
                    loaded = _load_rule_direct(rule_acc)
                    if loaded:
                        st.success(
                            f"Rule found · chunk=€{loaded.get('chunk_size',0):,.0f}"
                            f" · total={loaded.get('total_position','bottom')}"
                        )
                    else:
                        st.error("No rule found in GitHub for this account.")
            with rb3:
                if existing and _is_admin() and st.button("🗑  Delete", key="rule_del",
                                          use_container_width=True):
                    ok, msg = delete_rule_github(rule_acc)
                    if ok:
                        invalidate_rule_cache()
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)


    # ── ACCOUNT GROUPS ────────────────────────────────────────────────────────
    with st.expander("📋  Account groups (combined downloads)"):
        st.caption(
            "Group multiple accounts into one combined Excel file. "
            "Upload a template for the primary account first, then define the group here."
        )

        if github_configured():
            existing_groups = list_account_groups()
            if existing_groups:
                st.markdown("**Saved groups:**")
                for grp in existing_groups:
                    accs  = grp.get("accounts", [])
                    label = grp.get("label", "")
                    primary = accs[0] if accs else ""
                    g1, g2 = st.columns([4, 1])
                    with g1:
                        st.markdown(
                            f"**{label or primary}** — "
                            f"{', '.join(accs)}"
                        )
                    with g2:
                        if _is_admin() and st.button("🗑", key=f"del_grp_{primary}",
                                     use_container_width=True, help="Delete group"):
                            ok, msg = delete_account_group(primary)
                            if ok:
                                st.rerun()
                            else:
                                st.error(msg)

            st.markdown("**Create a new group:**")
            ng1, ng2 = st.columns(2)
            with ng1:
                grp_label = st.text_input(
                    "Group label (e.g. customer name)",
                    key="grp_label", placeholder="North and South Beverages Belgium"
                )
            with ng2:
                grp_accs_raw = st.text_input(
                    "Account numbers (comma-separated)",
                    key="grp_accs", placeholder="30172457, 30521289"
                )


            flat_merge = st.checkbox(
                "Flat merge — combine all accounts into one sheet (no separate tabs)",
                key="grp_flat",
                help="Use this when all accounts should appear as one combined list",
            )

            if _is_admin() and st.button("💾  Save group", key="grp_save", type="primary"):
                accs_list = [a.strip() for a in grp_accs_raw.split(",") if a.strip()]
                if len(accs_list) < 2:
                    st.error("Enter at least 2 account numbers separated by commas.")
                else:
                    ok, msg = save_account_group(accs_list[0], accs_list, grp_label.strip(), flat=flat_merge)
                    if ok:
                        st.success(f"Group saved: {msg}")
                        st.rerun()
                    else:
                        st.error(msg)
        else:
            st.warning("GitHub storage must be configured to save account groups.")


def _template_manager_session():
    """Session-only fallback when GitHub is not configured."""
    st.warning(
        "GitHub storage is not configured — templates last this session only. "
        "Follow the setup steps below to make them permanent."
    )

    with st.expander("⚙️  One-time setup — permanent templates"):
        st.markdown("""
**Step 1 — Create a GitHub token**
1. Go to github.com → Settings → Developer settings → Personal access tokens → Fine-grained tokens
2. Click **Generate new token** · Name: `AR Suite` · Expiration: **No expiration**
3. Repository access → Only select repositories → pick your repo
4. Permissions → **Contents → Read and write**
5. Generate and copy the token

**Step 2 — Add to Streamlit secrets**

Streamlit Cloud → your app → ⋮ → Settings → Secrets:
```toml
[github]
token = "github_pat_your_token_here"
repo  = "your-username/your-repo-name"
```
        """)

    saved_ids = _sess_list(st.session_state)
    if saved_ids:
        st.markdown(f"**{len(saved_ids)} template(s) saved this session:**")
        for acc_id in saved_ids:
            ca, cb = st.columns([4, 1])
            with ca:
                st.markdown(f"**Account {acc_id}**")
            with cb:
                if _is_admin() and st.button("Delete", key=f"sess_del_{acc_id}"):
                    _sess_del(st.session_state, acc_id)
                    st.rerun()
        st.download_button(
            "💾  Download backup (restore next session)",
            data=_sess_export(st.session_state),
            file_name="ar_suite_templates_backup.json",
            mime="application/json",
            key="sess_export",
        )

    with st.expander("➕  Upload a template (session only)"):
        acc = st.text_input("Account number", key="sess_acc", placeholder="e.g. 30113601")
        f   = st.file_uploader("Template (.xlsx)", type=["xlsx", "xls"], key="sess_file")
        if f and acc:
            raw = f.read()
            try:
                info = template_preview(raw)
                st.caption(f"{info['layout_type']} · {info['max_col']} columns")
            except Exception:
                pass
            if _is_admin() and st.button("Save", key="sess_save"):
                _sess_save(st.session_state, acc, raw)
                st.success(f"Saved for account {acc} (this session).")
                st.rerun()

    with st.expander("📂  Restore from backup"):
        bf = st.file_uploader("Backup .json", type=["json"], key="sess_restore")
        if bf:
            if st.button("Restore", key="sess_restore_btn"):
                n = _sess_import(st.session_state, bf.read())
                st.success(f"Restored {n} template(s).")
                st.rerun()
